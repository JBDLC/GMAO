import datetime as dt
import os
import csv
import json
import threading
import time
from pathlib import Path
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user

from flask import Flask, redirect, render_template, request, url_for, flash, Response, make_response, abort, send_from_directory, jsonify, session
from werkzeug.utils import secure_filename
from translations import get_translation, get_language_from_session, TRANSLATIONS
from openpyxl import Workbook, load_workbook
import qrcode
from PIL import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO, StringIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import CheckConstraint, inspect, text, func
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy.exc import IntegrityError
from functools import wraps

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_FOLDER = BASE_DIR / "uploads" / "machine_documents"
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
MAINTENANCE_PHOTOS_FOLDER = BASE_DIR / "uploads" / "maintenance_photos"
MAINTENANCE_PHOTOS_FOLDER.mkdir(parents=True, exist_ok=True)
REPORT_PHOTOS_FOLDER = BASE_DIR / "uploads" / "report_photos"
REPORT_PHOTOS_FOLDER.mkdir(parents=True, exist_ok=True)
EXCEL_FILES_FOLDER = BASE_DIR / "uploads" / "excel_files"
EXCEL_FILES_FOLDER.mkdir(parents=True, exist_ok=True)
ALLOWED_EXTENSIONS = {'pdf'}
ALLOWED_IMAGE_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

app = Flask(__name__)

# Configuration de la base de données
# Sur Render, utiliser PostgreSQL via DATABASE_URL, sinon SQLite en local
database_url = os.environ.get("DATABASE_URL")
if database_url:
    # Render fournit DATABASE_URL au format postgresql://user:pass@host/dbname
    # SQLAlchemy attend postgresql:// mais certaines versions utilisent postgres://
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql+psycopg://", 1)
    elif database_url.startswith("postgresql://"):
        # Forcer l'utilisation de psycopg (v3) au lieu de psycopg2
        database_url = database_url.replace("postgresql://", "postgresql+psycopg://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
else:
    # Mode local avec SQLite
    DB_PATH = BASE_DIR / "app.db"
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-me-in-production")

db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "splash"
login_manager.login_message = None
login_manager.login_message_category = "info"

# Fonctions de permissions (doivent être définies avant le context processor)
def can_view_params():
    """Vérifie si l'utilisateur peut voir les paramètres"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type == "admin"


def can_edit_machines():
    """Vérifie si l'utilisateur peut modifier/supprimer des machines"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type == "admin"


def can_create_checklist():
    """Vérifie si l'utilisateur peut créer une check list"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type in ["admin"]


def can_create_preventive_template():
    """Vérifie si l'utilisateur peut créer un modèle de maintenance préventive"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type in ["admin"]


def can_create_corrective_maintenance():
    """Vérifie si l'utilisateur peut créer une maintenance corrective"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type in ["admin", "technicien"]


def can_delete_machines():
    """Vérifie si l'utilisateur peut supprimer des machines"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type == "admin"


def can_delete_stocks():
    """Vérifie si l'utilisateur peut supprimer des stocks"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type == "admin"


def can_delete_products():
    """Vérifie si l'utilisateur peut supprimer des produits"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type == "admin"


def can_edit_stocks_products():
    """Vérifie si l'utilisateur peut modifier des stocks/produits"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type in ["admin", "gestionnaire"]


def can_edit_machines_maintenances():
    """Vérifie si l'utilisateur peut modifier des machines/maintenances"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type in ["admin", "technicien"]


def can_add_documentation():
    """Vérifie si l'utilisateur peut ajouter de la documentation"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type in ["admin"]


def can_view_documentation():
    """Vérifie si l'utilisateur peut voir la documentation"""
    return current_user.is_authenticated


def can_access_chat():
    """Vérifie si l'utilisateur peut accéder au chat"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type in ["admin", "spectateur", "gestionnaire", "technicien"]


def can_access_qrcode():
    """Vérifie si l'utilisateur peut accéder au QR code"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type in ["admin", "spectateur", "gestionnaire", "technicien"]


def is_readonly_machines_maintenances():
    """Vérifie si l'utilisateur est en mode lecture seule sur machines/maintenances"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type == "gestionnaire"


def is_readonly_stocks_products():
    """Vérifie si l'utilisateur est en mode lecture seule sur stocks/produits"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type == "technicien"


def is_spectator():
    """Vérifie si l'utilisateur est un spectateur (lecture seule partout)"""
    if not current_user.is_authenticated:
        return False
    return current_user.user_type == "spectateur"


# Context processor pour rendre les traductions disponibles dans tous les templates
@app.context_processor
def inject_translations():
    """Injecte les fonctions de traduction dans tous les templates"""
    lang = get_language_from_session(session)
    def t(key):
        """Fonction de traduction pour les templates"""
        return get_translation(key, lang)
    
    # Injecter les fonctions de permissions dans les templates
    return dict(
        t=t, 
        current_lang=lang, 
        available_languages=['fr', 'es', 'en', 'it'],
        can_view_params=can_view_params,
        can_edit_machines=can_edit_machines,
        can_create_checklist=can_create_checklist,
        can_create_preventive_template=can_create_preventive_template,
        can_create_corrective_maintenance=can_create_corrective_maintenance,
        can_delete_machines=can_delete_machines,
        can_delete_stocks=can_delete_stocks,
        can_delete_products=can_delete_products,
        can_edit_stocks_products=can_edit_stocks_products,
        can_edit_machines_maintenances=can_edit_machines_maintenances,
        can_add_documentation=can_add_documentation,
        can_view_documentation=can_view_documentation,
        can_access_chat=can_access_chat,
        can_access_qrcode=can_access_qrcode,
        is_readonly_machines_maintenances=is_readonly_machines_maintenances,
        is_readonly_stocks_products=is_readonly_stocks_products,
        is_spectator=is_spectator
    )


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# Décorateurs pour la gestion des rôles
def admin_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.user_type != "admin":
            flash("Accès refusé : cette fonctionnalité est réservée aux administrateurs.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated_function


def admin_or_technician_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.user_type not in ["admin", "technicien"]:
            flash("Accès refusé : cette fonctionnalité est réservée aux administrateurs et techniciens.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated_function


def admin_or_manager_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.user_type not in ["admin", "gestionnaire"]:
            flash("Accès refusé : cette fonctionnalité est réservée aux administrateurs et gestionnaires.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated_function


def can_edit_maintenance_entry(entry):
    """Vérifie si l'utilisateur peut modifier/supprimer un rapport de maintenance"""
    if current_user.user_type == "admin":
        return True
    if current_user.user_type == "technicien" and entry.user_id == current_user.id:
        return True
    return False


# Les fonctions de permissions sont définies au début du fichier (après le context processor)


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    user_type = db.Column(db.String(20), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Machine(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    code = db.Column(db.String(50), unique=True, nullable=False)
    parent_id = db.Column(db.Integer, db.ForeignKey("machine.id"), index=True)
    hour_counter_enabled = db.Column(db.Boolean, default=False)
    hours = db.Column(db.Float, default=0.0)
    counter_unit = db.Column(db.String(20), nullable=True)  # Unité du compteur (h, cycles, anneaux, etc.)
    stock_id = db.Column(db.Integer, db.ForeignKey("stock.id"), nullable=True, index=True)  # Stock associé par défaut
    color_index = db.Column(db.Integer, default=0)  # Index de couleur (0-9) pour les machines racines

    parent = db.relationship("Machine", remote_side=[id], backref="children")
    stock = db.relationship("Stock")
    counters = db.relationship("Counter", back_populates="machine", cascade="all, delete-orphan")

    def depth(self):
        depth = 0
        parent = self.parent
        while parent:
            depth += 1
            parent = parent.parent
        return depth
    
    def is_root(self):
        """Vérifie si la machine est une machine racine"""
        return self.parent_id is None


class FollowedMachine(db.Model):
    """Machines suivies par les utilisateurs"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=dt.datetime.utcnow)

    user = db.relationship("User", backref="followed_machines")
    machine = db.relationship("Machine", backref="followers")

    __table_args__ = (db.UniqueConstraint('user_id', 'machine_id', name='unique_user_machine_follow'),)


class Counter(db.Model):
    """Compteur pour les machines racines (peut y en avoir plusieurs)"""
    id = db.Column(db.Integer, primary_key=True)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False, index=True)
    name = db.Column(db.String(120), nullable=False)  # Nom du compteur
    value = db.Column(db.Float, default=0.0)  # Valeur actuelle
    unit = db.Column(db.String(20), nullable=True)  # Unité (h, cycles, anneaux, etc.)

    machine = db.relationship("Machine", back_populates="counters")
    
    __table_args__ = (
        db.UniqueConstraint("machine_id", "name", name="uq_counter_machine_name"),
    )


class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    code = db.Column(db.String(60), unique=True, nullable=False)
    price = db.Column(db.Float, nullable=False)
    supplier_name = db.Column(db.String(120))
    supplier_reference = db.Column(db.String(120))
    location_code = db.Column(db.String(120))
    minimum_stock = db.Column(db.Float, default=0.0)


class Stock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    code = db.Column(db.String(60), unique=True, nullable=False)

    items = db.relationship("StockProduct", back_populates="stock", cascade="all, delete-orphan")


class StockProduct(db.Model):
    __tablename__ = "stock_product"

    id = db.Column(db.Integer, primary_key=True)
    stock_id = db.Column(db.Integer, db.ForeignKey("stock.id"), nullable=False, index=True)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=False, index=True)
    quantity = db.Column(db.Float, nullable=False, default=0.0)

    stock = db.relationship("Stock", back_populates="items")
    product = db.relationship("Product")

    __table_args__ = (
        db.UniqueConstraint("stock_id", "product_id", name="uq_stock_product"),
        CheckConstraint("quantity >= 0", name="ck_stock_product_positive_qty"),
    )


class Inventory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    stock_id = db.Column(db.Integer, db.ForeignKey("stock.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    name = db.Column(db.String(200), nullable=True)  # Nom de l'inventaire (généré automatiquement)
    created_at = db.Column(db.DateTime, nullable=False, default=dt.datetime.utcnow)
    
    stock = db.relationship("Stock", backref="inventories")
    user = db.relationship("User")
    items = db.relationship("InventoryItem", back_populates="inventory", cascade="all, delete-orphan")


class InventoryItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    inventory_id = db.Column(db.Integer, db.ForeignKey("inventory.id"), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=False)
    previous_quantity = db.Column(db.Float, nullable=False)
    new_quantity = db.Column(db.Float, nullable=False)
    comment = db.Column(db.Text, nullable=True)
    
    inventory = db.relationship("Inventory", back_populates="items")
    product = db.relationship("Product")


class Movement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.String(20), nullable=False)
    source_stock_id = db.Column(db.Integer, db.ForeignKey("stock.id"), index=True)
    dest_stock_id = db.Column(db.Integer, db.ForeignKey("stock.id"), index=True)
    created_at = db.Column(db.DateTime, nullable=False, default=dt.datetime.utcnow, index=True)

    source_stock = db.relationship("Stock", foreign_keys=[source_stock_id])
    dest_stock = db.relationship("Stock", foreign_keys=[dest_stock_id])
    items = db.relationship("MovementItem", back_populates="movement", cascade="all, delete-orphan")


class MovementItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    movement_id = db.Column(db.Integer, db.ForeignKey("movement.id"), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)

    movement = db.relationship("Movement", back_populates="items")
    product = db.relationship("Product")


class PreventiveReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False, index=True)
    counter_id = db.Column(db.Integer, db.ForeignKey("counter.id"), nullable=True, index=True)  # Pour les machines racines avec plusieurs compteurs
    periodicity = db.Column(db.Integer, nullable=False)

    machine = db.relationship("Machine", backref="preventive_reports")
    counter = db.relationship("Counter", backref="preventive_reports")
    components = db.relationship(
        "PreventiveComponent",
        back_populates="report",
        cascade="all, delete-orphan",
        order_by="PreventiveComponent.id",
    )


class PreventiveComponent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("preventive_report.id"), nullable=False)
    label = db.Column(db.String(150), nullable=False)
    field_type = db.Column(db.String(20), nullable=False)

    report = db.relationship("PreventiveReport", back_populates="components")


class MaintenanceEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False, index=True)
    report_id = db.Column(db.Integer, db.ForeignKey("preventive_report.id"), nullable=False, index=True)
    stock_id = db.Column(db.Integer, db.ForeignKey("stock.id"), index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), index=True)
    performed_hours = db.Column(db.Float, default=0.0)
    hours_before_maintenance = db.Column(db.Float)
    created_at = db.Column(db.DateTime, nullable=False, default=dt.datetime.utcnow, index=True)

    machine = db.relationship("Machine", backref="maintenance_entries")
    report = db.relationship("PreventiveReport")
    stock = db.relationship("Stock")
    user = db.relationship("User")
    values = db.relationship(
        "MaintenanceEntryValue",
        back_populates="entry",
        cascade="all, delete-orphan",
        order_by="MaintenanceEntryValue.id",
    )


class MaintenanceEntryValue(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    entry_id = db.Column(db.Integer, db.ForeignKey("maintenance_entry.id"), nullable=False)
    component_id = db.Column(db.Integer, db.ForeignKey("preventive_component.id"), nullable=False)
    value_text = db.Column(db.Text)
    value_number = db.Column(db.Float)
    value_bool = db.Column(db.Boolean)

    entry = db.relationship("MaintenanceEntry", back_populates="values")
    component = db.relationship("PreventiveComponent")


class MaintenanceProgress(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False, index=True)
    counter_id = db.Column(db.Integer, db.ForeignKey("counter.id"), nullable=True, index=True)  # Pour les machines racines avec compteurs multiples
    report_id = db.Column(db.Integer, db.ForeignKey("preventive_report.id"), nullable=False, index=True)
    hours_since = db.Column(db.Float, nullable=False, default=0.0)

    machine = db.relationship("Machine", backref="maintenance_progress")
    counter = db.relationship("Counter", backref="maintenance_progress")
    report = db.relationship("PreventiveReport")

    __table_args__ = (
        db.UniqueConstraint("machine_id", "report_id", name="uq_progress_machine_report"),
        db.UniqueConstraint("counter_id", "report_id", name="uq_progress_counter_report"),
    )


class CorrectiveMaintenance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False, index=True)
    stock_id = db.Column(db.Integer, db.ForeignKey("stock.id"), index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), index=True)
    comment = db.Column(db.Text)
    hours = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, nullable=False, default=dt.datetime.utcnow, index=True)

    machine = db.relationship("Machine", backref="corrective_maintenances")
    stock = db.relationship("Stock")
    user = db.relationship("User")
    products = db.relationship(
        "CorrectiveMaintenanceProduct",
        back_populates="maintenance",
        cascade="all, delete-orphan",
        order_by="CorrectiveMaintenanceProduct.id",
    )


class CorrectiveMaintenanceProduct(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    maintenance_id = db.Column(db.Integer, db.ForeignKey("corrective_maintenance.id"), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("product.id"), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)

    maintenance = db.relationship("CorrectiveMaintenance", back_populates="products")
    product = db.relationship("Product")


class MachineDocument(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    uploaded_at = db.Column(db.DateTime, nullable=False, default=dt.datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"))

    machine = db.relationship("Machine", backref="documents")
    user = db.relationship("User")


class MaintenancePhoto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    # Lié soit à une maintenance préventive, soit à une maintenance corrective
    maintenance_entry_id = db.Column(db.Integer, db.ForeignKey("maintenance_entry.id"), nullable=True)
    corrective_maintenance_id = db.Column(db.Integer, db.ForeignKey("corrective_maintenance.id"), nullable=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    uploaded_at = db.Column(db.DateTime, nullable=False, default=dt.datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"))

    maintenance_entry = db.relationship("MaintenanceEntry", backref="photos")
    corrective_maintenance = db.relationship("CorrectiveMaintenance", backref="photos")
    user = db.relationship("User")


class CounterLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False, index=True)
    counter_id = db.Column(db.Integer, db.ForeignKey("counter.id"), nullable=True, index=True)  # None pour compteur machine, ID pour compteur multiple
    previous_hours = db.Column(db.Float, nullable=False)
    new_hours = db.Column(db.Float, nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=dt.datetime.utcnow, index=True)

    machine = db.relationship("Machine", backref="counter_logs")
    counter = db.relationship("Counter", backref="counter_logs")


class ChecklistTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False, index=True)
    name = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=dt.datetime.utcnow)

    machine = db.relationship("Machine", backref="checklist_templates")
    items = db.relationship(
        "ChecklistItem",
        back_populates="template",
        cascade="all, delete-orphan",
        order_by="ChecklistItem.id",
    )
    instances = db.relationship(
        "ChecklistInstance",
        back_populates="template",
        cascade="all, delete-orphan",
        order_by="ChecklistInstance.created_at.desc()",
    )


class ChecklistItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    template_id = db.Column(db.Integer, db.ForeignKey("checklist_template.id"), nullable=False)
    label = db.Column(db.String(300), nullable=False)
    order = db.Column(db.Integer, default=0)

    template = db.relationship("ChecklistTemplate", back_populates="items")


class ChecklistInstance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    template_id = db.Column(db.Integer, db.ForeignKey("checklist_template.id"), nullable=False, index=True)
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), index=True)
    created_at = db.Column(db.DateTime, nullable=False, default=dt.datetime.utcnow, index=True)
    comment = db.Column(db.Text)

    template = db.relationship("ChecklistTemplate", back_populates="instances")
    machine = db.relationship("Machine", backref="checklist_instances")
    user = db.relationship("User")
    values = db.relationship(
        "ChecklistInstanceValue",
        back_populates="instance",
        cascade="all, delete-orphan",
    )


class ChecklistInstanceValue(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    instance_id = db.Column(db.Integer, db.ForeignKey("checklist_instance.id"), nullable=False)
    item_id = db.Column(db.Integer, db.ForeignKey("checklist_item.id"), nullable=False)
    checked = db.Column(db.Boolean, default=False)

    instance = db.relationship("ChecklistInstance", back_populates="values")
    item = db.relationship("ChecklistItem")


class ChatMessage(db.Model):
    """Messages du chat (manuels et automatiques)"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True, index=True)  # None pour messages auto
    message_type = db.Column(db.String(20), nullable=False, index=True)  # 'manual' ou 'auto'
    content = db.Column(db.Text, nullable=False)
    link_url = db.Column(db.String(500), nullable=True)  # Lien vers la tâche réalisée
    machine_id = db.Column(db.Integer, db.ForeignKey("machine.id"), nullable=True, index=True)
    reply_to_id = db.Column(db.Integer, db.ForeignKey("chat_message.id"), nullable=True, index=True)  # Message auquel on répond
    edited_at = db.Column(db.DateTime, nullable=True)  # Date de dernière modification
    deleted_at = db.Column(db.DateTime, nullable=True, index=True)  # Date de suppression (soft delete)
    created_at = db.Column(db.DateTime, default=dt.datetime.utcnow, nullable=False, index=True)

    user = db.relationship("User", backref="chat_messages")
    machine = db.relationship("Machine")
    reply_to = db.relationship("ChatMessage", remote_side=[id], backref="replies")


class ChatReadStatus(db.Model):
    """Suivi des messages lus par chaque utilisateur"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), unique=True, nullable=False)
    last_read_at = db.Column(db.DateTime, default=dt.datetime.utcnow, nullable=False)

    user = db.relationship("User", backref="chat_read_status")


class Report(db.Model):
    """Rapports de poste/jour créés par les utilisateurs"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=dt.datetime.utcnow, nullable=False, index=True)
    edited_at = db.Column(db.DateTime, nullable=True)  # Date de dernière modification
    deleted_at = db.Column(db.DateTime, nullable=True, index=True)  # Date de suppression (soft delete)

    user = db.relationship("User", backref="reports")
    photos = db.relationship(
        "ReportPhoto",
        back_populates="report",
        cascade="all, delete-orphan",
        order_by="ReportPhoto.id"
    )


class ReportPhoto(db.Model):
    """Photos associées aux rapports"""
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("report.id"), nullable=False)
    file_path = db.Column(db.String(500), nullable=True)  # Gardé pour compatibilité, mais optionnel maintenant
    original_filename = db.Column(db.String(255), nullable=False)
    photo_data = db.Column(db.LargeBinary, nullable=True)  # Stockage BLOB des données de l'image
    content_type = db.Column(db.String(50), nullable=True)  # Type MIME (image/jpeg, image/png, etc.)

    report = db.relationship("Report", back_populates="photos")


class ExcelFile(db.Model):
    """Fichiers Excel uploadés par les administrateurs"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)  # Nom personnalisé donné par l'utilisateur
    filename = db.Column(db.String(255), nullable=False)  # Nom du fichier sur le disque
    original_filename = db.Column(db.String(255), nullable=False)  # Nom original du fichier uploadé
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=dt.datetime.utcnow, nullable=False)
    
    user = db.relationship("User", backref="uploaded_excel_files")

with app.app_context():
    db.create_all()
    # Ajouter les index manquants pour optimiser les performances
    try:
        inspector = inspect(db.engine)
        # Créer les index si nécessaire (pour les bases existantes)
        # PostgreSQL supporte "CREATE INDEX IF NOT EXISTS" depuis la version 9.5
        # Pour SQLite, cette syntaxe est supportée depuis la version 3.9.0
        with db.engine.connect() as conn:
            # Index pour Machine
            try:
                # Vérifier si l'index existe déjà (compatible avec PostgreSQL et SQLite)
                indexes = [idx['name'] for idx in inspector.get_indexes('machine')]
                if 'ix_machine_parent_id' not in indexes:
                    conn.execute(text("CREATE INDEX ix_machine_parent_id ON machine(parent_id)"))
                if 'ix_machine_stock_id' not in indexes:
                    conn.execute(text("CREATE INDEX ix_machine_stock_id ON machine(stock_id)"))
                conn.commit()
            except Exception:
                # Fallback: utiliser CREATE INDEX IF NOT EXISTS si la vérification échoue
                try:
                    conn.execute(text("CREATE INDEX IF NOT EXISTS ix_machine_parent_id ON machine(parent_id)"))
                    conn.execute(text("CREATE INDEX IF NOT EXISTS ix_machine_stock_id ON machine(stock_id)"))
                    conn.commit()
                except Exception:
                    pass
            
            # Index pour FollowedMachine
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_followed_machine_user_id ON followed_machine(user_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_followed_machine_machine_id ON followed_machine(machine_id)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour Counter
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_counter_machine_id ON counter(machine_id)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour Movement
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_movement_source_stock_id ON movement(source_stock_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_movement_dest_stock_id ON movement(dest_stock_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_movement_created_at ON movement(created_at)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour MaintenanceEntry
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_maintenance_entry_machine_id ON maintenance_entry(machine_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_maintenance_entry_report_id ON maintenance_entry(report_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_maintenance_entry_stock_id ON maintenance_entry(stock_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_maintenance_entry_user_id ON maintenance_entry(user_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_maintenance_entry_created_at ON maintenance_entry(created_at)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour MaintenanceProgress
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_maintenance_progress_machine_id ON maintenance_progress(machine_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_maintenance_progress_counter_id ON maintenance_progress(counter_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_maintenance_progress_report_id ON maintenance_progress(report_id)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour CorrectiveMaintenance
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_corrective_maintenance_machine_id ON corrective_maintenance(machine_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_corrective_maintenance_stock_id ON corrective_maintenance(stock_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_corrective_maintenance_user_id ON corrective_maintenance(user_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_corrective_maintenance_created_at ON corrective_maintenance(created_at)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour CounterLog
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_counter_log_machine_id ON counter_log(machine_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_counter_log_counter_id ON counter_log(counter_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_counter_log_created_at ON counter_log(created_at)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour ChecklistTemplate
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_checklist_template_machine_id ON checklist_template(machine_id)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour ChecklistInstance
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_checklist_instance_template_id ON checklist_instance(template_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_checklist_instance_machine_id ON checklist_instance(machine_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_checklist_instance_user_id ON checklist_instance(user_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_checklist_instance_created_at ON checklist_instance(created_at)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour PreventiveReport
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_preventive_report_machine_id ON preventive_report(machine_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_preventive_report_counter_id ON preventive_report(counter_id)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour ChatMessage
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_chat_message_user_id ON chat_message(user_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_chat_message_message_type ON chat_message(message_type)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_chat_message_machine_id ON chat_message(machine_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_chat_message_reply_to_id ON chat_message(reply_to_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_chat_message_deleted_at ON chat_message(deleted_at)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_chat_message_created_at ON chat_message(created_at)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour Report
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_report_user_id ON report(user_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_report_created_at ON report(created_at)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_report_deleted_at ON report(deleted_at)"))
                conn.commit()
            except Exception:
                pass
            
            # Index pour StockProduct
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_stock_product_stock_id ON stock_product(stock_id)"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_stock_product_product_id ON stock_product(product_id)"))
                conn.commit()
            except Exception:
                pass
    except Exception:
        pass
    
    try:
        inspector = inspect(db.engine)
        columns = {col["name"] for col in inspector.get_columns("maintenance_entry")}
        if "stock_id" not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE maintenance_entry ADD COLUMN stock_id INTEGER"))
        if "performed_hours" not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE maintenance_entry ADD COLUMN performed_hours FLOAT DEFAULT 0"))
        if "user_id" not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE maintenance_entry ADD COLUMN user_id INTEGER"))
        if "hours_before_maintenance" not in columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE maintenance_entry ADD COLUMN hours_before_maintenance FLOAT"))
    except Exception:
        pass
    try:
        inspector = inspect(db.engine)
        corrective_columns = {col["name"] for col in inspector.get_columns("corrective_maintenance")}
        if "user_id" not in corrective_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE corrective_maintenance ADD COLUMN user_id INTEGER"))
        if "hours" not in corrective_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE corrective_maintenance ADD COLUMN hours FLOAT DEFAULT 0"))
    except Exception:
        pass
    # Migration pour la table chat_message
    try:
        inspector = inspect(db.engine)
        if "chat_message" in inspector.get_table_names():
            chat_message_columns = {col["name"] for col in inspector.get_columns("chat_message")}
            if "reply_to_id" not in chat_message_columns:
                with db.engine.connect() as conn:
                    conn.execute(text("ALTER TABLE chat_message ADD COLUMN reply_to_id INTEGER"))
                    conn.commit()
            if "edited_at" not in chat_message_columns:
                with db.engine.connect() as conn:
                    # Utiliser TIMESTAMP pour PostgreSQL, DATETIME pour SQLite
                    if database_url and "postgresql" in database_url:
                        conn.execute(text("ALTER TABLE chat_message ADD COLUMN edited_at TIMESTAMP"))
                    else:
                        conn.execute(text("ALTER TABLE chat_message ADD COLUMN edited_at DATETIME"))
                    conn.commit()
            if "deleted_at" not in chat_message_columns:
                with db.engine.connect() as conn:
                    # Utiliser TIMESTAMP pour PostgreSQL, DATETIME pour SQLite
                    if database_url and "postgresql" in database_url:
                        conn.execute(text("ALTER TABLE chat_message ADD COLUMN deleted_at TIMESTAMP"))
                    else:
                        conn.execute(text("ALTER TABLE chat_message ADD COLUMN deleted_at DATETIME"))
                    conn.commit()
    except Exception as exc:
        print(f"Error migrating chat_message table: {exc}")
        pass
    # Migration pour la table maintenance_photo
    try:
        inspector = inspect(db.engine)
        if "maintenance_photo" not in inspector.get_table_names():
            # La table sera créée automatiquement par db.create_all()
            pass
    except Exception:
        pass
    try:
        inspector = inspect(db.engine)
        machine_columns = {col["name"] for col in inspector.get_columns("machine")}
        if "counter_unit" not in machine_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE machine ADD COLUMN counter_unit VARCHAR(20)"))
                # Pour les machines existantes avec compteur horaire, définir l'unité par défaut à "h"
                # Note: PostgreSQL utilise TRUE/FALSE au lieu de 1/0 pour les booléens
                if database_url and "postgresql" in database_url:
                    conn.execute(text("UPDATE machine SET counter_unit = 'h' WHERE hour_counter_enabled = TRUE AND counter_unit IS NULL"))
                else:
                    conn.execute(text("UPDATE machine SET counter_unit = 'h' WHERE hour_counter_enabled = 1 AND counter_unit IS NULL"))
                conn.commit()
        if "stock_id" not in machine_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE machine ADD COLUMN stock_id INTEGER"))
                conn.commit()
    except Exception:
        pass
    # Migration pour la table counter
    try:
        inspector = inspect(db.engine)
        if "counter" not in [tbl["name"] for tbl in inspector.get_table_names()]:
            db.create_all()
    except Exception:
        pass
    # Migration pour ajouter counter_id à preventive_report
    try:
        inspector = inspect(db.engine)
        preventive_report_columns = {col["name"] for col in inspector.get_columns("preventive_report")}
        if "counter_id" not in preventive_report_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE preventive_report ADD COLUMN counter_id INTEGER"))
                conn.commit()
    except Exception:
        pass
    # Migration pour ajouter counter_id à maintenance_progress
    try:
        inspector = inspect(db.engine)
        maintenance_progress_columns = {col["name"] for col in inspector.get_columns("maintenance_progress")}
        if "counter_id" not in maintenance_progress_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE maintenance_progress ADD COLUMN counter_id INTEGER"))
                conn.commit()
    except Exception:
        pass
    try:
        inspector = inspect(db.engine)
        product_columns = {col["name"] for col in inspector.get_columns("product")}
        if "supplier_name" not in product_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE product ADD COLUMN supplier_name VARCHAR(120)"))
        if "supplier_reference" not in product_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE product ADD COLUMN supplier_reference VARCHAR(120)"))
        if "location_code" not in product_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE product ADD COLUMN location_code VARCHAR(120)"))
        if "minimum_stock" not in product_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE product ADD COLUMN minimum_stock FLOAT DEFAULT 0"))
    except Exception:
        pass
    # Migration pour ajouter counter_id à counter_log
    try:
        inspector = inspect(db.engine)
        counter_log_columns = {col["name"] for col in inspector.get_columns("counter_log")}
        if "counter_id" not in counter_log_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE counter_log ADD COLUMN counter_id INTEGER"))
                conn.commit()
    except Exception:
        pass
    # Migration pour créer les tables Report et ReportPhoto
    try:
        inspector = inspect(db.engine)
        tables = {table["name"] for table in inspector.get_table_names()}
        if "report" not in tables:
            db.create_all()
    except Exception:
        pass
    
    # Migration pour ajouter photo_data et content_type à report_photo
    try:
        inspector = inspect(db.engine)
        if "report_photo" in inspector.get_table_names():
            report_photo_columns = {col["name"] for col in inspector.get_columns("report_photo")}
            
            # Ajouter photo_data si elle n'existe pas
            if "photo_data" not in report_photo_columns:
                with db.engine.connect() as conn:
                    # Pour PostgreSQL, utiliser BYTEA pour les BLOB
                    # Pour SQLite, utiliser BLOB
                    if db.engine.dialect.name == 'postgresql':
                        conn.execute(text("ALTER TABLE report_photo ADD COLUMN photo_data BYTEA"))
                    else:
                        conn.execute(text("ALTER TABLE report_photo ADD COLUMN photo_data BLOB"))
                    conn.commit()
                    print("Colonne photo_data ajoutée à report_photo")
            
            # Ajouter content_type si elle n'existe pas
            if "content_type" not in report_photo_columns:
                with db.engine.connect() as conn:
                    conn.execute(text("ALTER TABLE report_photo ADD COLUMN content_type VARCHAR(50)"))
                    conn.commit()
                    print("Colonne content_type ajoutée à report_photo")
            
            # Rendre file_path nullable si ce n'est pas déjà le cas
            # (Cette partie peut échouer si la colonne est déjà nullable, c'est normal)
            try:
                with db.engine.connect() as conn:
                    if db.engine.dialect.name == 'postgresql':
                        conn.execute(text("ALTER TABLE report_photo ALTER COLUMN file_path DROP NOT NULL"))
                    else:
                        # SQLite ne supporte pas ALTER COLUMN, on ignore
                        pass
                    conn.commit()
            except Exception:
                pass
    except Exception as exc:
        print(f"Erreur lors de la migration report_photo: {exc}")
        pass
    updated = PreventiveComponent.query.filter_by(field_type="increment").update({"field_type": "number"})
    if updated:
        db.session.commit()
    try:
        # Migration: rendre machine_id obligatoire pour PreventiveReport
        inspector = inspect(db.engine)
        preventive_report_columns = {col["name"]: col for col in inspector.get_columns("preventive_report")}
        if "machine_id" in preventive_report_columns:
            # Supprimer les rapports sans machine_id (s'ils existent)
            reports_without_machine = PreventiveReport.query.filter_by(machine_id=None).all()
            if reports_without_machine:
                for report in reports_without_machine:
                    db.session.delete(report)
                db.session.commit()
            # Vérifier si la colonne est nullable et la rendre NOT NULL
            # Note: SQLite ne supporte pas ALTER COLUMN, donc on ne peut pas changer directement
            # Mais comme on a supprimé les NULL, le modèle SQLAlchemy avec nullable=False devrait fonctionner
    except Exception:
        db.session.rollback()
    try:
        machines_with_counter = Machine.query.filter_by(hour_counter_enabled=True).all()
        created = False
        for machine in machines_with_counter:
            reports = PreventiveReport.query.filter_by(machine_id=machine.id).all()
            existing_ids = {
                record.report_id for record in MaintenanceProgress.query.filter_by(machine_id=machine.id).all()
            }
            for report in reports:
                if report.id not in existing_ids:
                    initial_hours = report.periodicity
                    db.session.add(MaintenanceProgress(machine=machine, report=report, hours_since=initial_hours))
                    created = True
        if created:
            db.session.commit()
    except Exception:
        db.session.rollback()
    # Créer le compte admin par défaut s'il n'existe pas
    try:
        admin_user = User.query.filter_by(username="admin123").first()
        if not admin_user:
            admin_user = User(username="admin123", user_type="admin")
            admin_user.set_password("123")
            db.session.add(admin_user)
            db.session.commit()
    except Exception:
        db.session.rollback()


@app.route("/")
@login_required
def index():
    # Récupérer les données de maintenances avec seuil par défaut de 10%
    threshold_ratio = 0.10
    overdue = []
    warning = []

    # Charger progress_records une seule fois avec eager loading
    progress_records = (
        MaintenanceProgress.query
        .options(joinedload(MaintenanceProgress.machine), joinedload(MaintenanceProgress.report))
        .join(MaintenanceProgress.machine)
        .join(MaintenanceProgress.report)
        .all()
    )

    last_entry_rows = (
        db.session.query(
            MaintenanceEntry.machine_id.label("machine_id"),
            MaintenanceEntry.report_id.label("report_id"),
            func.max(MaintenanceEntry.created_at).label("last_date"),
        )
        .group_by(MaintenanceEntry.machine_id, MaintenanceEntry.report_id)
        .all()
    )
    last_map = {(row.machine_id, row.report_id): row.last_date for row in last_entry_rows}

    for record in progress_records:
        machine = record.machine
        report = record.report
        if not machine or not machine.hour_counter_enabled:
            continue
        remaining = record.hours_since
        last_performed = last_map.get((machine.id, report.id))
        if remaining <= 0:
            overdue.append(
                {
                    "machine": machine,
                    "report": report,
                    "remaining": remaining,
                    "last_performed": last_performed,
                }
            )
        elif remaining <= report.periodicity * threshold_ratio:
            warning.append(
                {
                    "machine": machine,
                    "report": report,
                    "remaining": remaining,
                    "last_performed": last_performed,
                }
            )

    # Optimiser le calcul du stock minimum : une seule requête avec jointure
    first_stock = Stock.query.order_by(Stock.id).first()
    low_stock_count = 0
    if first_stock:
        # Utiliser une requête SQL optimisée au lieu d'une boucle
        low_stock_products = (
            db.session.query(Product.id)
            .outerjoin(StockProduct, db.and_(
                StockProduct.product_id == Product.id,
                StockProduct.stock_id == first_stock.id
            ))
            .filter(
                Product.minimum_stock > 0,
                db.or_(
                    StockProduct.quantity.is_(None),
                    StockProduct.quantity < Product.minimum_stock
                )
            )
            .count()
        )
        low_stock_count = low_stock_products

    # Statistiques supplémentaires
    # Nombre total de maintenances (préventives + correctives)
    total_maintenances = MaintenanceEntry.query.count() + CorrectiveMaintenance.query.count()
    
    # Nombre de relevés du jour
    today_start = dt.datetime.combine(dt.date.today(), dt.time.min)
    today_end = dt.datetime.combine(dt.date.today(), dt.time.max)
    counter_logs_today = CounterLog.query.filter(
        CounterLog.created_at >= today_start,
        CounterLog.created_at <= today_end
    ).count()
    
    # Nombre de mouvements du jour
    movements_today = Movement.query.filter(
        Movement.created_at >= today_start,
        Movement.created_at <= today_end
    ).count()
    
    # Nombre de mouvements de la semaine (7 derniers jours)
    week_start = dt.datetime.now() - dt.timedelta(days=7)
    movements_week = Movement.query.filter(
        Movement.created_at >= week_start
    ).count()

    # Récupérer les machines suivies par l'utilisateur
    followed_machines_data = []
    followed_machines = FollowedMachine.query.filter_by(user_id=current_user.id).all()
    
    # Calculer l'état de maintenance pour chaque machine (réutiliser progress_records déjà chargé)
    machine_status = {}
    
    for record in progress_records:
        machine = record.machine
        report = record.report
        if not machine:
            continue
        
        if machine.hour_counter_enabled:
            if machine.hours <= 0:
                continue
            remaining = record.hours_since
            if remaining <= 0:
                machine_status[machine.id] = 'danger'
            elif remaining <= report.periodicity * threshold_ratio:
                if machine.id not in machine_status:
                    machine_status[machine.id] = 'warning'
        elif machine.is_root() and machine.counters:
            for counter in machine.counters:
                if counter.value <= 0:
                    continue
                remaining = record.hours_since
                if remaining <= 0:
                    machine_status[machine.id] = 'danger'
                elif remaining <= report.periodicity * threshold_ratio:
                    if machine.id not in machine_status:
                        machine_status[machine.id] = 'warning'
    
    # Récupérer uniquement les machines suivies directement (pas via un parent)
    # Pour éviter les doublons dans l'affichage
    directly_followed_ids = {
        fm.machine_id for fm in followed_machines
    }
    
    # Fonction pour vérifier si une machine est suivie via un parent
    def is_followed_via_parent(machine):
        """Vérifie si une machine est suivie via un de ses parents"""
        current = machine.parent
        while current:
            if current.id in directly_followed_ids:
                return True
            current = current.parent
        return False
    
    # Récupérer TOUTES les machines racines UNE SEULE FOIS avec eager loading des compteurs
    all_roots = (
        Machine.query
        .filter_by(parent_id=None)
        .options(joinedload(Machine.counters), joinedload(Machine.parent))
        .order_by(Machine.name)
        .all()
    )
    
    # Créer un mapping de couleur basé sur le color_index
    root_color_map = {root.id: (root.color_index if root.color_index is not None else 0) for root in all_roots}
    
    # Charger toutes les machines suivies en une seule requête pour éviter N+1
    followed_machine_ids_list = [fm.machine_id for fm in followed_machines]
    followed_machines_dict = {}
    if followed_machine_ids_list:
        followed_machines_loaded = (
            Machine.query
            .filter(Machine.id.in_(followed_machine_ids_list))
            .options(joinedload(Machine.parent))
            .all()
        )
        followed_machines_dict = {m.id: m for m in followed_machines_loaded}
    
    # Identifier quelles machines racines ont des machines suivies dans leur arborescence
    roots_with_followed = set()
    for followed in followed_machines:
        machine = followed_machines_dict.get(followed.machine_id)
        if machine and not is_followed_via_parent(machine):
            # Trouver la machine racine de cette arborescence
            root_machine = machine
            while root_machine.parent:
                root_machine = root_machine.parent
            roots_with_followed.add(root_machine.id)
    
    # Créer les données pour toutes les machines racines
    for root_machine in all_roots:
        # Précharger les compteurs si c'est une machine racine
        if root_machine.is_root() and root_machine.counters:
            # Les compteurs sont déjà chargés via la relation
            pass
        
        color_index = root_color_map.get(root_machine.id, 0)  # Utiliser le mapping basé sur l'ordre alphabétique
        followed_machines_data.append({
            'root_machine': root_machine,
            'has_followed': root_machine.id in roots_with_followed,
            'color_index': color_index
        })
    
    # Fonction helper pour vérifier si une machine est suivie
    def is_machine_followed(machine):
        """Vérifie si une machine est suivie directement ou via un de ses parents"""
        current = machine
        while current:
            if current.id in directly_followed_ids:
                return True
            current = current.parent
        return False
    
    # Récupérer les machines suivies pour l'affichage des étoiles (réutiliser all_roots)
    followed_machine_ids = set()
    for root in all_roots:
        for node, level in build_machine_tree(root):
            if is_machine_followed(node):
                followed_machine_ids.add(node.id)

    # Calculer le nombre de maintenances en retard pour les machines suivies
    followed_overdue_count = 0
    if followed_machine_ids:
        # Charger toutes les machines suivies en une seule requête
        all_followed_machines_loaded = (
            Machine.query
            .filter(Machine.id.in_(list(followed_machine_ids)))
            .all()
        )
        
        # Récupérer toutes les machines suivies (y compris les enfants)
        all_followed_machines = []
        for machine in all_followed_machines_loaded:
            all_followed_machines.extend(get_all_descendants(machine))
        
        # Dédupliquer les machines
        unique_followed_machines = {m.id: m for m in all_followed_machines}
        
        # Compter les maintenances en retard pour ces machines
        for record in progress_records:
            machine = record.machine
            if machine and machine.id in unique_followed_machines:
                if not machine.hour_counter_enabled:
                    continue
                remaining = record.hours_since
                if remaining <= 0:
                    followed_overdue_count += 1
    
    # Vérifier si on doit afficher toutes les machines
    show_all = request.args.get('show_all', 'false').lower() == 'true'
    
    return render_template(
        "index.html",
        overdue_count=len(overdue),
        warning_count=len(warning),
        low_stock_count=low_stock_count,
        total_maintenances=total_maintenances,
        counter_logs_today=counter_logs_today,
        movements_today=movements_today,
        movements_week=movements_week,
        followed_machines_data=followed_machines_data,
        machine_status=machine_status,
        followed_machine_ids=followed_machine_ids,
        show_all=show_all,
        followed_overdue_count=followed_overdue_count,
    )


@app.route("/splash")
def splash():
    """Page de splash screen avec logo animé avant la connexion"""
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    return render_template("splash.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        lang = get_language_from_session(session)
        if not username or not password:
            flash(get_translation("Identifiant et mot de passe requis", lang), "danger")
            return redirect(request.url)
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            next_page = request.args.get("next")
            return redirect(next_page or url_for("index"))
        else:
            flash(get_translation("Identifiant ou mot de passe incorrect", lang), "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    lang = get_language_from_session(session)
    flash(get_translation("Vous avez été déconnecté", lang), "info")
    return redirect(url_for("login"))

@app.route("/set-language/<lang>")
def set_language(lang):
    """Route pour changer la langue de l'interface (accessible même sans être connecté)"""
    if lang in ['fr', 'es', 'en', 'it']:
        session['language'] = lang
    # Rediriger vers la page précédente ou l'accueil
    return redirect(request.referrer or url_for('index'))


@app.route("/users")
@admin_required
def users():
    all_users = User.query.order_by(User.username).all()
    return render_template("users.html", users=all_users)


@app.route("/permissions-summary")
@admin_required
def permissions_summary():
    """Page de récapitulatif des droits d'accès"""
    return render_template("permissions_summary.html")


@app.route("/users/new", methods=["GET", "POST"])
@admin_required
def new_user():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user_type = request.form.get("user_type", "").strip()
        if not username or not password or not user_type:
            flash("Tous les champs sont requis", "danger")
            return redirect(request.url)
        if user_type not in {"admin", "gestionnaire", "spectateur", "technicien"}:
            flash("Type d'utilisateur invalide", "danger")
            return redirect(request.url)
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash("Cet identifiant existe déjà", "danger")
            return redirect(request.url)
        user = User(username=username, user_type=user_type)
        user.set_password(password)
        db.session.add(user)
        try:
            db.session.commit()
            flash("Utilisateur créé avec succès", "success")
            return redirect(url_for("users"))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)
    return render_template("user_form.html")


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def delete_user(user_id):
    if user_id == current_user.id:
        flash("Vous ne pouvez pas supprimer votre propre compte", "danger")
        return redirect(url_for("users"))
    user = User.query.get_or_404(user_id)
    if user.username == "admin123":
        flash("Le compte admin par défaut ne peut pas être supprimé", "danger")
        return redirect(url_for("users"))
    db.session.delete(user)
    try:
        db.session.commit()
        flash("Utilisateur supprimé", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur: {exc}", "danger")
    return redirect(url_for("users"))


@app.route("/machines")
@login_required
def machines():
    # Charger les machines racines avec eager loading des compteurs
    roots = (
        Machine.query
        .filter_by(parent_id=None)
        .options(joinedload(Machine.counters))
        .order_by(Machine.name)
        .all()
    )
    
    # Calculer l'état de maintenance pour chaque machine
    threshold_ratio = 0.10
    machine_status = {}  # machine_id -> 'danger' (dépassé), 'warning' (proche), ou None
    
    # Charger progress_records avec eager loading
    progress_records = (
        MaintenanceProgress.query
        .options(joinedload(MaintenanceProgress.machine), joinedload(MaintenanceProgress.report))
        .join(MaintenanceProgress.machine)
        .join(MaintenanceProgress.report)
        .all()
    )
    
    for record in progress_records:
        machine = record.machine
        report = record.report
        if not machine:
            continue
        
        # Gérer les machines avec compteur classique
        if machine.hour_counter_enabled:
            # Ne pas afficher d'alerte si la machine n'a pas encore d'heures
            if machine.hours <= 0:
                continue
            
            remaining = record.hours_since
            
            # Si hours_since est nulle ou négative, afficher le point rouge
            if remaining <= 0:
                # Périodicité dépassée ou atteinte - priorité rouge
                machine_status[machine.id] = 'danger'
            elif remaining <= report.periodicity * threshold_ratio:
                # Proche de l'échéance - orange seulement si pas déjà rouge
                if machine.id not in machine_status:
                    machine_status[machine.id] = 'warning'
        # Gérer les machines racines avec compteurs multiples
        elif machine.is_root() and record.counter_id:
            counter = Counter.query.get(record.counter_id)
            if counter and counter.value > 0:
                remaining = record.hours_since
                if remaining <= 0:
                    machine_status[machine.id] = 'danger'
                elif remaining <= report.periodicity * threshold_ratio:
                    if machine.id not in machine_status:
                        machine_status[machine.id] = 'warning'
    
    # Récupérer les machines suivies par l'utilisateur (directement)
    directly_followed_ids = {
        fm.machine_id for fm in FollowedMachine.query.filter_by(user_id=current_user.id).all()
    }
    
    # Fonction pour vérifier si une machine est suivie (directement ou via un parent)
    def is_machine_followed(machine):
        """Vérifie si une machine est suivie directement ou via un de ses parents"""
        current = machine
        while current:
            if current.id in directly_followed_ids:
                return True
            current = current.parent
        return False
    
    # Créer un set avec toutes les machines suivies (directement ou indirectement)
    followed_machine_ids = set()
    for root in roots:
        for node, level in build_machine_tree(root):
            if is_machine_followed(node):
                followed_machine_ids.add(node.id)
    
    # Créer un dictionnaire pour mapper chaque machine racine à son color_index
    # Utiliser le color_index stocké dans la base de données
    all_roots_ordered = Machine.query.filter_by(parent_id=None).order_by(Machine.name).all()
    machine_color_map = {}
    for root in all_roots_ordered:
        machine_color_map[root.id] = root.color_index if root.color_index is not None else 0
    
    # Debug: s'assurer que toutes les machines racines ont une couleur
    print(f"DEBUG machines(): {len(all_roots_ordered)} machines racines, {len(machine_color_map)} dans le mapping")
    for root in roots:
        if root.id not in machine_color_map:
            print(f"ERREUR: Machine racine {root.id} ({root.name}) n'est pas dans machine_color_map!")
            machine_color_map[root.id] = 0  # Fallback
    
    return render_template("machines.html", roots=roots, machine_status=machine_status, followed_machine_ids=followed_machine_ids, machine_color_map=machine_color_map)


@app.route("/machines/<int:machine_id>/toggle-follow", methods=["POST"])
@login_required
def toggle_follow_machine(machine_id):
    """Toggle le suivi d'une machine par l'utilisateur connecté et tous ses descendants"""
    machine = Machine.query.get_or_404(machine_id)
    
    # Vérifier si la machine est déjà suivie
    followed = FollowedMachine.query.filter_by(
        user_id=current_user.id,
        machine_id=machine_id
    ).first()
    
    if followed:
        # Retirer le suivi de cette machine et tous ses descendants
        descendants = get_all_descendants(machine)
        machine_ids_to_remove = [m.id for m in descendants]
        FollowedMachine.query.filter(
            FollowedMachine.user_id == current_user.id,
            FollowedMachine.machine_id.in_(machine_ids_to_remove)
        ).delete(synchronize_session=False)
        is_followed = False
    else:
        # Ajouter le suivi de cette machine et tous ses descendants
        descendants = get_all_descendants(machine)
        existing_followed_ids = {
            fm.machine_id for fm in FollowedMachine.query.filter_by(user_id=current_user.id).all()
        }
        
        for descendant in descendants:
            if descendant.id not in existing_followed_ids:
                followed_machine = FollowedMachine(
                    user_id=current_user.id,
                    machine_id=descendant.id
                )
                db.session.add(followed_machine)
        is_followed = True
    
    try:
        db.session.commit()
        return json.dumps({"success": True, "is_followed": is_followed})
    except Exception as exc:
        db.session.rollback()
        return json.dumps({"success": False, "error": str(exc)}), 500


def get_machine_detail_url(machine_id, tab=None):
    """Helper function to generate machine_detail URL with tab parameter"""
    if tab:
        # Stocker dans la session pour cette machine
        session_key = f'machine_{machine_id}_tab'
        session[session_key] = tab
        return url_for("machine_detail", machine_id=machine_id, tab=tab)
    # Essayer d'extraire le tab depuis le referrer
    referrer = request.referrer
    if referrer:
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(referrer)
        query_params = parse_qs(parsed.query)
        if 'tab' in query_params:
            tab = query_params['tab'][0]
            if tab in ['checklists', 'corrective', 'preventive', 'documentation']:
                session_key = f'machine_{machine_id}_tab'
                session[session_key] = tab
                return url_for("machine_detail", machine_id=machine_id, tab=tab)
    # Essayer depuis la session
    session_key = f'machine_{machine_id}_tab'
    if session_key in session:
        tab = session[session_key]
        if tab in ['checklists', 'corrective', 'preventive', 'documentation']:
            return url_for("machine_detail", machine_id=machine_id, tab=tab)
    return url_for("machine_detail", machine_id=machine_id)


@app.route("/machines/<int:machine_id>")
@login_required
def machine_detail(machine_id):
    # Charger la machine avec eager loading des relations fréquemment utilisées
    machine = (
        Machine.query
        .options(
            joinedload(Machine.parent),
            joinedload(Machine.children),
            joinedload(Machine.counters)
        )
        .get_or_404(machine_id)
    )
    
    # Charger les entries avec eager loading
    entries = (
        MaintenanceEntry.query
        .filter_by(machine_id=machine.id)
        .options(joinedload(MaintenanceEntry.report), joinedload(MaintenanceEntry.user))
        .order_by(MaintenanceEntry.created_at.desc())
        .all()
    )
    
    # Charger les maintenances correctives avec eager loading
    corrective_maintenances = (
        CorrectiveMaintenance.query
        .filter_by(machine_id=machine.id)
        .options(joinedload(CorrectiveMaintenance.user), joinedload(CorrectiveMaintenance.products))
        .order_by(CorrectiveMaintenance.created_at.desc())
        .all()
    )
    
    # Récupérer les templates de maintenance
    templates = []
    template_progress = []
    
    # Trouver la machine racine pour vérifier si elle a des compteurs
    root_machine = machine
    while root_machine.parent:
        root_machine = root_machine.parent
    
    # Récupérer les templates si la machine a un compteur OU si la machine racine a des compteurs
    has_own_counter = machine.hour_counter_enabled
    has_root_counters = root_machine.is_root() and root_machine.counters
    has_counter = has_own_counter or has_root_counters
    
    if has_counter:
        # Charger les templates avec eager loading
        templates = (
            PreventiveReport.query
            .filter_by(machine_id=machine.id)
            .options(joinedload(PreventiveReport.components), joinedload(PreventiveReport.counter))
            .order_by(PreventiveReport.name)
            .all()
        )
    
    children = sorted(machine.children, key=lambda c: c.name)
    
    # S'assurer que tous les MaintenanceProgress existent pour cette machine
    if has_counter:
        ensure_all_progress_for_machine(machine)
        # Ne pas modifier les valeurs existantes lors de la simple consultation
        # Les valeurs ne doivent être modifiées que lors de l'enregistrement d'une maintenance
        # ou lors de la modification du compteur
    
    # Construire template_progress si la machine a un compteur (classique ou multiples)
    # Et regrouper les entries par report_id pour affichage sous chaque modèle
    entries_by_report = {}
    if has_counter:
        # Charger progress_records avec eager loading
        progress_records = (
            MaintenanceProgress.query
            .filter_by(machine_id=machine.id)
            .options(joinedload(MaintenanceProgress.report), joinedload(MaintenanceProgress.counter))
            .all()
        )
        progress_map = {(record.report_id, record.counter_id): record for record in progress_records}
        for report in templates:
            # Utiliser counter_id comme clé si disponible, sinon None
            key = (report.id, report.counter_id)
            progress = progress_map.get(key)
            if progress:
                hours_remaining = progress.hours_since
            else:
                # Fallback: calculate initial value (ne devrait normalement pas arriver si ensure_all_progress_for_machine fonctionne)
                hours_remaining = report.periodicity
            template_progress.append({"report": report, "hours_since": hours_remaining})
            
            # Regrouper les entries par report_id
            report_entries = [e for e in entries if e.report_id == report.id]
            entries_by_report[report.id] = sorted(report_entries, key=lambda e: e.created_at, reverse=True)
    
    # Récupérer les documents de la machine
    documents = (
        MachineDocument.query
        .filter_by(machine_id=machine.id)
        .options(joinedload(MachineDocument.user))
        .order_by(MachineDocument.uploaded_at.desc())
        .all()
    )
    
    # Récupérer les modèles de check list avec eager loading
    checklist_templates = (
        ChecklistTemplate.query
        .filter_by(machine_id=machine.id)
        .options(
            joinedload(ChecklistTemplate.items),
            selectinload(ChecklistTemplate.instances).joinedload(ChecklistInstance.user)
        )
        .order_by(ChecklistTemplate.name)
        .all()
    )

    # Récupérer les instances de check lists pour cette machine (historique)
    checklist_instances = (
        ChecklistInstance.query.filter_by(machine_id=machine.id)
        .order_by(ChecklistInstance.created_at.desc())
        .limit(50)
        .all()
    )

    # Récupérer l'onglet actif depuis l'URL ou la session
    active_tab = request.args.get('tab')
    if not active_tab:
        # Essayer depuis la session
        session_key = f'machine_{machine_id}_tab'
        active_tab = session.get(session_key, 'checklists')
    # Valider que l'onglet est valide
    valid_tabs = ['checklists', 'corrective', 'preventive', 'documentation']
    if active_tab not in valid_tabs:
        active_tab = 'checklists'
    # Stocker dans la session
    session_key = f'machine_{machine_id}_tab'
    session[session_key] = active_tab
    
    return render_template(
        "machine_detail.html",
        machine=machine,
        root_machine=root_machine,
        entries=entries,
        entries_by_report=entries_by_report,
        corrective_maintenances=corrective_maintenances,
        templates=templates,
        template_progress=template_progress,
        children=children,
        active_tab=active_tab,
        documents=documents,
        checklist_templates=checklist_templates,
        checklist_instances=checklist_instances,
    )


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


@app.route("/machines/<int:machine_id>/documents/upload", methods=["POST"])
@admin_required
def upload_machine_document(machine_id):
    machine = Machine.query.get_or_404(machine_id)
    
    if 'file' not in request.files:
        flash("Aucun fichier sélectionné", "danger")
        return redirect(get_machine_detail_url(machine_id, 'documentation'))
    
    file = request.files['file']
    if file.filename == '':
        flash("Aucun fichier sélectionné", "danger")
        return redirect(get_machine_detail_url(machine_id, 'documentation'))
    
    if file and allowed_file(file.filename):
        # Sécuriser le nom du fichier
        original_filename = file.filename
        filename = secure_filename(original_filename)
        
        # Ajouter un timestamp pour éviter les collisions
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S_")
        safe_filename = timestamp + filename
        file_path = UPLOAD_FOLDER / safe_filename
        
        # Sauvegarder le fichier
        file.save(str(file_path))
        
        # Créer l'enregistrement en base de données
        document = MachineDocument(
            machine_id=machine_id,
            filename=safe_filename,
            original_filename=original_filename,
            file_path=str(file_path),
            user_id=current_user.id
        )
        db.session.add(document)
        try:
            db.session.commit()
            flash("Document uploadé avec succès", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur lors de l'upload: {exc}", "danger")
    else:
        flash("Seuls les fichiers PDF sont autorisés", "danger")
    
    return redirect(get_machine_detail_url(machine_id, 'documentation'))


@app.route("/machines/<int:machine_id>/documents/<int:document_id>/delete", methods=["POST"])
@admin_required
def delete_machine_document(machine_id, document_id):
    machine = Machine.query.get_or_404(machine_id)
    document = MachineDocument.query.get_or_404(document_id)
    
    if document.machine_id != machine_id:
        flash("Ce document n'appartient pas à cette machine", "danger")
        return redirect(get_machine_detail_url(machine_id, 'documentation'))
    
    # Supprimer le fichier physique
    try:
        if os.path.exists(document.file_path):
            os.remove(document.file_path)
    except Exception as exc:
        flash(f"Erreur lors de la suppression du fichier: {exc}", "warning")
    
    # Supprimer l'enregistrement en base de données
    db.session.delete(document)
    try:
        db.session.commit()
        flash("Document supprimé avec succès", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur lors de la suppression: {exc}", "danger")
    
    return redirect(get_machine_detail_url(machine_id, 'documentation'))


@app.route("/machines/<int:machine_id>/documents/<int:document_id>/download")
@login_required
def download_machine_document(machine_id, document_id):
    machine = Machine.query.get_or_404(machine_id)
    document = MachineDocument.query.get_or_404(document_id)
    
    if document.machine_id != machine_id:
        abort(404)
    
    if not os.path.exists(document.file_path):
        flash("Le fichier n'existe plus", "danger")
        return redirect(url_for("machine_detail", machine_id=machine_id))
    
    return send_from_directory(
        str(UPLOAD_FOLDER),
        document.filename,
        as_attachment=False,
        download_name=document.original_filename
    )


@app.route("/machines/<int:machine_id>/qrcode")
@login_required
def machine_qrcode(machine_id):
    """Génère un QR code pour une machine"""
    machine = Machine.query.get_or_404(machine_id)
    return render_template("machine_qrcode.html", machine=machine)


@app.route("/machines/<int:machine_id>/qrcode/image")
@login_required
def machine_qrcode_image(machine_id):
    """Génère l'image du QR code pour une machine"""
    machine = Machine.query.get_or_404(machine_id)
    
    # Générer le QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(str(machine.id))  # Encoder l'ID de la machine
    qr.make(fit=True)
    
    # Créer l'image
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convertir en BytesIO pour l'envoyer
    img_io = BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)
    
    return Response(img_io.getvalue(), mimetype='image/png')


@app.route("/scan-qrcode")
@login_required
def scan_qrcode():
    """Page pour scanner un QR code"""
    return render_template("scan_qrcode.html")


@app.route("/machines/qrcode/<int:machine_id>")
@login_required
def qrcode_redirect(machine_id):
    """Redirige vers la page machine après scan du QR code"""
    machine = Machine.query.get_or_404(machine_id)
    return redirect(url_for("machine_detail", machine_id=machine.id))


# Routes pour les check lists
@app.route("/machines/<int:machine_id>/checklists/new", methods=["GET", "POST"])
@admin_required
def new_checklist_template(machine_id):
    machine = Machine.query.get_or_404(machine_id)
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Le nom de la check list est requis", "danger")
            return redirect(request.url)
        
        # Créer le template
        template = ChecklistTemplate(machine_id=machine_id, name=name)
        db.session.add(template)
        db.session.flush()  # Pour obtenir l'ID
        
        # Récupérer les items depuis le formulaire
        item_labels = request.form.getlist("item_label")
        for idx, label in enumerate(item_labels):
            label = label.strip()
            if label:
                item = ChecklistItem(template_id=template.id, label=label, order=idx)
                db.session.add(item)
        
        try:
            db.session.commit()
            flash("Check list créée avec succès", "success")
            return redirect(url_for("machine_detail", machine_id=machine_id))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)
    
    # Récupérer toutes les checklists existantes pour le dropdown
    all_checklists = ChecklistTemplate.query.order_by(ChecklistTemplate.name).all()
    
    return render_template("checklist_template_form.html", machine=machine, template=None, all_checklists=all_checklists)


@app.route("/machines/<int:machine_id>/checklists/<int:template_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_checklist_template(machine_id, template_id):
    machine = Machine.query.get_or_404(machine_id)
    template = ChecklistTemplate.query.get_or_404(template_id)
    
    if template.machine_id != machine_id:
        flash("Cette check list n'appartient pas à cette machine", "danger")
        return redirect(url_for("machine_detail", machine_id=machine_id))
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Le nom de la check list est requis", "danger")
            return redirect(request.url)
        
        template.name = name
        
        # Supprimer les anciens items
        ChecklistItem.query.filter_by(template_id=template.id).delete()
        
        # Ajouter les nouveaux items
        item_labels = request.form.getlist("item_label")
        for idx, label in enumerate(item_labels):
            label = label.strip()
            if label:
                item = ChecklistItem(template_id=template.id, label=label, order=idx)
                db.session.add(item)
        
        try:
            db.session.commit()
            flash("Check list modifiée avec succès", "success")
            return redirect(get_machine_detail_url(machine_id, 'checklists'))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)
    
    # Récupérer toutes les checklists existantes pour le dropdown
    all_checklists = ChecklistTemplate.query.order_by(ChecklistTemplate.name).all()
    
    return render_template("checklist_template_form.html", machine=machine, template=template, all_checklists=all_checklists)


@app.route("/machines/<int:machine_id>/checklists/<int:template_id>/delete", methods=["POST"])
@admin_required
def delete_checklist_template(machine_id, template_id):
    machine = Machine.query.get_or_404(machine_id)
    template = ChecklistTemplate.query.get_or_404(template_id)
    
    if template.machine_id != machine_id:
        flash("Cette check list n'appartient pas à cette machine", "danger")
        return redirect(url_for("machine_detail", machine_id=machine_id))
    
    db.session.delete(template)
    try:
        db.session.commit()
        flash("Check list supprimée avec succès", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur: {exc}", "danger")
    
    return redirect(get_machine_detail_url(machine_id, 'checklists'))


@app.route("/machines/<int:machine_id>/checklists/<int:template_id>/fill", methods=["GET", "POST"])
@admin_or_technician_required
def fill_checklist(machine_id, template_id):
    machine = Machine.query.get_or_404(machine_id)
    template = ChecklistTemplate.query.get_or_404(template_id)
    
    if template.machine_id != machine_id:
        flash("Cette check list n'appartient pas à cette machine", "danger")
        return redirect(url_for("machine_detail", machine_id=machine_id))
    
    if request.method == "POST":
        # Créer l'instance
        comment = request.form.get("comment", "").strip()
        instance = ChecklistInstance(
            template_id=template_id,
            machine_id=machine_id,
            user_id=current_user.id,
            comment=comment if comment else None
        )
        db.session.add(instance)
        db.session.flush()
        
        # Récupérer les valeurs cochées
        for item in template.items:
            checked = request.form.get(f"item_{item.id}") == "on"
            value = ChecklistInstanceValue(
                instance_id=instance.id,
                item_id=item.id,
                checked=checked
            )
            db.session.add(value)
        
        try:
            db.session.commit()
            # Message automatique pour le chat
            create_chat_message(
                message_type="auto",
                content=f"{current_user.username} a rempli la checklist '{template.name}' sur la machine '{machine.name}'",
                link_url=url_for("checklist_instance_detail", machine_id=machine_id, template_id=template_id, instance_id=instance.id),
                machine_id=machine_id
            )
            flash("Check list remplie avec succès", "success")
            return redirect(get_machine_detail_url(machine_id, 'checklists'))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)
    
    return render_template("checklist_fill.html", machine=machine, template=template)


@app.route("/machines/<int:machine_id>/checklists/<int:template_id>/instances")
@login_required
def checklist_instances_list(machine_id, template_id):
    machine = Machine.query.get_or_404(machine_id)
    template = ChecklistTemplate.query.get_or_404(template_id)
    
    if template.machine_id != machine_id:
        flash("Cette check list n'appartient pas à cette machine", "danger")
        return redirect(url_for("machine_detail", machine_id=machine_id))
    
    instances = ChecklistInstance.query.filter_by(template_id=template_id, machine_id=machine_id).order_by(ChecklistInstance.created_at.desc()).all()
    
    return render_template("checklist_instances_list.html", machine=machine, template=template, instances=instances)


@app.route("/machines/<int:machine_id>/checklists/<int:template_id>/instances/<int:instance_id>")
@login_required
def checklist_instance_detail(machine_id, template_id, instance_id):
    machine = Machine.query.get_or_404(machine_id)
    template = ChecklistTemplate.query.get_or_404(template_id)
    instance = ChecklistInstance.query.get_or_404(instance_id)
    
    if instance.template_id != template_id or instance.machine_id != machine_id:
        abort(404)
    
    return render_template("checklist_instance_detail.html", machine=machine, template=template, instance=instance)


@app.route("/machines/new", methods=["GET", "POST"])
@admin_required
def new_machine():
    parents = Machine.query.order_by(Machine.name).all()
    stocks = Stock.query.order_by(Stock.name).all()
    if request.method == "POST":
        name = request.form["name"].strip()
        code = request.form["code"].strip()
        parent_id = request.form.get("parent_id")
        stock_id_raw = request.form.get("stock_id")
        
        try:
            stock_id = int(stock_id_raw) if stock_id_raw else None
        except (ValueError, TypeError):
            stock_id = None

        if not name or not code:
            flash("Nom et code requis", "danger")
            return redirect(request.url)

        # Vérifier si le code est déjà utilisé
        existing = Machine.query.filter_by(code=code).first()
        if existing:
            flash("Ce code est déjà utilisé par une autre machine", "danger")
            return redirect(request.url)

        parent = Machine.query.get(parent_id) if parent_id else None
        if parent and parent.depth() >= 4:
            flash("Profondeur maximale de 5 niveaux atteinte", "danger")
            return redirect(request.url)

        stock = Stock.query.get(stock_id) if stock_id else None
        
        # Déterminer si c'est une machine racine
        is_root = not parent_id
        
        # Gérer le color_index
        color_index = 0
        if is_root:
            # Pour les machines racines, récupérer le color_index du formulaire
            try:
                color_index = int(request.form.get("color_index", 0))
                if color_index < 0 or color_index > 9:
                    color_index = 0
            except (ValueError, TypeError):
                color_index = 0
        else:
            # Pour les sous-machines, hériter du color_index de la machine racine
            root_machine = parent
            while root_machine.parent:
                root_machine = root_machine.parent
            color_index = root_machine.color_index if root_machine.color_index is not None else 0
        
        # Gérer les compteurs multiples pour les machines racines
        counters_data = []
        if is_root:
            # Récupérer tous les compteurs depuis le formulaire
            counter_names = [k for k in request.form.keys() if k.startswith("counter_name_")]
            for key in counter_names:
                index = key.replace("counter_name_", "")
                counter_name = request.form.get(f"counter_name_{index}", "").strip()
                if counter_name:
                    counter_value = float(request.form.get(f"counter_value_{index}") or 0.0)
                    counter_unit = request.form.get(f"counter_unit_{index}", "").strip() or "h"
                    counters_data.append({
                        "name": counter_name,
                        "value": counter_value,
                        "unit": counter_unit
                    })
        
        # Gérer le compteur unique pour les sous-machines
        has_counter = request.form.get("hour_counter") == "on"
        initial_hours = float(request.form.get("initial_hours") or 0.0)
        counter_unit = request.form.get("counter_unit", "").strip() or None

        machine = Machine(
            name=name,
            code=code,
            parent=parent,
            hour_counter_enabled=has_counter if not is_root else False,  # Pas de compteur classique pour les racines avec compteurs multiples
            hours=initial_hours if has_counter and not is_root else 0.0,
            counter_unit=counter_unit if has_counter and not is_root else None,
            stock=stock,
            color_index=color_index,
        )
        db.session.add(machine)
        db.session.flush()  # Pour obtenir l'ID de la machine
        
        # Créer les compteurs multiples pour les machines racines
        if is_root and counters_data:
            counter_names_used = set()
            for counter_data in counters_data:
                if counter_data["name"] in counter_names_used:
                    flash(f"Le nom de compteur '{counter_data['name']}' est en double. Veuillez utiliser des noms uniques.", "danger")
                    db.session.rollback()
                    return redirect(request.url)
                counter_names_used.add(counter_data["name"])
                
                counter = Counter(
                    machine_id=machine.id,
                    name=counter_data["name"],
                    value=counter_data["value"],
                    unit=counter_data["unit"]
                )
                db.session.add(counter)
        
        try:
            db.session.commit()
            if is_root and counters_data:
                flash(f"Machine créée avec {len(counters_data)} compteur(s)", "success")
            else:
                flash("Machine créée", "success")
            return redirect(url_for("machines"))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)

    # Récupérer tous les modèles existants pour la liste déroulante
    all_reports = PreventiveReport.query.order_by(PreventiveReport.name).all()
    
    return render_template("machine_form.html", parents=parents, machine=None, is_edit=False, stocks=stocks)


@app.route("/machines/<int:machine_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_machine(machine_id):
    machine = Machine.query.get_or_404(machine_id)
    parents = Machine.query.filter(Machine.id != machine_id).order_by(Machine.name).all()
    stocks = Stock.query.order_by(Stock.name).all()
    
    if request.method == "POST":
        name = request.form["name"].strip()
        code = request.form["code"].strip()
        parent_id = request.form.get("parent_id")
        has_counter = request.form.get("hour_counter") == "on"
        initial_hours = float(request.form.get("initial_hours") or 0.0)
        counter_unit = request.form.get("counter_unit", "").strip() or None
        stock_id_raw = request.form.get("stock_id")
        
        try:
            stock_id = int(stock_id_raw) if stock_id_raw else None
        except (ValueError, TypeError):
            stock_id = None
        
        if not name or not code:
            flash("Nom et code requis", "danger")
            return redirect(request.url)
        
        # Vérifier si le code est déjà utilisé par une autre machine
        existing = Machine.query.filter_by(code=code).first()
        if existing and existing.id != machine.id:
            flash("Ce code est déjà utilisé par une autre machine", "danger")
            return redirect(request.url)
        
        parent = Machine.query.get(parent_id) if parent_id else None
        if parent:
            # Vérifier qu'on ne crée pas une boucle (la machine ne peut pas être son propre parent ou ancêtre)
            if parent.id == machine.id:
                flash("Une machine ne peut pas être son propre parent", "danger")
                return redirect(request.url)
            # Vérifier qu'on ne crée pas une boucle en vérifiant les ancêtres
            current = parent
            while current:
                if current.id == machine.id:
                    flash("Une machine ne peut pas être parente d'une de ses sous-machines", "danger")
                    return redirect(request.url)
                current = current.parent
            if parent.depth() >= 4:
                flash("Profondeur maximale de 5 niveaux atteinte", "danger")
                return redirect(request.url)
        
        stock = Stock.query.get(stock_id) if stock_id else None
        
        # Gérer le color_index
        is_root = not parent
        if is_root:
            # Pour les machines racines, récupérer le color_index du formulaire
            try:
                color_index = int(request.form.get("color_index", machine.color_index or 0))
                if color_index < 0 or color_index > 9:
                    color_index = machine.color_index or 0
            except (ValueError, TypeError):
                color_index = machine.color_index or 0
            machine.color_index = color_index
        else:
            # Pour les sous-machines, hériter du color_index de la machine racine
            root_machine = parent
            while root_machine.parent:
                root_machine = root_machine.parent
            machine.color_index = root_machine.color_index if root_machine.color_index is not None else 0
        
        machine.name = name
        machine.code = code
        machine.parent = parent
        machine.hour_counter_enabled = has_counter
        machine.counter_unit = counter_unit if has_counter else None
        machine.stock = stock
        
        if has_counter:
            # Si on active le compteur et qu'il n'y avait pas d'heures, utiliser initial_hours
            if machine.hours == 0.0:
                machine.hours = initial_hours
        else:
            # Si on désactive le compteur, on peut garder les heures ou les mettre à 0
            # On garde les heures pour l'instant
            pass
        
        # Gérer les compteurs multiples pour les machines racines
        is_root = not parent
        if is_root:
            # Récupérer tous les compteurs existants
            existing_counters = {c.id: c for c in Counter.query.filter_by(machine_id=machine_id).all()}
            
            # Traiter les compteurs du formulaire
            counter_names_used = set()
            counter_index = 0
            while True:
                counter_id_input = request.form.get(f"counter_id_{counter_index}")
                counter_name = request.form.get(f"counter_name_{counter_index}", "").strip()
                counter_value = request.form.get(f"counter_value_{counter_index}")
                counter_unit = request.form.get(f"counter_unit_{counter_index}", "").strip() or None
                counter_delete = request.form.get(f"counter_delete_{counter_id_input}") == "1" if counter_id_input else False
                
                # Si aucun champ n'existe pour cet index, on a fini
                if not counter_id_input and not counter_name:
                    break
                
                # Si le compteur est marqué pour suppression
                if counter_delete and counter_id_input:
                    counter_id = int(counter_id_input)
                    if counter_id in existing_counters:
                        db.session.delete(existing_counters[counter_id])
                        del existing_counters[counter_id]
                    counter_index += 1
                    continue
                
                # Si c'est un compteur existant à modifier
                if counter_id_input and counter_name:
                    counter_id = int(counter_id_input)
                    if counter_id in existing_counters:
                        counter = existing_counters[counter_id]
                        # Vérifier l'unicité du nom (sauf pour le compteur lui-même)
                        other_counter = Counter.query.filter_by(machine_id=machine_id, name=counter_name).first()
                        if other_counter and other_counter.id != counter_id:
                            flash(f"Le nom de compteur '{counter_name}' est déjà utilisé.", "danger")
                            db.session.rollback()
                            return redirect(request.url)
                        counter.name = counter_name
                        try:
                            counter.value = float(counter_value) if counter_value else 0.0
                        except (ValueError, TypeError):
                            counter.value = 0.0
                        counter.unit = counter_unit
                        del existing_counters[counter_id]  # Retirer de la liste des existants
                    counter_index += 1
                    continue
                
                # Si c'est un nouveau compteur
                if not counter_id_input and counter_name:
                    # Vérifier l'unicité du nom
                    if counter_name in counter_names_used:
                        flash(f"Le nom de compteur '{counter_name}' est en double dans le formulaire.", "danger")
                        db.session.rollback()
                        return redirect(request.url)
                    existing = Counter.query.filter_by(machine_id=machine_id, name=counter_name).first()
                    if existing:
                        flash(f"Un compteur avec le nom '{counter_name}' existe déjà.", "danger")
                        db.session.rollback()
                        return redirect(request.url)
                    counter_names_used.add(counter_name)
                    
                    try:
                        value = float(counter_value) if counter_value else 0.0
                    except (ValueError, TypeError):
                        value = 0.0
                    
                    counter = Counter(
                        machine_id=machine_id,
                        name=counter_name,
                        value=value,
                        unit=counter_unit
                    )
                    db.session.add(counter)
                counter_index += 1
            
            # Supprimer les compteurs qui n'ont pas été modifiés (supprimés du formulaire)
            for counter_id, counter in existing_counters.items():
                db.session.delete(counter)
        
        try:
            db.session.commit()
            flash("Machine modifiée", "success")
            return redirect(get_machine_detail_url(machine.id))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)
    
    # Récupérer les compteurs existants si c'est une machine racine
    existing_counters = []
    if machine.is_root():
        existing_counters = Counter.query.filter_by(machine_id=machine_id).order_by(Counter.name).all()
    
    return render_template("machine_form.html", parents=parents, machine=machine, is_edit=True, stocks=stocks, existing_counters=existing_counters)


@app.route("/machines/<int:machine_id>/delete", methods=["POST"])
@admin_required
def delete_machine(machine_id):
    machine = Machine.query.get_or_404(machine_id)
    
    # Vérifier si la machine a des enfants (sous-machines)
    if machine.children:
        flash(f"Impossible de supprimer cette machine : elle contient {len(machine.children)} sous-machine(s).", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))
    
    # Vérifier si la machine a des maintenances préventives
    preventive_count = MaintenanceEntry.query.filter_by(machine_id=machine_id).count()
    if preventive_count > 0:
        flash(f"Impossible de supprimer cette machine : elle est utilisée dans {preventive_count} maintenance(s) préventive(s).", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))
    
    # Vérifier si la machine a des maintenances correctives
    corrective_count = CorrectiveMaintenance.query.filter_by(machine_id=machine_id).count()
    if corrective_count > 0:
        flash(f"Impossible de supprimer cette machine : elle est utilisée dans {corrective_count} maintenance(s) corrective(s).", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))
    
    # Vérifier si la machine a des relevés de compteur
    counter_logs_count = CounterLog.query.filter_by(machine_id=machine_id).count()
    if counter_logs_count > 0:
        flash(f"Impossible de supprimer cette machine : elle a {counter_logs_count} relevé(s) de compteur.", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))
    
    # Supprimer les plans de maintenance associés (cascade)
    PreventiveReport.query.filter_by(machine_id=machine_id).delete()
    MaintenanceProgress.query.filter_by(machine_id=machine_id).delete()
    
    # Supprimer la machine
    db.session.delete(machine)
    
    try:
        db.session.commit()
        flash("Machine supprimée", "success")
        return redirect(url_for("machines"))
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur lors de la suppression : {exc}", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))


@app.route("/machines/<int:machine_id>/counters")
@admin_required
def machine_counters(machine_id):
    """Gestion des compteurs multiples pour une machine racine"""
    machine = Machine.query.get_or_404(machine_id)
    
    if not machine.is_root():
        flash("Les compteurs multiples sont uniquement disponibles pour les machines racines.", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))
    
    counters = Counter.query.filter_by(machine_id=machine_id).order_by(Counter.name).all()
    return render_template("machine_counters.html", machine=machine, counters=counters)


@app.route("/machines/<int:machine_id>/counters/new", methods=["GET", "POST"])
@admin_required
def new_counter(machine_id):
    """Créer un nouveau compteur pour une machine racine"""
    machine = Machine.query.get_or_404(machine_id)
    
    if not machine.is_root():
        flash("Les compteurs multiples sont uniquement disponibles pour les machines racines.", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        initial_value = float(request.form.get("initial_value") or 0.0)
        unit = request.form.get("unit", "").strip() or None
        
        if not name:
            flash("Le nom du compteur est requis", "danger")
            return redirect(request.url)
        
        # Vérifier si un compteur avec ce nom existe déjà
        existing = Counter.query.filter_by(machine_id=machine_id, name=name).first()
        if existing:
            flash(f"Un compteur avec le nom '{name}' existe déjà pour cette machine.", "danger")
            return redirect(request.url)
        
        counter = Counter(
            machine_id=machine_id,
            name=name,
            value=initial_value,
            unit=unit
        )
        db.session.add(counter)
        try:
            db.session.commit()
            flash("Compteur créé", "success")
            return redirect(url_for("machine_counters", machine_id=machine_id))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)
    
    return render_template("counter_form.html", machine=machine, counter=None, is_edit=False)


@app.route("/machines/<int:machine_id>/counters/<int:counter_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_counter(machine_id, counter_id):
    """Modifier un compteur"""
    machine = Machine.query.get_or_404(machine_id)
    counter = Counter.query.get_or_404(counter_id)
    
    if counter.machine_id != machine_id:
        flash("Ce compteur n'appartient pas à cette machine", "danger")
        return redirect(url_for("machine_detail", machine_id=machine_id))
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        unit = request.form.get("unit", "").strip() or None
        
        if not name:
            flash("Le nom du compteur est requis", "danger")
            return redirect(request.url)
        
        # Vérifier si un autre compteur avec ce nom existe déjà
        existing = Counter.query.filter_by(machine_id=machine_id, name=name).first()
        if existing and existing.id != counter_id:
            flash(f"Un compteur avec le nom '{name}' existe déjà pour cette machine.", "danger")
            return redirect(request.url)
        
        counter.name = name
        counter.unit = unit
        
        try:
            db.session.commit()
            flash("Compteur modifié", "success")
            return redirect(url_for("machine_counters", machine_id=machine_id))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)
    
    return render_template("counter_form.html", machine=machine, counter=counter, is_edit=True)


@app.route("/machines/<int:machine_id>/counters/<int:counter_id>/delete", methods=["POST"])
@admin_required
def delete_counter(machine_id, counter_id):
    """Supprimer un compteur"""
    machine = Machine.query.get_or_404(machine_id)
    counter = Counter.query.get_or_404(counter_id)
    
    if counter.machine_id != machine_id:
        flash("Ce compteur n'appartient pas à cette machine", "danger")
        return redirect(url_for("machine_detail", machine_id=machine_id))
    
    # Vérifier si le compteur est utilisé dans des maintenances
    reports_using_counter = PreventiveReport.query.filter_by(counter_id=counter_id).count()
    if reports_using_counter > 0:
        flash(f"Impossible de supprimer ce compteur : il est utilisé dans {reports_using_counter} modèle(s) de maintenance.", "danger")
        return redirect(url_for("machine_counters", machine_id=machine_id))
    
    db.session.delete(counter)
    try:
        db.session.commit()
        flash("Compteur supprimé", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur: {exc}", "danger")
    return redirect(url_for("machine_counters", machine_id=machine_id))


@app.route("/products", methods=["GET", "POST"])
@login_required
def products():
    if request.method == "POST":
        # Seul l'admin peut créer/modifier des produits
        if current_user.user_type != "admin":
            flash("Accès refusé : cette fonctionnalité est réservée aux administrateurs.", "danger")
            return redirect(url_for("products"))
        name = request.form["name"].strip()
        code = request.form["code"].strip()
        price = float(request.form.get("price") or 0.0)
        supplier_name = request.form.get("supplier_name", "").strip()
        supplier_reference = request.form.get("supplier_reference", "").strip()
        location_code = request.form.get("location_code", "").strip()
        minimum_stock = float(request.form.get("minimum_stock") or 0.0)
        if not name or not code:
            flash("Nom et code requis", "danger")
        else:
            # Vérifier si le code existe déjà
            existing = Product.query.filter_by(code=code).first()
            if existing:
                flash(f"Le code '{code}' est déjà utilisé par un autre produit", "danger")
            else:
                product = Product(
                    name=name,
                    code=code,
                    price=price,
                    supplier_name=supplier_name if supplier_name else None,
                    supplier_reference=supplier_reference if supplier_reference else None,
                    location_code=location_code if location_code else None,
                    minimum_stock=minimum_stock,
                )
                db.session.add(product)
                
                # Gérer le stock si sélectionné
                stock_id_raw = request.form.get("stock_id", "").strip()
                initial_quantity_raw = request.form.get("initial_quantity", "0").strip()
                
                if stock_id_raw:
                    try:
                        stock_id = int(stock_id_raw)
                        stock = Stock.query.get(stock_id)
                        if stock:
                            try:
                                initial_quantity = float(initial_quantity_raw) if initial_quantity_raw else 0.0
                                if initial_quantity > 0:
                                    # Commit d'abord pour avoir l'ID du produit
                                    db.session.flush()
                                    
                                    # Vérifier si le produit existe déjà dans ce stock
                                    existing_stock_product = StockProduct.query.filter_by(
                                        stock_id=stock_id, product_id=product.id
                                    ).first()
                                    
                                    if existing_stock_product:
                                        # Ajouter à la quantité existante
                                        existing_stock_product.quantity += initial_quantity
                                    else:
                                        # Créer une nouvelle entrée
                                        stock_product = StockProduct(
                                            stock_id=stock_id,
                                            product_id=product.id,
                                            quantity=initial_quantity
                                        )
                                        db.session.add(stock_product)
                            except (ValueError, TypeError):
                                pass  # Ignorer si la quantité n'est pas valide
                    except (ValueError, TypeError):
                        pass  # Ignorer si le stock_id n'est pas valide
                
                try:
                    db.session.commit()
                    flash("Produit créé", "success")
                except Exception as exc:
                    db.session.rollback()
                    flash(f"Erreur: {exc}", "danger")
        return redirect(request.url)

    # Récupérer les paramètres de filtrage
    filter_name = request.args.get('filter_name', '').strip().lower()
    filter_code = request.args.get('filter_code', '').strip().lower()
    filter_supplier = request.args.get('filter_supplier', '').strip().lower()
    filter_min_stock = request.args.get('filter_min_stock', '').strip()
    filter_low_stock = request.args.get('filter_low_stock', '').strip() == '1'
    filter_stock_id_raw = request.args.get('filter_stock_id', '').strip()
    try:
        filter_stock_id = int(filter_stock_id_raw) if filter_stock_id_raw else None
    except (ValueError, TypeError):
        filter_stock_id = None
    
    # Charger tous les produits avec eager loading des stock_products
    all_products = (
        Product.query
        .order_by(Product.name)
        .all()
    )
    all_stocks = Stock.query.order_by(Stock.name).all()
    # Identifier le stock principal (même logique que la page stocks : premier stock par ID)
    first_stock = Stock.query.order_by(Stock.id).first()
    main_stock_id = first_stock.id if first_stock else None
    
    # Charger toutes les quantités de stock en une seule requête pour éviter N+1
    stock_products_dict = {}
    if first_stock:
        stock_products = (
            StockProduct.query
            .filter_by(stock_id=first_stock.id)
            .all()
        )
        stock_products_dict = {sp.product_id: sp.quantity for sp in stock_products}
    
    # Charger toutes les quantités de stock en une seule requête pour éviter N+1
    all_stock_products = (
        StockProduct.query
        .filter(StockProduct.stock_id.in_([s.id for s in all_stocks]))
        .all()
    )
    
    # Créer un dictionnaire pour accéder rapidement aux quantités
    # Structure: {product_id: {stock_id: quantity}}
    quantities_by_product = {}
    total_by_product = {}  # Somme totale des quantités par produit
    
    # Initialiser les dictionnaires pour tous les produits
    for product in all_products:
        quantities_by_product[product.id] = {}
        total_by_product[product.id] = 0.0
    
    # Remplir avec les données chargées
    for sp in all_stock_products:
        if sp.product_id not in quantities_by_product:
            quantities_by_product[sp.product_id] = {}
        quantities_by_product[sp.product_id][sp.stock_id] = sp.quantity
        total_by_product[sp.product_id] = total_by_product.get(sp.product_id, 0.0) + sp.quantity
    
    # Appliquer les filtres
    filtered_products = all_products
    
    if filter_name:
        filtered_products = [p for p in filtered_products if filter_name in p.name.lower()]
    
    if filter_code:
        filtered_products = [p for p in filtered_products if filter_code in p.code.lower()]
    
    if filter_supplier:
        filtered_products = [p for p in filtered_products if p.supplier_name and filter_supplier in p.supplier_name.lower()]
    
    if filter_min_stock:
        try:
            min_stock_value = float(filter_min_stock)
            filtered_products = [p for p in filtered_products if p.minimum_stock >= min_stock_value]
        except (ValueError, TypeError):
            pass
    
    if filter_low_stock:
        # Filtrer les produits dont le stock total est inférieur au stock minimum
        filtered_products = [
            p for p in filtered_products
            if p.minimum_stock > 0 and total_by_product[p.id] < p.minimum_stock
        ]
    
    if filter_stock_id:
        # Filtrer les produits qui ont une quantité > 0 dans le stock sélectionné
        filtered_products = [
            p for p in filtered_products
            if quantities_by_product[p.id].get(filter_stock_id, 0.0) > 0
        ]
    
    return render_template(
        "products.html",
        products=filtered_products,
        stocks=all_stocks,
        quantities_by_product=quantities_by_product,
        total_by_product=total_by_product,
        filter_name=filter_name,
        filter_code=filter_code,
        filter_supplier=filter_supplier,
        filter_min_stock=filter_min_stock,
        filter_low_stock=filter_low_stock,
        filter_stock_id=filter_stock_id,
        main_stock_id=main_stock_id,
    )


@app.route("/products/export")
@login_required
def export_products():
    # Récupérer les mêmes filtres que la page produits
    filter_name = request.args.get('filter_name', '').strip().lower()
    filter_code = request.args.get('filter_code', '').strip().lower()
    filter_supplier = request.args.get('filter_supplier', '').strip().lower()
    filter_min_stock = request.args.get('filter_min_stock', '').strip()
    filter_low_stock = request.args.get('filter_low_stock', '').strip() == '1'
    filter_stock_id_raw = request.args.get('filter_stock_id', '').strip()
    try:
        filter_stock_id = int(filter_stock_id_raw) if filter_stock_id_raw else None
    except (ValueError, TypeError):
        filter_stock_id = None
    
    all_products = Product.query.order_by(Product.name).all()
    all_stocks = Stock.query.order_by(Stock.name).all()
    
    # Créer un dictionnaire pour accéder rapidement aux quantités
    quantities_by_product = {}
    total_by_product = {}
    for product in all_products:
        quantities_by_product[product.id] = {}
        total = 0.0
        for stock in all_stocks:
            stock_product = StockProduct.query.filter_by(
                product_id=product.id, stock_id=stock.id
            ).first()
            qty = stock_product.quantity if stock_product else 0.0
            quantities_by_product[product.id][stock.id] = qty
            total += qty
        total_by_product[product.id] = total
    
    # Appliquer les mêmes filtres que la page produits
    filtered_products = all_products
    
    if filter_name:
        filtered_products = [p for p in filtered_products if filter_name in p.name.lower()]
    if filter_code:
        filtered_products = [p for p in filtered_products if filter_code in p.code.lower()]
    if filter_supplier:
        filtered_products = [p for p in filtered_products if p.supplier_name and filter_supplier in p.supplier_name.lower()]
    if filter_min_stock:
        try:
            min_stock_value = float(filter_min_stock)
            filtered_products = [p for p in filtered_products if p.minimum_stock >= min_stock_value]
        except (ValueError, TypeError):
            pass
    if filter_low_stock:
        filtered_products = [
            p for p in filtered_products
            if p.minimum_stock > 0 and total_by_product[p.id] < p.minimum_stock
        ]
    if filter_stock_id:
        filtered_products = [
            p for p in filtered_products
            if quantities_by_product[p.id].get(filter_stock_id, 0.0) > 0
        ]
    
    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Produits"
    
    # En-têtes
    headers = ["Nom", "Code", "Code emplacement", "Prix", "Fournisseur", "REF", "Stock min."]
    if filter_stock_id:
        for stock in all_stocks:
            if stock.id == filter_stock_id:
                headers.append(stock.name)
                break
    else:
        for stock in all_stocks:
            headers.append(stock.name)
    headers.append("Total stocks")
    
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    for product in filtered_products:
        row = [
            product.name,
            product.code,
            product.location_code or "-",
            product.price,
            product.supplier_name or "-",
            product.supplier_reference or "-",
            product.minimum_stock
        ]
        if filter_stock_id:
            row.append(quantities_by_product[product.id].get(filter_stock_id, 0.0))
        else:
            for stock in all_stocks:
                row.append(quantities_by_product[product.id][stock.id])
        row.append(total_by_product[product.id])
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=produits_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/api/products/<int:product_id>")
@login_required
def get_product_data(product_id):
    """API pour récupérer les données d'un produit"""
    product = Product.query.get_or_404(product_id)
    
    # Récupérer les stocks où ce produit est présent
    stock_products = StockProduct.query.filter_by(product_id=product_id).all()
    stocks_data = []
    for sp in stock_products:
        stocks_data.append({
            "stock_id": sp.stock_id,
            "quantity": sp.quantity
        })
    
    return json.dumps({
        "success": True,
        "product": {
            "id": product.id,
            "name": product.name,
            "code": product.code,
            "location_code": product.location_code or "",
            "price": product.price,
            "supplier_name": product.supplier_name or "",
            "supplier_reference": product.supplier_reference or "",
            "minimum_stock": product.minimum_stock,
            "stocks": stocks_data
        }
    })


@app.route("/products/<int:product_id>/edit", methods=["POST"])
@admin_required
def edit_product(product_id):
    product = Product.query.get_or_404(product_id)
    name = request.form["name"].strip()
    code = request.form["code"].strip()
    price = float(request.form.get("price") or 0.0)
    supplier_name = request.form.get("supplier_name", "").strip()
    supplier_reference = request.form.get("supplier_reference", "").strip()
    location_code = request.form.get("location_code", "").strip()
    minimum_stock = float(request.form.get("minimum_stock") or 0.0)
    
    if not name or not code:
        flash("Nom et code requis", "danger")
    else:
        # Vérifier si le code est déjà utilisé par un autre produit
        existing = Product.query.filter_by(code=code).first()
        if existing and existing.id != product.id:
            flash("Ce code est déjà utilisé par un autre produit", "danger")
        else:
            product.name = name
            product.code = code
            product.price = price
            product.supplier_name = supplier_name if supplier_name else None
            product.supplier_reference = supplier_reference if supplier_reference else None
            product.location_code = location_code if location_code else None
            product.minimum_stock = minimum_stock
            
            # Gérer le changement de stock
            current_stock_id_raw = request.form.get("current_stock_id", "").strip()
            if current_stock_id_raw:
                try:
                    new_stock_id = int(current_stock_id_raw)
                    new_stock = Stock.query.get(new_stock_id)
                    
                    if new_stock:
                        # Récupérer tous les stocks où ce produit est présent
                        all_stock_products = StockProduct.query.filter_by(product_id=product.id).all()
                        
                        # Si le produit est dans plusieurs stocks, on prend le premier avec quantité > 0
                        # Sinon, on prend le premier trouvé
                        old_stock_product = None
                        if all_stock_products:
                            # Chercher un stock avec quantité > 0
                            for sp in all_stock_products:
                                if sp.quantity > 0:
                                    old_stock_product = sp
                                    break
                            # Si aucun avec quantité > 0, prendre le premier
                            if not old_stock_product and all_stock_products:
                                old_stock_product = all_stock_products[0]
                        
                        if old_stock_product:
                            old_quantity = old_stock_product.quantity
                            
                            # Si on change de stock, déplacer la quantité
                            if old_stock_product.stock_id != new_stock_id:
                                # Vérifier si le produit existe déjà dans le nouveau stock
                                new_stock_product = StockProduct.query.filter_by(
                                    stock_id=new_stock_id, product_id=product.id
                                ).first()
                                
                                if new_stock_product:
                                    # Ajouter la quantité à l'existant
                                    new_stock_product.quantity += old_quantity
                                else:
                                    # Créer une nouvelle entrée
                                    new_stock_product = StockProduct(
                                        stock_id=new_stock_id,
                                        product_id=product.id,
                                        quantity=old_quantity
                                    )
                                    db.session.add(new_stock_product)
                                
                                # Supprimer l'ancienne entrée
                                db.session.delete(old_stock_product)
                        else:
                            # Le produit n'est dans aucun stock, créer une entrée dans le nouveau stock avec quantité 0
                            new_stock_product = StockProduct.query.filter_by(
                                stock_id=new_stock_id, product_id=product.id
                            ).first()
                            
                            if not new_stock_product:
                                new_stock_product = StockProduct(
                                    stock_id=new_stock_id,
                                    product_id=product.id,
                                    quantity=0.0
                                )
                                db.session.add(new_stock_product)
                except (ValueError, TypeError):
                    pass  # Ignorer si le stock_id n'est pas valide
            
            try:
                db.session.commit()
                flash("Produit modifié", "success")
            except Exception as exc:
                db.session.rollback()
                flash(f"Erreur: {exc}", "danger")
    
    return redirect(url_for("products"))


@app.route("/products/import", methods=["GET", "POST"])
@admin_required
def import_products():
    if request.method == "POST":
        if "file" not in request.files:
            flash("Aucun fichier sélectionné", "danger")
            return redirect(url_for("products"))
        
        file = request.files["file"]
        if not file or file.filename == "":
            flash("Aucun fichier sélectionné", "danger")
            return redirect(url_for("products"))
        
        # Vérifier l'extension du fichier
        filename = file.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
            flash("Le fichier doit être au format Excel (.xlsx ou .xls)", "danger")
            return redirect(url_for("products"))
        
        try:
            # Lire le fichier Excel depuis le stream
            # Convertir le FileStorage en BytesIO pour openpyxl
            file_content = file.read()
            if not file_content:
                flash("Le fichier est vide", "danger")
                return redirect(url_for("products"))
            
            file_stream = BytesIO(file_content)
            wb = load_workbook(file_stream, data_only=True)
            ws = wb.active
            
            if ws.max_row < 2:
                flash("Le fichier Excel doit contenir au moins une ligne de données (en plus des en-têtes)", "danger")
                return redirect(url_for("products"))
            
            # Parcourir les lignes (en supposant que la première ligne contient les en-têtes)
            imported_count = 0
            skipped_count = 0
            errors = []
            
            # Parcourir à partir de la ligne 2 (en supposant que la ligne 1 contient les en-têtes)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Ignorer les lignes vides
                if not row or not any(row):
                    continue
                
                # Extraire les valeurs (nom, code, prix, fournisseur, stock min)
                # Gérer les valeurs None et convertir en string
                name = None
                code = None
                if row[0] is not None:
                    name_str = str(row[0]).strip()
                    name = name_str if name_str else None
                if row[1] is not None:
                    code_str = str(row[1]).strip()
                    code = code_str if code_str else None
                
                price = 0.0
                supplier_name = None
                minimum_stock = 0.0
                
                # Vérifier que nom et code sont présents
                if not name or not code:
                    skipped_count += 1
                    errors.append(f"Ligne {row_idx}: Nom ou code manquant")
                    continue
                
                # Extraire le prix (colonne 2)
                if len(row) > 2 and row[2] is not None:
                    try:
                        price = float(row[2])
                    except (ValueError, TypeError):
                        price = 0.0
                
                # Extraire le fournisseur (colonne 3)
                if len(row) > 3 and row[3] is not None:
                    supplier_str = str(row[3]).strip()
                    supplier_name = supplier_str if supplier_str and supplier_str.lower() != "none" else None
                
                # Extraire le stock minimum (colonne 4)
                if len(row) > 4 and row[4] is not None:
                    try:
                        minimum_stock = float(row[4])
                    except (ValueError, TypeError):
                        minimum_stock = 0.0
                
                # Vérifier si le code existe déjà
                existing = Product.query.filter_by(code=code).first()
                if existing:
                    skipped_count += 1
                    errors.append(f"Ligne {row_idx}: Le code '{code}' existe déjà")
                    continue
                
                # Créer le produit
                product = Product(
                    name=name,
                    code=code,
                    price=price,
                    supplier_name=supplier_name,
                    minimum_stock=minimum_stock,
                )
                db.session.add(product)
                imported_count += 1
            
            # Commit toutes les modifications
            try:
                db.session.commit()
                if imported_count > 0:
                    flash(f"{imported_count} produit(s) importé(s) avec succès", "success")
                if skipped_count > 0:
                    flash(f"{skipped_count} produit(s) ignoré(s) (codes existants ou données invalides)", "warning")
                if errors:
                    # Afficher les erreurs dans un message plus détaillé si nécessaire
                    error_summary = "; ".join(errors[:5])  # Limiter à 5 erreurs pour l'affichage
                    if len(errors) > 5:
                        error_summary += f" ... et {len(errors) - 5} autre(s)"
                    flash(f"Détails des erreurs: {error_summary}", "info")
            except Exception as exc:
                db.session.rollback()
                flash(f"Erreur lors de l'import: {exc}", "danger")
        
        except Exception as exc:
            flash(f"Erreur lors de la lecture du fichier: {exc}", "danger")
        
        return redirect(url_for("products"))
    
    return redirect(url_for("products"))


@app.route("/products/<int:product_id>/delete", methods=["POST"])
@admin_required
def delete_product(product_id):
    product = Product.query.get_or_404(product_id)
    
    # Vérifier si le produit est utilisé dans des maintenances correctives
    corrective_maintenance_products = CorrectiveMaintenanceProduct.query.filter_by(product_id=product_id).all()
    has_corrective_maintenance = len(corrective_maintenance_products) > 0
    
    # Vérifier si le produit est utilisé dans des mouvements (qui peuvent être liés à des maintenances préventives)
    movement_items = MovementItem.query.filter_by(product_id=product_id).all()
    has_movements = len(movement_items) > 0
    
    # Supprimer uniquement les quantités en stock (StockProduct)
    # Cela permet de retirer le produit des stocks sans affecter les rapports de maintenance
    stock_products = StockProduct.query.filter_by(product_id=product_id).all()
    for sp in stock_products:
        db.session.delete(sp)
    
    # Si le produit est référencé dans des maintenances, on ne supprime pas le produit lui-même
    # mais on supprime uniquement les quantités en stock
    if has_corrective_maintenance or has_movements:
        try:
            db.session.commit()
            if has_corrective_maintenance and has_movements:
                flash("Produit retiré des stocks. Le produit est conservé car il est référencé dans des rapports de maintenance (corrective et préventive).", "info")
            elif has_corrective_maintenance:
                flash("Produit retiré des stocks. Le produit est conservé car il est référencé dans des rapports de maintenance corrective.", "info")
            else:
                flash("Produit retiré des stocks. Le produit est conservé car il est référencé dans des mouvements liés à des maintenances.", "info")
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur lors de la suppression des quantités en stock : {exc}", "danger")
    else:
        # Si le produit n'est pas dans des maintenances, on peut le supprimer complètement
        db.session.delete(product)
        try:
            db.session.commit()
            flash("Produit supprimé", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur lors de la suppression : {exc}", "danger")
    
    return redirect(url_for("products"))


@app.route("/stocks")
@login_required
def stocks():
    all_stocks = Stock.query.order_by(Stock.name).all()
    # Identifier le stock principal (premier stock par ID)
    first_stock = Stock.query.order_by(Stock.id).first()
    main_stock_id = first_stock.id if first_stock else None
    return render_template("stocks.html", stocks=all_stocks, main_stock_id=main_stock_id)


@app.route("/stocks/new", methods=["GET", "POST"])
@admin_required
def new_stock():
    if request.method == "POST":
        # Vérifier la limite de 10 stocks
        existing_stocks_count = Stock.query.count()
        if existing_stocks_count >= 10:
            flash("Le nombre maximum de stocks est limité à 10", "danger")
            return redirect(url_for("stocks"))
        
        name = request.form["name"].strip()
        code = request.form["code"].strip()
        if not name or not code:
            flash("Nom et code requis", "danger")
        else:
            # Vérifier si le code existe déjà
            existing_stock = Stock.query.filter_by(code=code).first()
            if existing_stock:
                flash(f"Un stock avec le code '{code}' existe déjà. Veuillez choisir un code différent.", "danger")
            else:
                stock = Stock(name=name, code=code)
                db.session.add(stock)
                try:
                    db.session.commit()
                    flash("Stock créé", "success")
                    return redirect(url_for("stocks"))
                except IntegrityError as exc:
                    db.session.rollback()
                    if "UNIQUE constraint failed: stock.code" in str(exc) or "stock.code" in str(exc):
                        flash(f"Un stock avec le code '{code}' existe déjà. Veuillez choisir un code différent.", "danger")
                    else:
                        flash(f"Erreur lors de la création du stock : {exc}", "danger")
                except Exception as exc:
                    db.session.rollback()
                    flash(f"Erreur lors de la création du stock : {exc}", "danger")
    
    # Vérifier la limite avant d'afficher le formulaire
    existing_stocks_count = Stock.query.count()
    can_create = existing_stocks_count < 10
    
    return render_template("stock_form.html", can_create=can_create, existing_stocks_count=existing_stocks_count)


def _get_or_create_stock_product(stock_id, product_id):
    record = StockProduct.query.filter_by(stock_id=stock_id, product_id=product_id).one_or_none()
    if not record:
        record = StockProduct(stock_id=stock_id, product_id=product_id, quantity=0.0)
        db.session.add(record)
    return record


@app.route("/stocks/<int:stock_id>", methods=["GET", "POST"])
@login_required
def manage_stock(stock_id):
    stock = Stock.query.get_or_404(stock_id)
    products = Product.query.order_by(Product.name).all()
    if request.method == "POST":
        # Vérifier les permissions pour modifier les stocks
        if not can_edit_stocks_products() or is_readonly_stocks_products():
            flash("Accès refusé : vous n'avez pas les droits pour modifier ce stock.", "danger")
            return redirect(url_for("manage_stock", stock_id=stock_id))
        action = request.form["action"]
        product_id = int(request.form["product_id"])

        if action == "delete":
            item = StockProduct.query.filter_by(stock_id=stock_id, product_id=product_id).one_or_none()
            if not item:
                flash("Produit introuvable dans le stock", "danger")
                return redirect(request.url)
            db.session.delete(item)
            try:
                db.session.commit()
                flash("Produit retiré du stock", "success")
            except Exception as exc:
                db.session.rollback()
                flash(f"Erreur: {exc}", "danger")
        return redirect(request.url)

    # Récupérer les produits du stock avec leurs informations
    stock_products = StockProduct.query.filter_by(stock_id=stock_id).all()
    products_in_stock = []
    for sp in stock_products:
        products_in_stock.append({
            'product': sp.product,
            'quantity': sp.quantity
        })
    # Trier par nom de produit
    products_in_stock.sort(key=lambda x: x['product'].name)
    
    return render_template("stock_detail.html", stock=stock, products_in_stock=products_in_stock)


@app.route("/stocks/<int:stock_id>/inventory", methods=["GET", "POST"])
@login_required
def create_inventory(stock_id):
    stock = Stock.query.get_or_404(stock_id)
    
    if request.method == "POST":
        # Compter le nombre d'inventaires existants pour ce stock pour générer le nom
        existing_inventories_count = Inventory.query.filter_by(stock_id=stock_id).count()
        inventory_number = existing_inventories_count + 1
        inventory_name = f"{stock.name} #{inventory_number}"
        
        # Créer l'inventaire
        inventory = Inventory(
            stock_id=stock_id,
            user_id=current_user.id,
            name=inventory_name,
            created_at=dt.datetime.utcnow()
        )
        
        # Récupérer les produits du stock avec leurs quantités actuelles
        stock_products = StockProduct.query.filter_by(stock_id=stock_id).all()
        product_quantities = {sp.product_id: sp.quantity for sp in stock_products}
        
        # Récupérer toutes les quantités modifiées depuis le formulaire
        product_ids = request.form.getlist("product_id")
        new_quantities = request.form.getlist("new_quantity")
        comments = request.form.getlist("comment")
        
        has_changes = False
        
        for product_id_str, new_qty_str, comment in zip(product_ids, new_quantities, comments):
            try:
                product_id = int(product_id_str)
                new_quantity = float(new_qty_str) if new_qty_str else 0.0
                previous_quantity = product_quantities.get(product_id, 0.0)
                
                # Ignorer si la quantité n'a pas changé
                if abs(new_quantity - previous_quantity) < 0.01:
                    continue
                
                has_changes = True
                
                # Créer l'item d'inventaire
                item = InventoryItem(
                    inventory=inventory,
                    product_id=product_id,
                    previous_quantity=previous_quantity,
                    new_quantity=new_quantity,
                    comment=comment.strip() if comment else None
                )
                inventory.items.append(item)
                
                # Mettre à jour la quantité dans le stock
                stock_product = StockProduct.query.filter_by(stock_id=stock_id, product_id=product_id).first()
                if stock_product:
                    stock_product.quantity = new_quantity
                else:
                    # Créer une nouvelle entrée si le produit n'était pas dans le stock
                    stock_product = StockProduct(stock_id=stock_id, product_id=product_id, quantity=new_quantity)
                    db.session.add(stock_product)
            
            except (ValueError, TypeError):
                continue
        
        if not has_changes:
            flash("Aucune modification de quantité détectée", "warning")
            return redirect(url_for("create_inventory", stock_id=stock_id))
        
        try:
            db.session.add(inventory)
            db.session.commit()
            # Message automatique pour le chat
            # Récupérer la machine associée au stock si elle existe
            machine_name = ""
            machine_id_for_msg = None
            # Chercher une machine qui utilise ce stock
            machine_with_stock = Machine.query.filter_by(stock_id=stock.id).first()
            if machine_with_stock:
                machine_name = f" sur la machine '{machine_with_stock.name}'"
                machine_id_for_msg = machine_with_stock.id
            
            create_chat_message(
                message_type="auto",
                content=f"{current_user.username} a créé l'inventaire '{inventory.name}'{machine_name}",
                link_url=url_for("inventory_detail", inventory_id=inventory.id),
                machine_id=machine_id_for_msg
            )
            flash("Inventaire enregistré avec succès", "success")
            return redirect(url_for("inventories_list"))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur lors de l'enregistrement de l'inventaire: {exc}", "danger")
            return redirect(url_for("create_inventory", stock_id=stock_id))
    
    # GET: Afficher le formulaire d'inventaire
    # Récupérer tous les produits du stock
    stock_products = StockProduct.query.filter_by(stock_id=stock_id).order_by(StockProduct.product_id).all()
    
    # Créer une liste de tous les produits avec leurs quantités actuelles
    products_data = []
    for sp in stock_products:
        products_data.append({
            'product': sp.product,
            'quantity': sp.quantity
        })
    
    return render_template("inventory_form.html", stock=stock, products_data=products_data)


@app.route("/stocks/<int:stock_id>/inventory/import", methods=["GET", "POST"])
@login_required
def import_inventory(stock_id):
    stock = Stock.query.get_or_404(stock_id)
    
    if request.method == "POST":
        if "file" not in request.files:
            flash("Aucun fichier sélectionné", "danger")
            return redirect(url_for("stocks"))
        
        file = request.files["file"]
        if not file or file.filename == "":
            flash("Aucun fichier sélectionné", "danger")
            return redirect(url_for("stocks"))
        
        # Vérifier l'extension du fichier
        filename = file.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
            flash("Le fichier doit être au format Excel (.xlsx ou .xls)", "danger")
            return redirect(url_for("stocks"))
        
        try:
            # Lire le fichier Excel depuis le stream
            file_content = file.read()
            if not file_content:
                flash("Le fichier est vide", "danger")
                return redirect(url_for("stocks"))
            
            file_stream = BytesIO(file_content)
            wb = load_workbook(file_stream, data_only=True)
            ws = wb.active
            
            if ws.max_row < 2:
                flash("Le fichier Excel doit contenir au moins une ligne de données (en plus des en-têtes)", "danger")
                return redirect(url_for("stocks"))
            
            # Compter le nombre d'inventaires existants pour ce stock pour générer le nom
            existing_inventories_count = Inventory.query.filter_by(stock_id=stock_id).count()
            inventory_number = existing_inventories_count + 1
            inventory_name = f"{stock.name} #{inventory_number}"
            
            # Créer l'inventaire
            inventory = Inventory(
                stock_id=stock_id,
                user_id=current_user.id,
                name=inventory_name,
                created_at=dt.datetime.utcnow()
            )
            
            # Récupérer les produits du stock avec leurs quantités actuelles
            stock_products = StockProduct.query.filter_by(stock_id=stock_id).all()
            product_quantities = {sp.product_id: sp.quantity for sp in stock_products}
            
            imported_count = 0
            skipped_count = 0
            errors = []
            has_changes = False
            
            # Parcourir à partir de la ligne 2 (en supposant que la ligne 1 contient les en-têtes)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Ignorer les lignes vides
                if not row or not any(row):
                    continue
                
                # Extraire les valeurs (nom, code, prix, fournisseur, stock min, quantité)
                name = None
                code = None
                if row[0] is not None:
                    name_str = str(row[0]).strip()
                    name = name_str if name_str else None
                if row[1] is not None:
                    code_str = str(row[1]).strip()
                    code = code_str if code_str else None
                
                price = 0.0
                supplier_name = None
                minimum_stock = 0.0
                quantity = 0.0
                
                # Vérifier que nom et code sont présents
                if not name or not code:
                    skipped_count += 1
                    errors.append(f"Ligne {row_idx}: Nom ou code manquant")
                    continue
                
                # Extraire le prix (colonne 3)
                if len(row) > 2 and row[2] is not None:
                    try:
                        price = float(row[2])
                    except (ValueError, TypeError):
                        price = 0.0
                
                # Extraire le fournisseur (colonne 4)
                if len(row) > 3 and row[3] is not None:
                    supplier_str = str(row[3]).strip()
                    supplier_name = supplier_str if supplier_str and supplier_str.lower() != "none" else None
                
                # Extraire le stock minimum (colonne 5)
                if len(row) > 4 and row[4] is not None:
                    try:
                        minimum_stock = float(row[4])
                    except (ValueError, TypeError):
                        minimum_stock = 0.0
                
                # Extraire la quantité (colonne 6) - obligatoire pour l'inventaire
                if len(row) > 5 and row[5] is not None:
                    try:
                        quantity = float(row[5])
                    except (ValueError, TypeError):
                        skipped_count += 1
                        errors.append(f"Ligne {row_idx}: Quantité invalide")
                        continue
                else:
                    skipped_count += 1
                    errors.append(f"Ligne {row_idx}: Quantité manquante")
                    continue
                
                # Chercher le produit par code
                product = Product.query.filter_by(code=code).first()
                
                # Si le produit n'existe pas, le créer et l'ajouter à la base de données
                # Les nouveaux produits seront visibles dans la page produits après le commit
                if not product:
                    product = Product(
                        name=name,
                        code=code,
                        price=price,
                        supplier_name=supplier_name,
                        minimum_stock=minimum_stock,
                    )
                    db.session.add(product)
                    db.session.flush()  # Pour obtenir l'ID du produit et s'assurer qu'il est dans la session
                    imported_count += 1
                
                # Récupérer la quantité précédente
                previous_quantity = product_quantities.get(product.id, 0.0)
                
                # Ignorer si la quantité n'a pas changé
                if abs(quantity - previous_quantity) < 0.01:
                    continue
                
                has_changes = True
                
                # Créer l'item d'inventaire
                item = InventoryItem(
                    inventory=inventory,
                    product_id=product.id,
                    previous_quantity=previous_quantity,
                    new_quantity=quantity,
                    comment=None
                )
                inventory.items.append(item)
                
                # Mettre à jour la quantité dans le stock
                stock_product = StockProduct.query.filter_by(stock_id=stock_id, product_id=product.id).first()
                if stock_product:
                    stock_product.quantity = quantity
                else:
                    # Créer une nouvelle entrée si le produit n'était pas dans le stock
                    stock_product = StockProduct(stock_id=stock_id, product_id=product.id, quantity=quantity)
                    db.session.add(stock_product)
            
            if not has_changes:
                flash("Aucune modification de quantité détectée", "warning")
                return redirect(url_for("stocks"))
            
            # Commit toutes les modifications
            try:
                db.session.add(inventory)
                db.session.commit()
                
                # Message automatique pour le chat
                machine_name = ""
                machine_id_for_msg = None
                machine_with_stock = Machine.query.filter_by(stock_id=stock.id).first()
                if machine_with_stock:
                    machine_name = f" sur la machine '{machine_with_stock.name}'"
                    machine_id_for_msg = machine_with_stock.id
                
                create_chat_message(
                    message_type="auto",
                    content=f"{current_user.username} a créé l'inventaire '{inventory.name}'{machine_name}",
                    link_url=url_for("inventory_detail", inventory_id=inventory.id),
                    machine_id=machine_id_for_msg
                )
                
                if imported_count > 0:
                    flash(f"Inventaire créé avec succès. {imported_count} nouveau(x) produit(s) créé(s).", "success")
                else:
                    flash("Inventaire créé avec succès", "success")
                if skipped_count > 0:
                    flash(f"{skipped_count} ligne(s) ignorée(s) (données invalides)", "warning")
                if errors:
                    error_summary = "; ".join(errors[:5])
                    if len(errors) > 5:
                        error_summary += f" ... et {len(errors) - 5} autre(s)"
                    flash(f"Détails des erreurs: {error_summary}", "info")
                
                return redirect(url_for("inventories_list"))
            except Exception as exc:
                db.session.rollback()
                flash(f"Erreur lors de l'import: {exc}", "danger")
        
        except Exception as exc:
            flash(f"Erreur lors de la lecture du fichier: {exc}", "danger")
        
        return redirect(url_for("stocks"))
    
    return redirect(url_for("stocks"))


@app.route("/inventories")
@admin_or_manager_required
def inventories_list():
    inventories = Inventory.query.order_by(Inventory.created_at.desc()).all()
    return render_template("inventories_list.html", inventories=inventories)


@app.route("/inventories/<int:inventory_id>")
@admin_or_manager_required
def inventory_detail(inventory_id):
    inventory = Inventory.query.get_or_404(inventory_id)
    return render_template("inventory_detail.html", inventory=inventory)


@app.route("/inventories/export")
@admin_or_manager_required
def export_inventories():
    inventories = Inventory.query.order_by(Inventory.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaires"
    
    # En-têtes
    headers = ["Date", "Stock", "Code Stock", "Utilisateur", "Produit", "Code Produit", "Quantité précédente", "Nouvelle quantité", "Différence", "Commentaire"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Données
    for inventory in inventories:
        if inventory.items:
            for item in inventory.items:
                row = [
                    inventory.created_at.strftime("%d/%m/%Y %H:%M"),
                    inventory.stock.name,
                    inventory.stock.code,
                    inventory.user.username,
                    item.product.name,
                    item.product.code,
                    item.previous_quantity,
                    item.new_quantity,
                    item.new_quantity - item.previous_quantity,
                    item.comment or ""
                ]
                ws.append(row)
        else:
            # Inventaire sans modifications
            row = [
                inventory.created_at.strftime("%d/%m/%Y %H:%M"),
                inventory.stock.name,
                inventory.stock.code,
                inventory.user.username,
                "",
                "",
                "",
                "",
                "",
                "Aucune modification"
            ]
            ws.append(row)
    
    # Ajuster la largeur des colonnes
    column_widths = [18, 20, 15, 15, 25, 15, 18, 18, 12, 30]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=inventaires_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/inventories/<int:inventory_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_inventory(inventory_id):
    """Modifier un inventaire existant"""
    inventory = Inventory.query.get_or_404(inventory_id)
    stock = inventory.stock
    
    if request.method == "POST":
        # Restaurer les quantités précédentes
        for item in inventory.items:
            stock_product = StockProduct.query.filter_by(stock_id=stock.id, product_id=item.product_id).first()
            if stock_product:
                stock_product.quantity = item.previous_quantity
            else:
                # Si le produit n'existe plus dans le stock, on le recrée avec l'ancienne quantité
                stock_product = StockProduct(stock_id=stock.id, product_id=item.product_id, quantity=item.previous_quantity)
                db.session.add(stock_product)
        
        # Supprimer les anciens items
        for item in inventory.items:
            db.session.delete(item)
        inventory.items = []
        
        # Récupérer les nouvelles quantités depuis le formulaire
        product_ids = request.form.getlist("product_id")
        new_quantities = request.form.getlist("new_quantity")
        comments = request.form.getlist("comment")
        
        # Récupérer les quantités actuelles (qui ont été restaurées)
        stock_products = StockProduct.query.filter_by(stock_id=stock.id).all()
        product_quantities = {sp.product_id: sp.quantity for sp in stock_products}
        
        has_changes = False
        
        for product_id_str, new_qty_str, comment in zip(product_ids, new_quantities, comments):
            try:
                product_id = int(product_id_str)
                new_quantity = float(new_qty_str) if new_qty_str else 0.0
                previous_quantity = product_quantities.get(product_id, 0.0)
                
                # Ignorer si la quantité n'a pas changé
                if abs(new_quantity - previous_quantity) < 0.01:
                    continue
                
                has_changes = True
                
                # Créer le nouvel item d'inventaire
                item = InventoryItem(
                    inventory=inventory,
                    product_id=product_id,
                    previous_quantity=previous_quantity,
                    new_quantity=new_quantity,
                    comment=comment.strip() if comment else None
                )
                inventory.items.append(item)
                
                # Mettre à jour la quantité dans le stock
                stock_product = StockProduct.query.filter_by(stock_id=stock.id, product_id=product_id).first()
                if stock_product:
                    stock_product.quantity = new_quantity
                else:
                    stock_product = StockProduct(stock_id=stock.id, product_id=product_id, quantity=new_quantity)
                    db.session.add(stock_product)
            
            except (ValueError, TypeError):
                continue
        
        if not has_changes:
            flash("Aucune modification de quantité détectée", "warning")
            return redirect(url_for("edit_inventory", inventory_id=inventory_id))
        
        # Mettre à jour la date de modification
        inventory.created_at = dt.datetime.utcnow()
        
        try:
            db.session.commit()
            # Message automatique pour le chat
            # Récupérer la machine associée au stock si elle existe
            machine_name = ""
            machine_id_for_msg = None
            # Chercher une machine qui utilise ce stock
            machine_with_stock = Machine.query.filter_by(stock_id=stock.id).first()
            if machine_with_stock:
                machine_name = f" sur la machine '{machine_with_stock.name}'"
                machine_id_for_msg = machine_with_stock.id
            
            create_chat_message(
                message_type="auto",
                content=f"{current_user.username} a modifié l'inventaire du stock '{stock.name}'{machine_name}",
                link_url=url_for("inventory_detail", inventory_id=inventory.id),
                machine_id=machine_id_for_msg
            )
            flash("Inventaire modifié avec succès", "success")
            return redirect(url_for("inventory_detail", inventory_id=inventory.id))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur lors de la modification de l'inventaire: {exc}", "danger")
            return redirect(url_for("edit_inventory", inventory_id=inventory_id))
    
    # GET: Afficher le formulaire de modification
    # Récupérer tous les produits du stock avec leurs quantités actuelles
    stock_products = StockProduct.query.filter_by(stock_id=stock.id).order_by(StockProduct.product_id).all()
    
    # Créer une liste de tous les produits avec leurs quantités actuelles
    # Pour les produits modifiés dans l'inventaire, utiliser la nouvelle quantité
    products_data = []
    inventory_items_dict = {item.product_id: item for item in inventory.items}
    
    for sp in stock_products:
        # Si le produit a été modifié dans cet inventaire, utiliser la nouvelle quantité
        if sp.product_id in inventory_items_dict:
            current_qty = inventory_items_dict[sp.product_id].new_quantity
        else:
            current_qty = sp.quantity
        
        products_data.append({
            'product': sp.product,
            'quantity': current_qty
        })
    
    return render_template("inventory_form.html", stock=stock, products_data=products_data, inventory=inventory, is_edit=True)


@app.route("/inventories/<int:inventory_id>/delete", methods=["POST"])
@admin_required
def delete_inventory(inventory_id):
    """Supprimer un inventaire et restaurer les quantités précédentes"""
    inventory = Inventory.query.get_or_404(inventory_id)
    stock = inventory.stock
    
    # Restaurer les quantités précédentes pour chaque item
    for item in inventory.items:
        stock_product = StockProduct.query.filter_by(stock_id=stock.id, product_id=item.product_id).first()
        if stock_product:
            stock_product.quantity = item.previous_quantity
        else:
            # Si le produit n'existe plus dans le stock, on le recrée avec l'ancienne quantité
            stock_product = StockProduct(stock_id=stock.id, product_id=item.product_id, quantity=item.previous_quantity)
            db.session.add(stock_product)
    
    # Supprimer l'inventaire (les items seront supprimés en cascade)
    db.session.delete(inventory)
    
    try:
        db.session.commit()
        flash("Inventaire supprimé avec succès. Les quantités ont été restaurées.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur lors de la suppression: {exc}", "danger")
    
    return redirect(url_for("inventories_list"))


@app.route("/stocks/<int:stock_id>/delete", methods=["POST"])
@admin_required
def delete_stock(stock_id):
    stock = Stock.query.get_or_404(stock_id)
    
    # Vérifier si le stock est utilisé dans des mouvements
    movements_as_source = Movement.query.filter_by(source_stock_id=stock_id).count()
    movements_as_dest = Movement.query.filter_by(dest_stock_id=stock_id).count()
    
    # Vérifier si le stock est utilisé dans des maintenances préventives
    preventive_maintenances = MaintenanceEntry.query.filter_by(stock_id=stock_id).count()
    
    # Vérifier si le stock est utilisé dans des maintenances correctives
    corrective_maintenances = CorrectiveMaintenance.query.filter_by(stock_id=stock_id).count()
    
    if movements_as_source > 0 or movements_as_dest > 0:
        flash(f"Impossible de supprimer ce stock : il est utilisé dans {movements_as_source + movements_as_dest} mouvement(s).", "danger")
        return redirect(url_for("stocks"))
    
    if preventive_maintenances > 0:
        flash(f"Impossible de supprimer ce stock : il est utilisé dans {preventive_maintenances} maintenance(s) préventive(s).", "danger")
        return redirect(url_for("stocks"))
    
    if corrective_maintenances > 0:
        flash(f"Impossible de supprimer ce stock : il est utilisé dans {corrective_maintenances} maintenance(s) corrective(s).", "danger")
        return redirect(url_for("stocks"))
    
    # Supprimer tous les produits du stock (cascade)
    StockProduct.query.filter_by(stock_id=stock_id).delete()
    
    # Supprimer le stock
    db.session.delete(stock)
    
    try:
        db.session.commit()
        flash("Stock supprimé", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur lors de la suppression : {exc}", "danger")
    
    return redirect(url_for("stocks"))


@app.route("/movements", methods=["GET", "POST"])
@login_required
def movements():
    stocks = Stock.query.order_by(Stock.name).all()
    products = Product.query.order_by(Product.name).all()
    if request.method == "POST":
        # Seul l'admin ou le gestionnaire peut créer/modifier des mouvements
        if current_user.user_type not in ["admin", "gestionnaire"]:
            flash("Accès refusé : cette fonctionnalité est réservée aux administrateurs et gestionnaires.", "danger")
            return redirect(url_for("movements"))
        move_type = request.form["type"]
        timestamp = request.form.get("created_at")
        created_at = dt.datetime.fromisoformat(timestamp) if timestamp else dt.datetime.utcnow()
        source_id = request.form.get("source_stock_id") or None
        dest_id = request.form.get("dest_stock_id") or None

        movement = Movement(type=move_type, source_stock_id=source_id, dest_stock_id=dest_id, created_at=created_at)

        product_ids = request.form.getlist("product_id")
        quantities = request.form.getlist("quantity")

        if not product_ids or not any(pid for pid in product_ids):
            flash("Sélectionnez au moins un produit", "danger")
            return redirect(request.url)

        items = []
        for pid, qty in zip(product_ids, quantities):
            if not pid:
                continue
            try:
                qty_value = int(float(qty or 0.0))
            except (TypeError, ValueError):
                continue
            if qty_value <= 0:
                continue
            item = MovementItem(product_id=int(pid), quantity=qty_value)
            items.append(item)
            movement.items.append(item)

        if not items:
            flash("Aucune quantité valide fournie", "danger")
            return redirect(request.url)

        error = apply_movement_rules(movement)
        if error:
            flash(error, "danger")
            db.session.rollback()
            return redirect(request.url)

        db.session.add(movement)
        try:
            db.session.commit()
            # Message automatique pour le chat
            move_type_label = {"entree": "entrée", "sortie": "sortie", "transfert": "transfert"}.get(move_type, move_type)
            stock_names = []
            source_stock = Stock.query.get(source_id) if source_id else None
            dest_stock = Stock.query.get(dest_id) if dest_id else None
            if source_stock:
                stock_names.append(source_stock.name)
            if dest_stock:
                stock_names.append(dest_stock.name)
            stock_info = " → ".join(stock_names) if stock_names else ""
            # Récupérer la machine associée si le mouvement est lié à une maintenance
            machine_name = ""
            machine_id_for_msg = None
            # Chercher dans une fenêtre de 5 minutes pour trouver une maintenance associée
            time_window_start = created_at - dt.timedelta(minutes=5)
            time_window_end = created_at + dt.timedelta(minutes=5)
            
            # Vérifier maintenance préventive
            related_entry = MaintenanceEntry.query.filter(
                MaintenanceEntry.stock_id == source_id,
                MaintenanceEntry.created_at >= time_window_start,
                MaintenanceEntry.created_at <= time_window_end
            ).first()
            
            if related_entry:
                machine_name = f" sur la machine '{related_entry.machine.name}'"
                machine_id_for_msg = related_entry.machine_id
            else:
                # Vérifier maintenance corrective
                related_corrective = CorrectiveMaintenance.query.filter(
                    CorrectiveMaintenance.stock_id == source_id,
                    CorrectiveMaintenance.created_at >= time_window_start,
                    CorrectiveMaintenance.created_at <= time_window_end
                ).first()
                
                if related_corrective:
                    machine_name = f" sur la machine '{related_corrective.machine.name}'"
                    machine_id_for_msg = related_corrective.machine_id
            
            create_chat_message(
                message_type="auto",
                content=f"{current_user.username} a effectué un mouvement ({move_type_label})" + (f" : {stock_info}" if stock_info else "") + machine_name,
                link_url=url_for("movements"),
                machine_id=machine_id_for_msg
            )
            flash("Mouvement enregistré", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
        return redirect(request.url)

    recent_movements = Movement.query.order_by(Movement.created_at.desc()).limit(20).all()
    
    # Marquer les mouvements liés aux maintenances et récupérer les informations
    for move in recent_movements:
        move.is_maintenance_related = False
        move.maintenance_info = None
        # Vérifier si le mouvement est lié à une maintenance préventive ou corrective
        if move.type == "sortie" and move.source_stock_id:
            # Chercher dans une fenêtre de 5 minutes
            time_window_start = move.created_at - dt.timedelta(minutes=5)
            time_window_end = move.created_at + dt.timedelta(minutes=5)
            
            # Vérifier maintenance préventive
            related_entry = MaintenanceEntry.query.filter(
                MaintenanceEntry.stock_id == move.source_stock_id,
                MaintenanceEntry.created_at >= time_window_start,
                MaintenanceEntry.created_at <= time_window_end
            ).first()
            
            if related_entry:
                move.is_maintenance_related = True
                move.maintenance_info = {
                    'type': 'préventive',
                    'name': related_entry.report.name,
                    'machine_name': related_entry.machine.name,
                    'machine_code': related_entry.machine.code
                }
            else:
                # Vérifier maintenance corrective
                related_corrective = CorrectiveMaintenance.query.filter(
                    CorrectiveMaintenance.stock_id == move.source_stock_id,
                    CorrectiveMaintenance.created_at >= time_window_start,
                    CorrectiveMaintenance.created_at <= time_window_end
                ).first()
                
                if related_corrective:
                    move.is_maintenance_related = True
                    move.maintenance_info = {
                        'type': 'corrective',
                        'name': 'Maintenance corrective',
                        'machine_name': related_corrective.machine.name,
                        'machine_code': related_corrective.machine.code
                    }
    
    return render_template("movements.html", movements=recent_movements, stocks=stocks, products=products)


@app.route("/movements/export")
@login_required
def export_movements():
    movements = Movement.query.order_by(Movement.created_at.desc()).all()
    
    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Mouvements"
    
    # En-têtes
    headers = ["Date", "Type", "Stock source", "Stock destination", "Produits", "Quantités"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    for movement in movements:
        products_list = []
        quantities_list = []
        for item in movement.items:
            product_name = item.product.name if item.product else "Produit inconnu"
            product_code = item.product.code if item.product else "-"
            products_list.append(f"{product_name} ({product_code})")
            quantities_list.append(str(item.quantity))
        
        row = [
            movement.created_at.strftime("%d/%m/%Y %H:%M"),
            movement.type.capitalize(),
            movement.source_stock.name if movement.source_stock else "-",
            movement.dest_stock.name if movement.dest_stock else "-",
            ", ".join(products_list),
            ", ".join(quantities_list)
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=mouvements_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/movements/<int:movement_id>/edit", methods=["GET", "POST"])
@admin_or_manager_required
def edit_movement(movement_id):
    movement = Movement.query.get_or_404(movement_id)
    
    # Vérifier si le mouvement est lié à une maintenance
    is_maintenance_related = False
    if movement.type == "sortie" and movement.source_stock_id:
        time_window_start = movement.created_at - dt.timedelta(minutes=5)
        time_window_end = movement.created_at + dt.timedelta(minutes=5)
        related_entry = MaintenanceEntry.query.filter(
            MaintenanceEntry.stock_id == movement.source_stock_id,
            MaintenanceEntry.created_at >= time_window_start,
            MaintenanceEntry.created_at <= time_window_end
        ).first()
        if related_entry:
            is_maintenance_related = True
    
    if is_maintenance_related:
        flash("Ce mouvement est lié à une maintenance et ne peut pas être modifié", "danger")
        return redirect(url_for("movements"))
    
    stocks = Stock.query.order_by(Stock.name).all()
    products = Product.query.order_by(Product.name).all()
    
    if request.method == "POST":
        move_type = request.form["type"]
        timestamp = request.form.get("created_at")
        try:
            created_at = dt.datetime.fromisoformat(timestamp) if timestamp else dt.datetime.utcnow()
        except (ValueError, TypeError):
            created_at = dt.datetime.utcnow()
        source_id = request.form.get("source_stock_id") or None
        dest_id = request.form.get("dest_stock_id") or None
        
        # Sauvegarder les anciens items avant de les supprimer
        old_items = [(item.product_id, item.quantity) for item in movement.items]
        old_type = movement.type
        old_source_id = movement.source_stock_id
        old_dest_id = movement.dest_stock_id
        
        # Inverser l'ancien mouvement
        try:
            reverse_movement_rules(movement)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(request.url)
        
        # Mettre à jour le mouvement
        movement.type = move_type
        movement.source_stock_id = source_id
        movement.dest_stock_id = dest_id
        movement.created_at = created_at
        
        # Supprimer les anciens items
        for item in movement.items:
            db.session.delete(item)
        movement.items = []
        
        # Créer les nouveaux items
        product_ids = request.form.getlist("product_id")
        quantities = request.form.getlist("quantity")
        
        if not product_ids or not any(pid for pid in product_ids):
            flash("Sélectionnez au moins un produit", "danger")
            # Restaurer l'ancien mouvement
            movement.type = old_type
            movement.source_stock_id = old_source_id
            movement.dest_stock_id = old_dest_id
            for pid, qty in old_items:
                item = MovementItem(product_id=pid, quantity=qty)
                movement.items.append(item)
            apply_movement_rules(movement)
            return redirect(request.url)
        
        items = []
        for pid, qty in zip(product_ids, quantities):
            if not pid:
                continue
            try:
                qty_value = int(float(qty or 0.0))
            except (TypeError, ValueError):
                continue
            if qty_value <= 0:
                continue
            item = MovementItem(product_id=int(pid), quantity=qty_value)
            items.append(item)
            movement.items.append(item)
        
        if not items:
            flash("Aucune quantité valide fournie", "danger")
            # Restaurer l'ancien mouvement
            movement.type = old_type
            movement.source_stock_id = old_source_id
            movement.dest_stock_id = old_dest_id
            for pid, qty in old_items:
                item = MovementItem(product_id=pid, quantity=qty)
                movement.items.append(item)
            apply_movement_rules(movement)
            return redirect(request.url)
        
        # Appliquer le nouveau mouvement
        error = apply_movement_rules(movement)
        if error:
            flash(error, "danger")
            db.session.rollback()
            # Restaurer l'ancien mouvement
            movement.type = old_type
            movement.source_stock_id = old_source_id
            movement.dest_stock_id = old_dest_id
            for item in movement.items:
                db.session.delete(item)
            movement.items = []
            for pid, qty in old_items:
                item = MovementItem(product_id=pid, quantity=qty)
                movement.items.append(item)
            try:
                apply_movement_rules(movement)
                db.session.commit()
            except:
                db.session.rollback()
            return redirect(request.url)
        
        try:
            db.session.commit()
            flash("Mouvement modifié", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
        return redirect(url_for("movements"))
    
    # Rediriger vers la page movements avec les données nécessaires
    # Les données seront chargées via JavaScript dans la modal
    return redirect(url_for("movements"))


@app.route("/movements/<int:movement_id>/delete", methods=["POST"])
@admin_or_manager_required
def delete_movement(movement_id):
    movement = Movement.query.get_or_404(movement_id)
    
    # Vérifier si le mouvement est lié à une maintenance
    is_maintenance_related = False
    if movement.type == "sortie" and movement.source_stock_id:
        time_window_start = movement.created_at - dt.timedelta(minutes=5)
        time_window_end = movement.created_at + dt.timedelta(minutes=5)
        related_entry = MaintenanceEntry.query.filter(
            MaintenanceEntry.stock_id == movement.source_stock_id,
            MaintenanceEntry.created_at >= time_window_start,
            MaintenanceEntry.created_at <= time_window_end
        ).first()
        if related_entry:
            is_maintenance_related = True
    
    if is_maintenance_related:
        flash("Ce mouvement est lié à une maintenance et ne peut pas être supprimé", "danger")
        return redirect(url_for("movements"))
    
    # Inverser le mouvement pour remettre les stocks à leur état initial
    try:
        reverse_movement_rules(movement)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("movements"))
    
    # Supprimer le mouvement (les items seront supprimés en cascade)
    db.session.delete(movement)
    
    try:
        db.session.commit()
        flash("Mouvement supprimé", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur lors de la suppression : {exc}", "danger")
    
    return redirect(url_for("movements"))


@app.route("/maintenance/new", methods=["GET", "POST"])
@admin_required
def new_maintenance():
    machine_id_param = request.args.get("machine_id")
    if request.method == "POST":
        name = request.form["name"].strip()
        periodicity_raw = request.form.get("periodicity")
        machine_id_raw = request.form.get("machine_id")
        component_labels = request.form.getlist("component_label")
        component_types = request.form.getlist("component_type")

        try:
            periodicity = int(periodicity_raw)
        except (TypeError, ValueError):
            periodicity = 0

        try:
            machine_id = int(machine_id_raw)
        except (TypeError, ValueError):
            machine_id = None

        if not name or periodicity <= 0:
            flash("Nom et périodicité valide requis", "danger")
            return redirect(request.url)

        if not machine_id:
            flash("Veuillez sélectionner une machine", "danger")
            return redirect(request.url)

        machine = Machine.query.get(machine_id)
        if not machine:
            flash("Machine introuvable", "danger")
            return redirect(request.url)

        # Trouver la machine racine
        root_machine = machine
        while root_machine.parent:
            root_machine = root_machine.parent
        
        # Vérifier que la machine a un compteur OU que la machine racine a des compteurs
        has_own_counter = machine.hour_counter_enabled
        has_root_counters = root_machine.is_root() and root_machine.counters
        
        if not has_own_counter and not has_root_counters:
            flash("La machine doit avoir un compteur ou la machine racine doit avoir des compteurs pour créer un modèle de maintenance", "danger")
            return redirect(request.url)

        # Gérer le counter_id sélectionné
        counter_id = None
        counter_id_raw = request.form.get("counter_id")
        
        # Si la machine n'a pas de compteur, un compteur de la machine racine est obligatoire
        if not has_own_counter:
            if not counter_id_raw or counter_id_raw == "" or counter_id_raw == "machine":
                flash("Vous devez sélectionner un compteur de la machine racine car cette machine n'a pas de compteur", "danger")
                return redirect(request.url)
        
        if counter_id_raw and counter_id_raw != "" and counter_id_raw != "machine":
            try:
                counter_id = int(counter_id_raw)
                # Vérifier que le compteur existe et appartient à la machine racine
                counter = Counter.query.get(counter_id)
                if counter:
                    if counter.machine_id != root_machine.id:
                        flash("Compteur invalide", "danger")
                        return redirect(request.url)
                else:
                    flash("Compteur introuvable", "danger")
                    return redirect(request.url)
            except (TypeError, ValueError):
                flash("Valeur de compteur invalide", "danger")
                return redirect(request.url)
        # Si counter_id_raw est "machine" ou vide et que la machine a un compteur, counter_id reste None (utilise le compteur de la machine)

        components = []
        for label, field_type in zip(component_labels, component_types):
            label = label.strip()
            if not label or field_type not in {"number", "text", "checkbox"}:
                continue
            components.append(PreventiveComponent(label=label, field_type=field_type))

        if not components:
            flash("Ajoutez au moins un élément de rapport", "danger")
            return redirect(request.url)

        report = PreventiveReport(name=name, periodicity=periodicity, machine_id=machine_id, counter_id=counter_id)
        report.components.extend(components)
        db.session.add(report)
        try:
            db.session.commit()
            # Créer le MaintenanceProgress pour initialiser "avant maintenance"
            ensure_all_progress_for_machine(machine)
            db.session.commit()
            flash("Modèle enregistré", "success")
            if machine_id:
                return redirect(get_machine_detail_url(machine_id, 'preventive'))
            return redirect(url_for("new_maintenance"))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)

    # Montrer toutes les machines qui ont un compteur OU qui ont une machine racine avec des compteurs
    machines_with_counter = []
    all_machines = Machine.query.order_by(Machine.name).all()
    
    for machine in all_machines:
        # Vérifier si la machine a son propre compteur
        has_own_counter = machine.hour_counter_enabled
        
        # Vérifier si la machine racine a des compteurs
        root_machine = machine
        while root_machine.parent:
            root_machine = root_machine.parent
        has_root_counters = root_machine.is_root() and root_machine.counters
        
        # Inclure la machine si elle a un compteur OU si sa machine racine a des compteurs
        if has_own_counter or has_root_counters:
            machines_with_counter.append(machine)
    
    # Vérifier que le machine_id_param correspond à une machine valide
    selected_machine_id = None
    selected_machine = None
    if machine_id_param:
        try:
            param_id = int(machine_id_param)
            param_machine = Machine.query.get(param_id)
            if param_machine:
                # Vérifier si la machine a un compteur ou si sa machine racine a des compteurs
                has_own_counter = param_machine.hour_counter_enabled
                root_machine = param_machine
                while root_machine.parent:
                    root_machine = root_machine.parent
                has_root_counters = root_machine.is_root() and root_machine.counters
                
                if has_own_counter or has_root_counters:
                    selected_machine_id = param_id
                    selected_machine = param_machine
        except (TypeError, ValueError):
            pass
    
    # Récupérer tous les modèles existants pour la liste déroulante
    all_reports = PreventiveReport.query.order_by(PreventiveReport.name).all()
    
    return render_template("maintenance_form.html", machines=machines_with_counter, selected_machine_id=selected_machine_id, selected_machine=selected_machine, all_reports=all_reports)


@app.route("/maintenance/<int:report_id>")
@login_required
def maintenance_detail(report_id):
    report = PreventiveReport.query.get_or_404(report_id)
    return render_template("maintenance_detail.html", report=report)


@app.route("/maintenance/<int:report_id>/export-pdf")
@login_required
def export_maintenance_pdf(report_id):
    """Exporte un modèle de maintenance préventive en PDF"""
    report = PreventiveReport.query.get_or_404(report_id)
    machine = report.machine
    
    # Créer le PDF en mémoire
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#34495e'),
        spaceAfter=12,
        spaceBefore=12
    )
    normal_style = styles['Normal']
    
    # Contenu du PDF
    story = []
    
    # Titre
    story.append(Paragraph("Modèle de Maintenance Préventive", title_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Informations générales
    story.append(Paragraph("<b>Nom du modèle :</b>", heading_style))
    story.append(Paragraph(report.name, normal_style))
    story.append(Spacer(1, 0.3*cm))
    
    story.append(Paragraph("<b>Machine :</b>", heading_style))
    story.append(Paragraph(f"{machine.name} ({machine.code})", normal_style))
    story.append(Spacer(1, 0.3*cm))
    
    # Informations sur le compteur si disponible
    if report.counter:
        story.append(Paragraph("<b>Compteur de référence :</b>", heading_style))
        story.append(Paragraph(f"{report.counter.name} ({report.counter.unit or 'h'})", normal_style))
        story.append(Spacer(1, 0.3*cm))
    elif machine.hour_counter_enabled:
        story.append(Paragraph("<b>Compteur de référence :</b>", heading_style))
        story.append(Paragraph(f"Compteur machine ({machine.counter_unit or 'h'})", normal_style))
        story.append(Spacer(1, 0.3*cm))
    
    story.append(Paragraph("<b>Périodicité :</b>", heading_style))
    unit = "h"
    if report.counter:
        unit = report.counter.unit or "h"
    elif machine.hour_counter_enabled:
        unit = machine.counter_unit or "h"
    story.append(Paragraph(f"{report.periodicity} {unit}", normal_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Éléments du rapport
    story.append(Paragraph("<b>Éléments de la maintenance :</b>", heading_style))
    story.append(Spacer(1, 0.3*cm))
    
    if report.components:
        # Créer un tableau pour les éléments
        data = [['Élément', 'Type']]
        for component in report.components:
            type_label = {
                'number': 'Nombre',
                'text': 'Texte',
                'checkbox': 'Case à cocher'
            }.get(component.field_type, component.field_type)
            data.append([component.label, type_label])
        
        table = Table(data, colWidths=[12*cm, 4*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("Aucun élément défini dans ce modèle.", normal_style))
    
    story.append(Spacer(1, 0.5*cm))
    
    # Date de génération
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(f"<i>Généré le {dt.datetime.now().strftime('%d/%m/%Y à %H:%M')}</i>", 
                          ParagraphStyle('DateStyle', parent=normal_style, fontSize=9, textColor=colors.grey, alignment=TA_CENTER)))
    
    # Générer le PDF
    doc.build(story)
    buffer.seek(0)
    
    # Créer la réponse
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=maintenance_{report.id}_{report.name.replace(" ", "_")}_{dt.datetime.now().strftime("%Y%m%d")}.pdf'
    return response


@app.route("/maintenance/<int:report_id>/delete", methods=["POST"])
@admin_required
def delete_maintenance_report(report_id):
    """Supprimer un modèle de maintenance préventive"""
    report = PreventiveReport.query.get_or_404(report_id)
    machine_id = report.machine_id
    
    # Vérifier si le rapport a des maintenances associées
    entries_count = MaintenanceEntry.query.filter_by(report_id=report_id).count()
    if entries_count > 0:
        flash(f"Impossible de supprimer ce modèle : il est utilisé dans {entries_count} maintenance(s) enregistrée(s).", "danger")
        return redirect(url_for("maintenance_detail", report_id=report_id))
    
    # Supprimer les progress associés
    MaintenanceProgress.query.filter_by(report_id=report_id).delete()
    
    # Supprimer le rapport (les composants seront supprimés en cascade)
    db.session.delete(report)
    
    try:
        db.session.commit()
        flash("Modèle de maintenance supprimé", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur lors de la suppression : {exc}", "danger")
        return redirect(url_for("maintenance_detail", report_id=report_id))
    
    return redirect(url_for("machine_detail", machine_id=machine_id))


@app.route("/api/stock/<int:stock_id>/product/<int:product_id>/quantity")
@login_required
def get_product_stock_quantity(stock_id, product_id):
    """API pour récupérer la quantité d'un produit dans un stock"""
    stock = Stock.query.get_or_404(stock_id)
    product = Product.query.get_or_404(product_id)
    
    stock_product = StockProduct.query.filter_by(stock_id=stock_id, product_id=product_id).first()
    quantity = stock_product.quantity if stock_product else 0
    
    return json.dumps({"quantity": float(quantity)})


@app.route("/api/stock/<int:stock_id>/products")
@login_required
def get_stock_products(stock_id):
    """API pour récupérer tous les produits disponibles dans un stock"""
    stock = Stock.query.get_or_404(stock_id)
    
    # Récupérer tous les produits qui ont une quantité > 0 dans ce stock
    stock_products = StockProduct.query.filter_by(stock_id=stock_id).filter(StockProduct.quantity > 0).all()
    
    products_data = []
    for sp in stock_products:
        products_data.append({
            "id": sp.product.id,
            "name": sp.product.name,
            "code": sp.product.code,
            "quantity": float(sp.quantity)
        })
    
    # Trier par nom
    products_data.sort(key=lambda x: x["name"])
    
    return jsonify({"products": products_data})


@app.route("/api/maintenance-report/<int:report_id>/components")
@login_required
def get_report_components(report_id):
    """Récupérer les composants d'un modèle de maintenance"""
    report = PreventiveReport.query.get_or_404(report_id)
    components = []
    for component in report.components:
        components.append({
            "label": component.label,
            "field_type": component.field_type
        })
    return json.dumps({"success": True, "components": components})


@app.route("/api/checklist-template/<int:template_id>/items")
@login_required
def get_checklist_template_items(template_id):
    """API pour récupérer les items d'un modèle de checklist"""
    template = ChecklistTemplate.query.get_or_404(template_id)
    items_data = []
    for item in template.items:
        items_data.append({
            "label": item.label
        })
    return jsonify(success=True, items=items_data)


@app.route("/api/machine/<int:machine_id>/available-counters")
@login_required
def get_available_counters(machine_id):
    """API pour récupérer tous les compteurs disponibles pour une machine (son propre compteur + compteurs de la machine racine)"""
    machine = Machine.query.get_or_404(machine_id)
    
    available_counters = []
    
    # 1. Si la machine a son propre compteur (hour_counter_enabled)
    if machine.hour_counter_enabled:
        available_counters.append({
            "id": None,  # None signifie utiliser le compteur de la machine (on utilisera "machine" comme valeur)
            "name": f"Compteur de {machine.name}",
            "machine_name": machine.name,
            "value": machine.hours,
            "unit": machine.counter_unit or "h",
            "type": "machine"
        })
    
    # 2. Récupérer la machine racine
    root_machine = machine
    while root_machine.parent:
        root_machine = root_machine.parent
    
    # 3. Si la machine racine a des compteurs multiples (toujours les proposer, même si la machine a son propre compteur)
    if root_machine.is_root() and root_machine.counters:
        for counter in root_machine.counters:
            available_counters.append({
                "id": counter.id,
                "name": f"{counter.name} ({root_machine.name})",
                "machine_name": root_machine.name,
                "value": counter.value,
                "unit": counter.unit or "h",
                "type": "root_counter"
            })
    
    return json.dumps({"counters": available_counters})


@app.route("/machines/<int:machine_id>/maintenance/<int:report_id>/fill", methods=["GET", "POST"])
@admin_or_technician_required
def fill_maintenance(machine_id, report_id):
    machine = Machine.query.get_or_404(machine_id)
    report = PreventiveReport.query.filter_by(id=report_id, machine_id=machine_id).first_or_404()
    
    # Vérifier que la machine a un compteur horaire
    if not machine.hour_counter_enabled:
        flash("Seules les machines avec compteur horaire peuvent avoir des plans de maintenance préventive", "danger")
        return redirect(get_machine_detail_url(machine_id, 'preventive'))
    
    stocks = Stock.query.order_by(Stock.name).all()
    products = Product.query.order_by(Product.name).all()
    
    # Déterminer le stock par défaut : celui de la machine, ou le premier stock disponible
    default_stock_id = None
    if machine.stock_id:
        default_stock_id = machine.stock_id
    elif stocks:
        default_stock_id = stocks[0].id

    if not report.components:
        flash("Le modèle sélectionné ne contient aucun élément", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))

    if request.method == "POST":
        # Récupérer le nombre d'heures saisi, ou utiliser machine.hours par défaut
        performed_hours_raw = request.form.get("performed_hours")
        try:
            performed_hours = float(performed_hours_raw) if performed_hours_raw else machine.hours
        except (TypeError, ValueError):
            performed_hours = machine.hours
        
        entry = MaintenanceEntry(
            machine=machine,
            report=report,
            performed_hours=performed_hours,
            user_id=current_user.id
        )
        errors = []

        stock_id_raw = request.form.get("stock_id")
        try:
            stock_id = int(stock_id_raw)
        except (TypeError, ValueError):
            stock_id = None

        stock = Stock.query.get(stock_id) if stock_id else None
        if stock:
            entry.stock = stock

        product_ids = request.form.getlist("stock_product_id")
        quantities = request.form.getlist("stock_product_qty")
        removal_items = []
        for pid, qty in zip(product_ids, quantities):
            if not pid:
                continue
            try:
                pid_int = int(pid)
            except (TypeError, ValueError):
                errors.append("Produit invalide sélectionné pour le stock.")
                continue
            try:
                qty_value = int(float(qty or 0.0))
            except (TypeError, ValueError):
                errors.append("Quantité invalide pour l'un des produits.")
                continue
            if qty_value <= 0:
                continue
            removal_items.append((pid_int, qty_value))

        for component in report.components:
            field_name = f"component_{component.id}"
            raw_value = request.form.get(field_name)
            if component.field_type == "checkbox":
                value_bool = request.form.get(field_name) == "on"
                entry.values.append(
                    MaintenanceEntryValue(
                        component=component,
                        value_bool=value_bool,
                    )
                )
            elif component.field_type == "number":
                try:
                    number_value = float(raw_value)
                except (TypeError, ValueError):
                    errors.append(f"Valeur numérique invalide pour {component.label}")
                    continue
                entry.values.append(
                    MaintenanceEntryValue(
                        component=component,
                        value_number=number_value,
                    )
                )
            else:
                if not raw_value:
                    errors.append(f"Veuillez renseigner {component.label}")
                    continue
                entry.values.append(
                    MaintenanceEntryValue(
                        component=component,
                        value_text=raw_value.strip(),
                    )
                )

        if errors:
            for msg in errors:
                flash(msg, "danger")
            return redirect(request.url)

        # Créer un mouvement de sortie uniquement si des produits sont fournis
        if removal_items and stock:
            movement = Movement(type="sortie", source_stock_id=stock.id, created_at=dt.datetime.utcnow())
            for pid, qty in removal_items:
                movement.items.append(MovementItem(product_id=pid, quantity=qty))

            movement_error = apply_movement_rules(movement)
            if movement_error:
                flash(movement_error, "danger")
                db.session.rollback()
                return redirect(request.url)
            db.session.add(movement)

        progress_record = None
        if machine.hour_counter_enabled:
            progress_record = get_or_create_progress(machine, report)
            # Stocker l'heure avant maintenance avant de la réinitialiser
            entry.hours_before_maintenance = progress_record.hours_since
            progress_record.hours_since = report.periodicity

        db.session.add(entry)
        try:
            db.session.commit()
            
            # Traiter les photos uploadées
            if 'photos' in request.files:
                photos = request.files.getlist('photos')
                for photo in photos:
                    if photo and photo.filename and allowed_image_file(photo.filename):
                        original_filename = photo.filename
                        filename = secure_filename(original_filename)
                        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S_")
                        safe_filename = f"{timestamp}{entry.id}_{filename}"
                        file_path = MAINTENANCE_PHOTOS_FOLDER / safe_filename
                        photo.save(str(file_path))
                        
                        maintenance_photo = MaintenancePhoto(
                            maintenance_entry_id=entry.id,
                            filename=safe_filename,
                            original_filename=original_filename,
                            file_path=str(file_path),
                            user_id=current_user.id
                        )
                        db.session.add(maintenance_photo)
            
            db.session.commit()
            # Message automatique pour le chat
            create_chat_message(
                message_type="auto",
                content=f"{current_user.username} a finalisé la maintenance préventive '{entry.report.name}' sur la machine '{entry.machine.name}'",
                link_url=url_for("maintenance_entry_detail", entry_id=entry.id),
                machine_id=entry.machine_id
            )
            flash("Rapport de maintenance enregistré", "success")
            return redirect(url_for("maintenance_entry_detail", entry_id=entry.id))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)

    return render_template(
        "maintenance_fill.html",
        machine=machine,
        report=report,
        stocks=stocks,
        products=products,
        default_stock_id=default_stock_id,
    )


@app.route("/maintenance-entry/<int:entry_id>")
@login_required
def maintenance_entry_detail(entry_id):
    entry = MaintenanceEntry.query.get_or_404(entry_id)
    
    # Récupérer le mouvement de sortie associé à cette maintenance (créé à peu près en même temps)
    movement = None
    if entry.stock:
        time_window_start = entry.created_at - dt.timedelta(minutes=5)
        time_window_end = entry.created_at + dt.timedelta(minutes=5)
        movements = Movement.query.filter(
            Movement.type == "sortie",
            Movement.source_stock_id == entry.stock.id,
            Movement.created_at >= time_window_start,
            Movement.created_at <= time_window_end
        ).all()
        # Prendre le premier mouvement trouvé (normalement il n'y en a qu'un)
        if movements:
            movement = movements[0]
    
    # Récupérer les photos
    photos = MaintenancePhoto.query.filter_by(maintenance_entry_id=entry_id).order_by(MaintenancePhoto.uploaded_at).all()
    
    return render_template("maintenance_entry_detail.html", entry=entry, movement=movement, photos=photos)


@app.route("/maintenance-entry/<int:entry_id>/delete", methods=["POST"])
@admin_required
def delete_maintenance_entry(entry_id):
    entry = MaintenanceEntry.query.get_or_404(entry_id)
    machine_id = entry.machine.id
    
    # Récupérer le mouvement de sortie associé à cette maintenance (créé à peu près en même temps)
    movement = None
    if entry.stock:
        time_window_start = entry.created_at - dt.timedelta(minutes=5)
        time_window_end = entry.created_at + dt.timedelta(minutes=5)
        movements = Movement.query.filter(
            Movement.type == "sortie",
            Movement.source_stock_id == entry.stock.id,
            Movement.created_at >= time_window_start,
            Movement.created_at <= time_window_end
        ).all()
        # Prendre le premier mouvement trouvé (normalement il n'y en a qu'un)
        if movements:
            movement = movements[0]
    
    # Si un mouvement est associé, inverser ses effets pour remettre les produits en stock
    if movement:
        try:
            reverse_movement_rules(movement)
            # Supprimer le mouvement
            db.session.delete(movement)
        except ValueError as exc:
            flash(f"Erreur lors de la restauration des stocks : {exc}", "danger")
            db.session.rollback()
            return redirect(url_for("maintenance_entry_detail", entry_id=entry.id))
    
    # Supprimer l'entrée de maintenance (les valeurs seront supprimées en cascade)
    db.session.delete(entry)
    
    try:
        db.session.commit()
        flash("Maintenance supprimée avec succès", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur lors de la suppression : {exc}", "danger")
        return redirect(url_for("maintenance_entry_detail", entry_id=entry.id))
    
    return redirect(url_for("machine_detail", machine_id=machine_id))


@app.route("/maintenance-entry/<int:entry_id>/edit", methods=["GET", "POST"])
@login_required
def edit_maintenance_entry(entry_id):
    entry = MaintenanceEntry.query.get_or_404(entry_id)
    
    # Vérifier les permissions : admin ou technicien qui a créé le rapport
    if not can_edit_maintenance_entry(entry):
        flash("Accès refusé : vous ne pouvez modifier que les rapports que vous avez créés.", "danger")
        return redirect(url_for("maintenance_entry_detail", entry_id=entry.id))
    
    machine = entry.machine
    report = entry.report
    
    # Vérifier que la machine a un compteur horaire
    if not machine.hour_counter_enabled:
        flash("Seules les machines avec compteur horaire peuvent avoir des plans de maintenance préventive", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))
    
    stocks = Stock.query.order_by(Stock.name).all()
    products = Product.query.order_by(Product.name).all()

    if not report.components:
        flash("Le modèle sélectionné ne contient aucun élément", "danger")
        return redirect(url_for("machine_detail", machine_id=machine.id))

    if request.method == "POST":
        errors = []

        # Mettre à jour le stock
        stock_id_raw = request.form.get("stock_id")
        try:
            stock_id = int(stock_id_raw)
        except (TypeError, ValueError):
            stock_id = None

        old_stock = entry.stock
        new_stock = Stock.query.get(stock_id) if stock_id else None
        entry.stock = new_stock

        # Récupérer les anciens produits pour annuler le mouvement si nécessaire
        old_product_quantities = {}
        if old_stock:
            # Chercher le mouvement de sortie associé (créé à peu près en même temps que le rapport)
            # On cherche dans une fenêtre de 5 minutes
            time_window_start = entry.created_at - dt.timedelta(minutes=5)
            time_window_end = entry.created_at + dt.timedelta(minutes=5)
            old_movements = Movement.query.filter(
                Movement.type == "sortie",
                Movement.source_stock_id == old_stock.id,
                Movement.created_at >= time_window_start,
                Movement.created_at <= time_window_end
            ).all()
            
            # Pour chaque mouvement, vérifier s'il correspond au rapport
            # (on ne peut pas être sûr à 100% sans lien direct, mais on fait de notre mieux)
            for mov in old_movements:
                for item in mov.items:
                    old_product_quantities[item.product_id] = old_product_quantities.get(item.product_id, 0) + item.quantity

        # Annuler l'ancien mouvement si nécessaire (remettre les produits en stock)
        if old_product_quantities and old_stock:
            for product_id, qty in old_product_quantities.items():
                stock_product = _get_or_create_stock_product(old_stock.id, product_id)
                stock_product.quantity += qty

        # Supprimer les anciennes valeurs
        for value in entry.values:
            db.session.delete(value)
        entry.values = []

        # Créer les nouvelles valeurs
        product_ids = request.form.getlist("stock_product_id")
        quantities = request.form.getlist("stock_product_qty")
        removal_items = []
        for pid, qty in zip(product_ids, quantities):
            if not pid:
                continue
            try:
                pid_int = int(pid)
            except (TypeError, ValueError):
                errors.append("Produit invalide sélectionné pour le stock.")
                continue
            try:
                qty_value = int(float(qty or 0.0))
            except (TypeError, ValueError):
                errors.append("Quantité invalide pour l'un des produits.")
                continue
            if qty_value <= 0:
                continue
            removal_items.append((pid_int, qty_value))

        for component in report.components:
            field_name = f"component_{component.id}"
            raw_value = request.form.get(field_name)
            if component.field_type == "checkbox":
                value_bool = request.form.get(field_name) == "on"
                entry.values.append(
                    MaintenanceEntryValue(
                        component=component,
                        value_bool=value_bool,
                    )
                )
            elif component.field_type == "number":
                try:
                    number_value = float(raw_value)
                except (TypeError, ValueError):
                    errors.append(f"Valeur numérique invalide pour {component.label}")
                    continue
                entry.values.append(
                    MaintenanceEntryValue(
                        component=component,
                        value_number=number_value,
                    )
                )
            else:
                if not raw_value:
                    errors.append(f"Veuillez renseigner {component.label}")
                    continue
                entry.values.append(
                    MaintenanceEntryValue(
                        component=component,
                        value_text=raw_value.strip(),
                    )
                )

        if errors:
            for msg in errors:
                flash(msg, "danger")
            return redirect(request.url)

        # Créer un nouveau mouvement de sortie si des produits sont fournis
        if removal_items and new_stock:
            movement = Movement(type="sortie", source_stock_id=new_stock.id, created_at=dt.datetime.utcnow())
            for pid, qty in removal_items:
                movement.items.append(MovementItem(product_id=pid, quantity=qty))

            movement_error = apply_movement_rules(movement)
            if movement_error:
                flash(movement_error, "danger")
                db.session.rollback()
                return redirect(request.url)
            db.session.add(movement)

        # Mettre à jour les heures effectuées
        performed_hours_raw = request.form.get("performed_hours")
        try:
            performed_hours = float(performed_hours_raw) if performed_hours_raw else machine.hours
        except (TypeError, ValueError):
            performed_hours = machine.hours
        entry.performed_hours = performed_hours

        try:
            db.session.commit()
            
            # Traiter les photos uploadées
            if 'photos' in request.files:
                photos = request.files.getlist('photos')
                for photo in photos:
                    if photo and photo.filename and allowed_image_file(photo.filename):
                        original_filename = photo.filename
                        filename = secure_filename(original_filename)
                        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S_")
                        safe_filename = f"{timestamp}{entry.id}_{filename}"
                        file_path = MAINTENANCE_PHOTOS_FOLDER / safe_filename
                        photo.save(str(file_path))
                        
                        maintenance_photo = MaintenancePhoto(
                            maintenance_entry_id=entry.id,
                            filename=safe_filename,
                            original_filename=original_filename,
                            file_path=str(file_path),
                            user_id=current_user.id
                        )
                        db.session.add(maintenance_photo)
            
            db.session.commit()
            flash("Rapport de maintenance modifié", "success")
            return redirect(url_for("maintenance_entry_detail", entry_id=entry.id))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)

    # Pré-remplir le formulaire avec les valeurs existantes
    existing_values = {value.component_id: value for value in entry.values}
    
    # Déterminer le stock par défaut : celui de la machine, ou le premier stock disponible
    default_stock_id = None
    if machine.stock_id:
        default_stock_id = machine.stock_id
    elif stocks:
        default_stock_id = stocks[0].id
    
    return render_template(
        "maintenance_fill.html",
        machine=machine,
        report=report,
        stocks=stocks,
        products=products,
        entry=entry,
        existing_values=existing_values,
        is_edit=True,
        default_stock_id=default_stock_id,
    )


@app.route("/machines/<int:machine_id>/corrective/new", methods=["GET", "POST"])
@admin_or_technician_required
def new_corrective_maintenance(machine_id):
    machine = Machine.query.get_or_404(machine_id)
    stocks = Stock.query.order_by(Stock.name).all()
    products = Product.query.order_by(Product.name).all()
    
    # Déterminer le stock par défaut : celui de la machine, ou le premier stock disponible
    default_stock_id = None
    if machine.stock_id:
        default_stock_id = machine.stock_id
    elif stocks:
        default_stock_id = stocks[0].id

    if request.method == "POST":
        comment = request.form.get("comment", "").strip()
        stock_id_raw = request.form.get("stock_id")
        created_at_str = request.form.get("created_at")
        hours_raw = request.form.get("hours")
        
        try:
            created_at = dt.datetime.fromisoformat(created_at_str) if created_at_str else dt.datetime.utcnow()
        except (ValueError, TypeError):
            created_at = dt.datetime.utcnow()

        try:
            stock_id = int(stock_id_raw) if stock_id_raw else None
        except (TypeError, ValueError):
            stock_id = None

        try:
            hours = float(hours_raw) if hours_raw else 0.0
        except (TypeError, ValueError):
            hours = 0.0

        stock = Stock.query.get(stock_id) if stock_id else None

        maintenance = CorrectiveMaintenance(
            machine=machine,
            stock=stock,
            comment=comment,
            hours=hours,
            created_at=created_at,
            user_id=current_user.id,
        )

        product_ids = request.form.getlist("product_id")
        quantities = request.form.getlist("quantity")
        removal_items = []
        for pid, qty in zip(product_ids, quantities):
            if not pid:
                continue
            try:
                pid_int = int(pid)
            except (TypeError, ValueError):
                continue
            try:
                qty_value = int(float(qty or 0.0))
            except (TypeError, ValueError):
                continue
            if qty_value <= 0:
                continue
            removal_items.append((pid_int, qty_value))

        for pid, qty in removal_items:
            maintenance.products.append(
                CorrectiveMaintenanceProduct(product_id=pid, quantity=qty)
            )

        if removal_items and stock:
            movement = Movement(type="sortie", source_stock_id=stock.id, created_at=created_at)
            for pid, qty in removal_items:
                movement.items.append(MovementItem(product_id=pid, quantity=qty))

            movement_error = apply_movement_rules(movement)
            if movement_error:
                flash(movement_error, "danger")
                db.session.rollback()
                return redirect(request.url)
            db.session.add(movement)

        db.session.add(maintenance)
        try:
            db.session.commit()
            # Message automatique pour le chat
            create_chat_message(
                message_type="auto",
                content=f"{current_user.username} a créé une maintenance corrective sur la machine '{machine.name}'",
                link_url=url_for("corrective_maintenance_detail", maintenance_id=maintenance.id),
                machine_id=machine.id
            )
            flash("Maintenance corrective enregistrée", "success")
            return redirect(get_machine_detail_url(machine.id, 'corrective'))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)

    return render_template(
        "corrective_maintenance_form.html",
        machine=machine,
        stocks=stocks,
        products=products,
        maintenance=None,
        existing_products=None,
        is_edit=False,
        default_stock_id=default_stock_id,
    )


@app.route("/corrective-maintenance/<int:maintenance_id>")
@login_required
def corrective_maintenance_detail(maintenance_id):
    maintenance = CorrectiveMaintenance.query.get_or_404(maintenance_id)
    # Récupérer les photos
    photos = MaintenancePhoto.query.filter_by(corrective_maintenance_id=maintenance_id).order_by(MaintenancePhoto.uploaded_at).all()
    return render_template("corrective_maintenance_detail.html", maintenance=maintenance, photos=photos)


@app.route("/corrective-maintenance/<int:maintenance_id>/edit", methods=["GET", "POST"])
@login_required
def edit_corrective_maintenance(maintenance_id):
    maintenance = CorrectiveMaintenance.query.get_or_404(maintenance_id)
    
    # Vérifier les permissions : admin ou technicien qui a créé le rapport
    if current_user.user_type != "admin" and (current_user.user_type != "technicien" or maintenance.user_id != current_user.id):
        flash("Accès refusé : vous ne pouvez modifier que les rapports que vous avez créés.", "danger")
        return redirect(url_for("corrective_maintenance_detail", maintenance_id=maintenance.id))
    machine = maintenance.machine
    stocks = Stock.query.order_by(Stock.name).all()
    products = Product.query.order_by(Product.name).all()

    if request.method == "POST":
        comment = request.form.get("comment", "").strip()
        stock_id_raw = request.form.get("stock_id")
        created_at_str = request.form.get("created_at")
        hours_raw = request.form.get("hours")
        
        try:
            created_at = dt.datetime.fromisoformat(created_at_str) if created_at_str else dt.datetime.utcnow()
        except (ValueError, TypeError):
            created_at = dt.datetime.utcnow()

        try:
            stock_id = int(stock_id_raw) if stock_id_raw else None
        except (TypeError, ValueError):
            stock_id = None

        try:
            hours = float(hours_raw) if hours_raw else 0.0
        except (TypeError, ValueError):
            hours = 0.0

        old_stock = maintenance.stock
        old_created_at = maintenance.created_at
        new_stock = Stock.query.get(stock_id) if stock_id else None
        maintenance.stock = new_stock
        maintenance.comment = comment
        maintenance.hours = hours
        maintenance.created_at = created_at

        # Récupérer les anciens produits pour annuler le mouvement si nécessaire
        old_product_quantities = {}
        if old_stock:
            # Chercher le mouvement de sortie associé (créé à peu près en même temps que le rapport)
            # On cherche dans une fenêtre de 5 minutes en utilisant l'ancienne date
            time_window_start = old_created_at - dt.timedelta(minutes=5)
            time_window_end = old_created_at + dt.timedelta(minutes=5)
            old_movements = Movement.query.filter(
                Movement.type == "sortie",
                Movement.source_stock_id == old_stock.id,
                Movement.created_at >= time_window_start,
                Movement.created_at <= time_window_end
            ).all()
            
            # Pour chaque mouvement, vérifier s'il correspond au rapport
            # (on ne peut pas être sûr à 100% sans lien direct, mais on fait de notre mieux)
            for mov in old_movements:
                for item in mov.items:
                    old_product_quantities[item.product_id] = old_product_quantities.get(item.product_id, 0) + item.quantity

        # Annuler l'ancien mouvement si nécessaire (remettre les produits en stock)
        if old_product_quantities and old_stock:
            for product_id, qty in old_product_quantities.items():
                stock_product = _get_or_create_stock_product(old_stock.id, product_id)
                stock_product.quantity += qty

        # Supprimer les anciens produits
        for product_item in maintenance.products:
            db.session.delete(product_item)
        maintenance.products = []

        # Récupérer les nouveaux produits
        product_ids = request.form.getlist("product_id")
        quantities = request.form.getlist("quantity")
        removal_items = []
        for pid, qty in zip(product_ids, quantities):
            if not pid:
                continue
            try:
                pid_int = int(pid)
            except (TypeError, ValueError):
                continue
            try:
                qty_value = int(float(qty or 0.0))
            except (TypeError, ValueError):
                continue
            if qty_value <= 0:
                continue
            removal_items.append((pid_int, qty_value))

        # Ajouter les nouveaux produits
        for pid, qty in removal_items:
            maintenance.products.append(
                CorrectiveMaintenanceProduct(product_id=pid, quantity=qty)
            )

        # Créer un nouveau mouvement de sortie si des produits sont fournis
        if removal_items and new_stock:
            movement = Movement(type="sortie", source_stock_id=new_stock.id, created_at=created_at)
            for pid, qty in removal_items:
                movement.items.append(MovementItem(product_id=pid, quantity=qty))

            movement_error = apply_movement_rules(movement)
            if movement_error:
                flash(movement_error, "danger")
                db.session.rollback()
                return redirect(request.url)
            db.session.add(movement)

        try:
            db.session.commit()
            
            # Traiter les photos uploadées
            if 'photos' in request.files:
                photos = request.files.getlist('photos')
                for photo in photos:
                    if photo and photo.filename and allowed_image_file(photo.filename):
                        original_filename = photo.filename
                        filename = secure_filename(original_filename)
                        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S_")
                        safe_filename = f"{timestamp}{maintenance.id}_{filename}"
                        file_path = MAINTENANCE_PHOTOS_FOLDER / safe_filename
                        photo.save(str(file_path))
                        
                        maintenance_photo = MaintenancePhoto(
                            corrective_maintenance_id=maintenance.id,
                            filename=safe_filename,
                            original_filename=original_filename,
                            file_path=str(file_path),
                            user_id=current_user.id
                        )
                        db.session.add(maintenance_photo)
            
            db.session.commit()
            flash("Maintenance corrective modifiée", "success")
            return redirect(url_for("corrective_maintenance_detail", maintenance_id=maintenance.id))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
            return redirect(request.url)

    # Pré-remplir le formulaire avec les valeurs existantes
    existing_products = [(item.product_id, item.quantity) for item in maintenance.products]
    
    # Déterminer le stock par défaut : celui de la machine, ou le premier stock disponible
    default_stock_id = None
    if machine.stock_id:
        default_stock_id = machine.stock_id
    elif stocks:
        default_stock_id = stocks[0].id
    
    return render_template(
        "corrective_maintenance_form.html",
        machine=machine,
        stocks=stocks,
        products=products,
        maintenance=maintenance,
        existing_products=existing_products,
        is_edit=True,
        default_stock_id=default_stock_id,
    )


@app.route("/maintenances")
@login_required
def maintenances_list():
    # Récupérer les paramètres de filtrage
    filter_type = request.args.get('filter_type', '').strip().lower()
    filter_name = request.args.get('filter_name', '').strip().lower()
    filter_date = request.args.get('filter_date', '').strip()
    filter_machine = request.args.get('filter_machine', '').strip().lower()
    filter_user = request.args.get('filter_user', '').strip().lower()
    
    # Récupérer toutes les maintenances préventives
    preventive_entries = MaintenanceEntry.query.order_by(MaintenanceEntry.created_at.desc()).all()
    
    # Récupérer toutes les maintenances correctives
    corrective_maintenances = CorrectiveMaintenance.query.order_by(CorrectiveMaintenance.created_at.desc()).all()
    
    # Créer une liste unifiée avec type, nom, date, machine
    all_maintenances = []
    
    for entry in preventive_entries:
        all_maintenances.append({
            'type': 'préventive',
            'type_badge': 'primary',
            'name': entry.report.name,
            'date': entry.created_at,
            'machine': entry.machine,
            'id': entry.id,
            'user': entry.user,
            'url': url_for('maintenance_entry_detail', entry_id=entry.id)
        })
    
    for maintenance in corrective_maintenances:
        all_maintenances.append({
            'type': 'corrective',
            'type_badge': 'warning',
            'name': 'Maintenance corrective',
            'date': maintenance.created_at,
            'machine': maintenance.machine,
            'id': maintenance.id,
            'user': maintenance.user,
            'url': url_for('corrective_maintenance_detail', maintenance_id=maintenance.id)
        })
    
    # Appliquer les filtres
    filtered_maintenances = all_maintenances
    
    if filter_type:
        filtered_maintenances = [m for m in filtered_maintenances if filter_type in m['type'].lower()]
    
    if filter_name:
        filtered_maintenances = [m for m in filtered_maintenances if filter_name in m['name'].lower()]
    
    if filter_date:
        try:
            # Essayer de parser la date (format attendu: YYYY-MM-DD ou DD/MM/YYYY)
            filter_date_obj = None
            if '/' in filter_date:
                # Format DD/MM/YYYY
                parts = filter_date.split('/')
                if len(parts) == 3:
                    filter_date_obj = dt.datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            else:
                # Format YYYY-MM-DD
                filter_date_obj = dt.datetime.strptime(filter_date, '%Y-%m-%d')
            
            if filter_date_obj:
                filtered_maintenances = [
                    m for m in filtered_maintenances
                    if m['date'].date() == filter_date_obj.date()
                ]
        except (ValueError, AttributeError):
            pass
    
    if filter_machine:
        # Rechercher dans toute l'arborescence de la machine
        filtered_maintenances = [
            m for m in filtered_maintenances
            if any(
                filter_machine in node.name.lower() or filter_machine in (node.code or '').lower()
                for node in machine_lineage(m['machine'])
            )
        ]
    
    if filter_user:
        filtered_maintenances = [
            m for m in filtered_maintenances
            if m['user'] and filter_user in (m['user'].username or '').lower()
        ]
    
    # Trier par date décroissante
    filtered_maintenances.sort(key=lambda x: x['date'], reverse=True)
    
    return render_template(
        "maintenances_list.html",
        maintenances=filtered_maintenances,
        filter_type=filter_type,
        filter_name=filter_name,
        filter_date=filter_date,
        filter_machine=filter_machine,
        filter_user=filter_user
    )


@app.route("/maintenances/export")
@login_required
def export_maintenances():
    # Récupérer les mêmes filtres que la page maintenances
    filter_type = request.args.get('filter_type', '').strip().lower()
    filter_name = request.args.get('filter_name', '').strip().lower()
    filter_date = request.args.get('filter_date', '').strip()
    filter_machine = request.args.get('filter_machine', '').strip().lower()
    filter_user = request.args.get('filter_user', '').strip().lower()
    
    # Récupérer toutes les maintenances préventives
    preventive_entries = MaintenanceEntry.query.order_by(MaintenanceEntry.created_at.desc()).all()
    
    # Récupérer toutes les maintenances correctives
    corrective_maintenances = CorrectiveMaintenance.query.order_by(CorrectiveMaintenance.created_at.desc()).all()
    
    # Créer une liste unifiée avec type, nom, date, machine
    all_maintenances = []
    
    for entry in preventive_entries:
        all_maintenances.append({
            'type': 'Préventive',
            'name': entry.report.name,
            'date': entry.created_at,
            'machine': entry.machine,
            'user': entry.user.username if entry.user else 'Non renseigné'
        })
    
    for maintenance in corrective_maintenances:
        all_maintenances.append({
            'type': 'Corrective',
            'name': 'Maintenance corrective',
            'date': maintenance.created_at,
            'machine': maintenance.machine,
            'user': maintenance.user.username if maintenance.user else 'Non renseigné'
        })
    
    # Appliquer les mêmes filtres que la page maintenances
    filtered_maintenances = all_maintenances
    
    if filter_type:
        filtered_maintenances = [m for m in filtered_maintenances if filter_type in m['type'].lower()]
    
    if filter_name:
        filtered_maintenances = [m for m in filtered_maintenances if filter_name in m['name'].lower()]
    
    if filter_date:
        try:
            filter_date_obj = None
            if '/' in filter_date:
                parts = filter_date.split('/')
                if len(parts) == 3:
                    filter_date_obj = dt.datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            else:
                filter_date_obj = dt.datetime.strptime(filter_date, '%Y-%m-%d')
            
            if filter_date_obj:
                filtered_maintenances = [
                    m for m in filtered_maintenances
                    if m['date'].date() == filter_date_obj.date()
                ]
        except (ValueError, AttributeError):
            pass
    
    if filter_machine:
        filtered_maintenances = [
            m for m in filtered_maintenances
            if any(
                filter_machine in node.name.lower() or filter_machine in (node.code or '').lower()
                for node in machine_lineage(m['machine'])
            )
        ]
    
    if filter_user:
        filtered_maintenances = [
            m for m in filtered_maintenances
            if filter_user in m['user'].lower()
        ]
    
    # Trier par date décroissante
    filtered_maintenances.sort(key=lambda x: x['date'], reverse=True)
    
    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenances"
    
    # En-têtes
    headers = ["Type", "Nom", "Date", "Machine / Sous-machine", "Identifiant"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    for maintenance in filtered_maintenances:
        # Construire le chemin de la machine
        lineage = machine_lineage(maintenance['machine'])
        machine_path = ' › '.join([node.name for node in lineage])
        
        row = [
            maintenance['type'],
            maintenance['name'],
            maintenance['date'].strftime("%d/%m/%Y %H:%M"),
            machine_path,
            maintenance['user']
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=maintenances_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/maintenance/manage")
@login_required
def maintenance_manage():
    try:
        selected = int(request.args.get("threshold", 10))
    except (TypeError, ValueError):
        selected = 10
    if selected not in {5, 10, 15, 20}:
        selected = 10
    threshold_ratio = selected / 100.0

    overdue = []
    warning = []

    progress_records = (
        MaintenanceProgress.query.join(MaintenanceProgress.machine).join(MaintenanceProgress.report).all()
    )

    last_entry_rows = (
        db.session.query(
            MaintenanceEntry.machine_id.label("machine_id"),
            MaintenanceEntry.report_id.label("report_id"),
            func.max(MaintenanceEntry.created_at).label("last_date"),
        )
        .group_by(MaintenanceEntry.machine_id, MaintenanceEntry.report_id)
        .all()
    )
    last_map = {(row.machine_id, row.report_id): row.last_date for row in last_entry_rows}

    for record in progress_records:
        machine = record.machine
        report = record.report
        if not machine or not machine.hour_counter_enabled:
            continue
        remaining = record.hours_since
        last_performed = last_map.get((machine.id, report.id))
        if remaining <= 0:
            overdue.append(
                {
                    "machine": machine,
                    "report": report,
                    "remaining": remaining,
                    "last_performed": last_performed,
                }
            )
        elif remaining <= report.periodicity * threshold_ratio:
            warning.append(
                {
                    "machine": machine,
                    "report": report,
                    "remaining": remaining,
                    "last_performed": last_performed,
                }
            )

    overdue.sort(key=lambda item: item["remaining"])
    warning.sort(key=lambda item: item["remaining"])

    return render_template(
        "maintenance_manage.html",
        threshold=selected,
        threshold_options=[5, 10, 15, 20],
        overdue=overdue,
        warning=warning,
    )


@app.route("/machines/counter-report", methods=["GET", "POST"])
@app.route("/machines/counter-report/<int:machine_id>", methods=["GET", "POST"])
@admin_or_technician_required
def counter_report(machine_id=None):
    root_machine = None
    machines_with_counters = []  # Liste des machines avec compteur classique
    root_counters = []  # Liste des compteurs de la machine racine (pour compatibilité)
    root_counters_by_machine = []  # Liste de tuples (machine_racine, compteur) pour gérer plusieurs machines racines
    
    # Structure hiérarchique pour l'affichage
    counter_hierarchy = []
    
    # Si machine_id est fourni, récupérer cette machine racine et construire sa hiérarchie
    if machine_id:
        root_machine = Machine.query.get_or_404(machine_id)
        if root_machine.parent_id is not None:
            flash("Cette route est réservée aux machines racines.", "danger")
            return redirect(url_for("machines"))
        
        # Construire la hiérarchie des compteurs
        counter_hierarchy = build_counter_hierarchy(root_machine, depth=0)
        
        if not counter_hierarchy:
            flash(f"Aucun compteur configuré dans l'arborescence de {root_machine.name}.", "warning")
            return redirect(url_for("machines"))
    else:
        # Comportement par défaut : toutes les machines racines avec leurs hiérarchies
        all_root_machines = Machine.query.filter_by(parent_id=None).order_by(Machine.name).all()
        for root in all_root_machines:
            root_hierarchy = build_counter_hierarchy(root, depth=0)
            counter_hierarchy.extend(root_hierarchy)
        
        if not counter_hierarchy:
            flash("Aucun compteur configuré.", "warning")
            return redirect(url_for("machines"))
        
        # Pour compatibilité avec le template (première machine racine trouvée)
        root_machine = all_root_machines[0] if all_root_machines else None

    if request.method == "POST":
        updated = 0
        machines_updated = []  # Liste des machines mises à jour pour le message
        
        # Traiter les compteurs depuis la hiérarchie
        for item in counter_hierarchy:
            if item['type'] == 'machine_single_counter':
                # Compteur classique pour sous-machine
                machine = item['machine']
                raw_value = request.form.get(f"machine_{machine.id}")
                if raw_value is None or raw_value.strip() == "":
                    continue
                try:
                    new_hours = float(raw_value)
                except ValueError:
                    flash(f"Valeur invalide pour {machine.name}", "danger")
                    return redirect(request.url)
                old_hours = machine.hours
                if new_hours < old_hours:
                    unit = machine.counter_unit or 'h'
                    flash(f"Le nouveau compteur pour {machine.name} doit être supérieur ou égal à {old_hours} {unit}", "danger")
                    return redirect(request.url)
                if new_hours == old_hours:
                    continue
                log = CounterLog(machine=machine, previous_hours=old_hours, new_hours=new_hours)
                delta = new_hours - old_hours
                machine.hours = new_hours
                db.session.add(log)
                
                # Mettre à jour uniquement les plans de maintenance de cette machine spécifique
                existing_progress_list = MaintenanceProgress.query.filter_by(machine_id=machine.id, counter_id=None).all()
                
                # Créer les progress manquants pour cette machine
                ensure_all_progress_for_machine(machine)
                
                # Mettre à jour les hours_since pour TOUS les progress qui existaient avant
                for progress in existing_progress_list:
                    progress.hours_since = progress.hours_since - delta
                    db.session.add(progress)
                updated += 1
                machines_updated.append(machine.name)
            
            elif item['type'] == 'machine_with_counters':
                # Compteurs multiples pour machine racine
                machine = item['machine']
                for counter in item['counters']:
                    raw_value = request.form.get(f"counter_{counter.id}")
                    if raw_value is None or raw_value.strip() == "":
                        continue
                    try:
                        new_value = float(raw_value)
                    except ValueError:
                        flash(f"Valeur invalide pour {counter.name}", "danger")
                        return redirect(request.url)
                    old_value = counter.value
                    if new_value < old_value:
                        unit = counter.unit or 'h'
                        flash(f"Le nouveau compteur pour {counter.name} doit être supérieur ou égal à {old_value} {unit}", "danger")
                        return redirect(request.url)
                    if new_value == old_value:
                        continue
                    delta = new_value - old_value
                    counter.value = new_value
                    db.session.add(counter)
                    
                    # Créer un CounterLog pour ce compteur multiple
                    log = CounterLog(
                        machine=machine,
                        counter_id=counter.id,
                        previous_hours=old_value,
                        new_hours=new_value
                    )
                    db.session.add(log)
                    
                    # Mettre à jour les progress de maintenance liés à ce compteur
                    # Pour toutes les machines de l'arborescence qui utilisent ce compteur
                    all_machines_in_tree = get_all_descendants(machine)
                    all_machines_in_tree.append(machine)
                    
                    for m in all_machines_in_tree:
                        existing_progress_list = MaintenanceProgress.query.filter_by(
                            machine_id=m.id, 
                            counter_id=counter.id
                        ).all()
                        
                        for progress in existing_progress_list:
                            progress.hours_since = progress.hours_since - delta
                            db.session.add(progress)
                    updated += 1
                    machines_updated.append(f"{counter.name} ({machine.name})")
        
        if updated == 0:
            flash("Aucune valeur saisie ou changement détecté.", "warning")
            return redirect(request.url)
        try:
            db.session.commit()
            # Message automatique pour le chat (un seul message pour tous les relevés)
            if updated > 0:
                machines_info = ""
                if machines_updated:
                    if len(machines_updated) == 1:
                        machines_info = f" sur la machine '{machines_updated[0]}'"
                    elif len(machines_updated) <= 3:
                        machines_list = ', '.join([f"'{m}'" for m in machines_updated])
                        machines_info = f" sur les machines : {machines_list}"
                    else:
                        machines_info = f" sur {len(machines_updated)} machines"
                
                create_chat_message(
                    message_type="auto",
                    content=f"{current_user.username} a effectué {updated} relevé(s) de compteur(s){machines_info}",
                    link_url=url_for("counter_logs"),
                    machine_id=root_machine.id if root_machine else None
                )
            flash("Relevé des compteurs enregistré", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur: {exc}", "danger")
        # Rediriger vers la page arborescence machine
        return redirect(url_for("machines"))

    recent_logs = CounterLog.query.order_by(CounterLog.created_at.desc()).limit(50).all()
    root_machine_name = root_machine.name if machine_id and root_machine else None
    return render_template("counter_report.html", 
                         counter_hierarchy=counter_hierarchy,
                         root_machine=root_machine,
                         logs=recent_logs, 
                         root_machine_name=root_machine_name)


@app.route("/machines/<int:machine_id>/edit-counter-report", methods=["GET", "POST"])
@admin_required
def edit_counter_report(machine_id):
    """Page pour modifier les relevés de compteurs (augmenter ou diminuer)"""
    machine = Machine.query.get_or_404(machine_id)
    
    # Récupérer toutes les machines de l'arborescence avec compteur
    all_machines_in_tree = get_all_descendants(machine)
    machines_with_counters = [m for m in all_machines_in_tree if m.hour_counter_enabled]
    machines_with_counters.sort(key=lambda m: m.name)
    
    # Récupérer les compteurs de la machine racine si c'est une machine racine
    root_counters_by_machine = []
    root_machine = machine
    while root_machine.parent:
        root_machine = root_machine.parent
    
    if root_machine.is_root() and root_machine.counters:
        root_counters = sorted(root_machine.counters, key=lambda c: c.name)
        root_counters_by_machine = [(root_machine, counter) for counter in root_counters]
    
    if not machines_with_counters and not root_counters_by_machine:
        flash(f"Aucun compteur configuré pour cette machine.", "warning")
        return redirect(get_machine_detail_url(machine_id))
    
    if request.method == "POST":
        updated = 0
        machines_updated = []
        
        # Traiter les compteurs des machines
        for m in machines_with_counters:
            raw_value = request.form.get(f"machine_{m.id}")
            if raw_value is None or raw_value.strip() == "":
                continue
            try:
                new_hours = float(raw_value)
            except ValueError:
                flash(f"Valeur invalide pour {m.name}", "danger")
                return redirect(request.url)
            
            old_hours = m.hours
            if new_hours == old_hours:
                continue
            
            delta = new_hours - old_hours
            m.hours = new_hours
            
            # Créer un log
            log = CounterLog(machine=m, previous_hours=old_hours, new_hours=new_hours)
            db.session.add(log)
            
            # Récupérer les progress existants AVANT de créer les nouveaux
            existing_progress_list = MaintenanceProgress.query.filter_by(machine_id=m.id, counter_id=None).all()
            
            # Créer les progress manquants pour cette machine
            ensure_all_progress_for_machine(m)
            
            # Mettre à jour les hours_since UNIQUEMENT pour les progress qui existaient avant
            # (les nouveaux progress créés auront la valeur initiale correcte)
            for progress in existing_progress_list:
                # Si on diminue le compteur (delta négatif), on augmente hours_since
                # Si on augmente le compteur (delta positif), on diminue hours_since
                progress.hours_since = progress.hours_since - delta
                db.session.add(progress)
            
            updated += 1
            machines_updated.append(m.name)
        
        # Traiter les compteurs multiples des machines racines
        for root_machine_for_counter, counter in root_counters_by_machine:
            raw_value = request.form.get(f"counter_{counter.id}")
            if raw_value is None or raw_value.strip() == "":
                continue
            try:
                new_value = float(raw_value)
            except ValueError:
                flash(f"Valeur invalide pour {counter.name}", "danger")
                return redirect(request.url)
            
            old_value = counter.value
            if new_value == old_value:
                continue
            
            delta = new_value - old_value
            counter.value = new_value
            
            # Créer un log
            log = CounterLog(
                machine=root_machine_for_counter,
                counter_id=counter.id,
                previous_hours=old_value,
                new_hours=new_value
            )
            db.session.add(log)
            
            # Mettre à jour les progress de maintenance pour toutes les machines de l'arborescence
            all_machines_in_tree = get_all_descendants(root_machine_for_counter)
            all_machines_in_tree.append(root_machine_for_counter)
            
            for m in all_machines_in_tree:
                # Récupérer les progress existants AVANT de créer les nouveaux
                existing_progress_list = MaintenanceProgress.query.filter_by(
                    machine_id=m.id, 
                    counter_id=counter.id
                ).all()
                
                # Créer les progress manquants pour cette machine
                ensure_all_progress_for_machine(m)
                
                # Mettre à jour les hours_since UNIQUEMENT pour les progress qui existaient avant
                # (les nouveaux progress créés auront la valeur initiale correcte)
                for progress in existing_progress_list:
                    # Si on diminue le compteur (delta négatif), on augmente hours_since
                    # Si on augmente le compteur (delta positif), on diminue hours_since
                    progress.hours_since = progress.hours_since - delta
                    db.session.add(progress)
            
            updated += 1
            machines_updated.append(f"{counter.name} ({root_machine_for_counter.name})")
        
        if updated == 0:
            flash("Aucune modification détectée.", "warning")
            return redirect(request.url)
        
        try:
            db.session.commit()
            flash(f"Relevé modifié pour {len(machines_updated)} compteur(s): {', '.join(machines_updated)}", "success")
            return redirect(get_machine_detail_url(machine_id))
        except Exception as exc:
            db.session.rollback()
            flash(f"Erreur lors de la modification: {exc}", "danger")
            return redirect(request.url)
    
    return render_template("edit_counter_report.html", 
                         machine=machine,
                         machines=machines_with_counters, 
                         root_counters_by_machine=root_counters_by_machine,
                         root_machine=root_machine)


@app.route("/counter-logs")
@login_required
def counter_logs():
    logs = CounterLog.query.order_by(CounterLog.created_at.desc()).all()
    return render_template("counter_logs.html", logs=logs)


@app.route("/maintenance-tracking")
@login_required
def maintenance_tracking():
    """Page de suivi des actions M&Ms (Maintenances & Machines)"""
    # Récupérer les paramètres de filtrage
    filter_type = request.args.get('filter_type', '').strip()
    filter_machine_id = request.args.get('filter_machine_id', '').strip()
    filter_date_start = request.args.get('filter_date_start', '').strip()
    filter_date_end = request.args.get('filter_date_end', '').strip()
    
    # Convertir filter_machine_id en int si fourni
    try:
        machine_id = int(filter_machine_id) if filter_machine_id else None
    except (ValueError, TypeError):
        machine_id = None
    
    # Convertir les dates
    date_start = None
    date_end = None
    if filter_date_start:
        try:
            date_start = dt.datetime.strptime(filter_date_start, '%Y-%m-%d')
        except ValueError:
            pass
    if filter_date_end:
        try:
            date_end = dt.datetime.strptime(filter_date_end, '%Y-%m-%d')
            # Ajouter 23h59 pour inclure toute la journée
            date_end = date_end.replace(hour=23, minute=59, second=59)
        except ValueError:
            pass
    
    # Récupérer toutes les actions
    all_actions = []
    
    # Maintenances préventives
    if not filter_type or filter_type == 'preventive':
        query = MaintenanceEntry.query
        if machine_id:
            query = query.filter_by(machine_id=machine_id)
        if date_start:
            query = query.filter(MaintenanceEntry.created_at >= date_start)
        if date_end:
            query = query.filter(MaintenanceEntry.created_at <= date_end)
        entries = query.order_by(MaintenanceEntry.created_at.desc()).all()
        for entry in entries:
            all_actions.append({
                'type': 'preventive',
                'type_label': 'Maintenance préventive',
                'name': entry.report.name,
                'date': entry.created_at,
                'machine': entry.machine,
                'user': entry.user,
                'url': url_for('maintenance_entry_detail', entry_id=entry.id)
            })
    
    # Maintenances correctives
    if not filter_type or filter_type == 'corrective':
        query = CorrectiveMaintenance.query
        if machine_id:
            query = query.filter_by(machine_id=machine_id)
        if date_start:
            query = query.filter(CorrectiveMaintenance.created_at >= date_start)
        if date_end:
            query = query.filter(CorrectiveMaintenance.created_at <= date_end)
        correctives = query.order_by(CorrectiveMaintenance.created_at.desc()).all()
        for maintenance in correctives:
            all_actions.append({
                'type': 'corrective',
                'type_label': 'Maintenance corrective',
                'name': 'Maintenance corrective',
                'date': maintenance.created_at,
                'machine': maintenance.machine,
                'user': maintenance.user,
                'url': url_for('corrective_maintenance_detail', maintenance_id=maintenance.id)
            })
    
    # Checklists
    if not filter_type or filter_type == 'checklist':
        query = ChecklistInstance.query
        if machine_id:
            query = query.filter_by(machine_id=machine_id)
        if date_start:
            query = query.filter(ChecklistInstance.created_at >= date_start)
        if date_end:
            query = query.filter(ChecklistInstance.created_at <= date_end)
        checklists = query.order_by(ChecklistInstance.created_at.desc()).all()
        for checklist in checklists:
            all_actions.append({
                'type': 'checklist',
                'type_label': 'Check-list',
                'name': checklist.template.name,
                'date': checklist.created_at,
                'machine': checklist.machine,
                'user': checklist.user,
                'url': url_for('checklist_instance_detail', machine_id=checklist.machine_id, template_id=checklist.template_id, instance_id=checklist.id)
            })
    
    # Relevés de compteurs
    if not filter_type or filter_type == 'counter':
        query = CounterLog.query
        if machine_id:
            query = query.filter_by(machine_id=machine_id)
        if date_start:
            query = query.filter(CounterLog.created_at >= date_start)
        if date_end:
            query = query.filter(CounterLog.created_at <= date_end)
        counter_logs = query.order_by(CounterLog.created_at.desc()).all()
        for log in counter_logs:
            counter_name = log.counter.name if log.counter else "Compteur machine"
            all_actions.append({
                'type': 'counter',
                'type_label': 'Relevé compteur',
                'name': counter_name,
                'date': log.created_at,
                'machine': log.machine,
                'user': None,  # CounterLog n'a pas de user_id
                'url': url_for('counter_logs')
            })
    
    # Trier par date décroissante
    all_actions.sort(key=lambda x: x['date'], reverse=True)
    
    # Récupérer toutes les machines racines pour le filtre hiérarchique
    root_machines = Machine.query.filter_by(parent_id=None).order_by(Machine.name).all()
    
    return render_template(
        "maintenance_tracking.html",
        actions=all_actions,
        root_machines=root_machines,
        filter_type=filter_type,
        filter_machine_id=machine_id,
        filter_date_start=filter_date_start,
        filter_date_end=filter_date_end
    )


@app.route("/stock-tracking")
@login_required
def stock_tracking():
    """Page de suivi des actions S&P (Stocks & Produits)"""
    # Récupérer les paramètres de filtrage
    filter_type = request.args.get('filter_type', '').strip()
    filter_date_start = request.args.get('filter_date_start', '').strip()
    filter_date_end = request.args.get('filter_date_end', '').strip()
    
    # Convertir les dates
    date_start = None
    date_end = None
    if filter_date_start:
        try:
            date_start = dt.datetime.strptime(filter_date_start, '%Y-%m-%d')
        except ValueError:
            pass
    if filter_date_end:
        try:
            date_end = dt.datetime.strptime(filter_date_end, '%Y-%m-%d')
            # Ajouter 23h59 pour inclure toute la journée
            date_end = date_end.replace(hour=23, minute=59, second=59)
        except ValueError:
            pass
    
    # Récupérer toutes les actions
    all_actions = []
    
    # Mouvements
    if not filter_type or filter_type == 'movement':
        query = Movement.query
        if date_start:
            query = query.filter(Movement.created_at >= date_start)
        if date_end:
            query = query.filter(Movement.created_at <= date_end)
        movements = query.order_by(Movement.created_at.desc()).all()
        for movement in movements:
            all_actions.append({
                'type': 'movement',
                'type_label': 'Mouvement',
                'name': f"Mouvement {movement.type}",
                'date': movement.created_at,
                'stock_source': movement.source_stock,
                'stock_dest': movement.dest_stock,
                'user': None,  # Movement n'a pas de user_id
                'url': url_for('movements')
            })
    
    # Inventaires
    if not filter_type or filter_type == 'inventory':
        query = Inventory.query
        if date_start:
            query = query.filter(Inventory.created_at >= date_start)
        if date_end:
            query = query.filter(Inventory.created_at <= date_end)
        inventories = query.order_by(Inventory.created_at.desc()).all()
        for inventory in inventories:
            # Utiliser le nom de l'inventaire s'il existe, sinon générer un nom par défaut
            inventory_name = inventory.name if inventory.name else f"Inventaire {inventory.stock.name}"
            all_actions.append({
                'type': 'inventory',
                'type_label': 'Inventaire',
                'name': inventory_name,
                'date': inventory.created_at,
                'stock': inventory.stock,
                'user': inventory.user,
                'url': url_for('inventory_detail', inventory_id=inventory.id)
            })
    
    # Trier par date décroissante
    all_actions.sort(key=lambda x: x['date'], reverse=True)
    
    return render_template(
        "stock_tracking.html",
        actions=all_actions,
        filter_type=filter_type,
        filter_date_start=filter_date_start,
        filter_date_end=filter_date_end
    )


@app.route("/database-export")
@admin_required
def database_export():
    """Page de gestion des exports de base de données"""
    excel_files = ExcelFile.query.order_by(ExcelFile.created_at.desc()).all()
    return render_template("database_export.html", excel_files=excel_files)


@app.route("/database-export/maintenances/excel")
@admin_required
def export_maintenances_excel():
    """Export Excel complet des maintenances avec tous les détails"""
    entries = MaintenanceEntry.query.order_by(MaintenanceEntry.created_at.desc()).all()
    corrective = CorrectiveMaintenance.query.order_by(CorrectiveMaintenance.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenances"
    
    # En-têtes complets
    headers = [
        "Type", "Nom", "Date", "Machine", "Code Machine", "Identifiant", "Stock", 
        "Compteur", "Heures avant maintenance", "Composants (Label|Type|Valeur)", 
        "Produits utilisés (Nom|Code|Quantité)", "Commentaire"
    ]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Maintenances préventives
    for entry in entries:
        machine_lineage_str = " > ".join([node.name for node in machine_lineage(entry.machine)])
        unit = entry.machine.counter_unit or 'h' if entry.machine.hour_counter_enabled else None
        counter_str = f"{entry.performed_hours} {unit}" if unit else "-"
        hours_before = f"{entry.hours_before_maintenance:.1f} {unit}" if entry.hours_before_maintenance is not None and unit else "-"
        
        # Récupérer les composants avec leurs valeurs
        components_list = []
        for value in entry.values:
            if value.component.field_type == "checkbox":
                val_str = "Oui" if value.value_bool else "Non"
            elif value.component.field_type == "number":
                val_str = str(value.value_number)
            else:
                val_str = value.value_text or ""
            components_list.append(f"{value.component.label}|{value.component.field_type}|{val_str}")
        components_str = " || ".join(components_list) if components_list else ""
        
        # Récupérer les produits utilisés via le mouvement associé
        products_list = []
        if entry.stock:
            time_window_start = entry.created_at - dt.timedelta(minutes=5)
            time_window_end = entry.created_at + dt.timedelta(minutes=5)
            movements = Movement.query.filter(
                Movement.type == "sortie",
                Movement.source_stock_id == entry.stock.id,
                Movement.created_at >= time_window_start,
                Movement.created_at <= time_window_end
            ).all()
            for movement in movements:
                for item in movement.items:
                    products_list.append(f"{item.product.name}|{item.product.code}|{item.quantity}")
        products_str = " || ".join(products_list) if products_list else ""
        
        ws.append([
            "Préventive",
            entry.report.name,
            entry.created_at.strftime("%d/%m/%Y %H:%M"),
            machine_lineage_str,
            entry.machine.code,
            entry.user.username if entry.user else "",
            entry.stock.name if entry.stock else "",
            counter_str,
            hours_before,
            components_str,
            products_str,
            ""  # Pas de commentaire pour les maintenances préventives
        ])
    
    # Maintenances correctives
    for maintenance in corrective:
        machine_lineage_str = " > ".join([node.name for node in machine_lineage(maintenance.machine)])
        
        # Récupérer les produits utilisés
        products_list = []
        for product_item in maintenance.products:
            products_list.append(f"{product_item.product.name}|{product_item.product.code}|{product_item.quantity}")
        products_str = " || ".join(products_list) if products_list else ""
        
        ws.append([
            "Corrective",
            "Maintenance corrective",
            maintenance.created_at.strftime("%d/%m/%Y %H:%M"),
            machine_lineage_str,
            maintenance.machine.code,
            maintenance.user.username if maintenance.user else "",
            maintenance.stock.name if maintenance.stock else "",
            "-",
            "-",
            "",  # Pas de composants pour les maintenances correctives
            products_str,
            maintenance.comment or ""
        ])
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=maintenances_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/database-export/modeles/excel")
@admin_required
def export_modeles_excel():
    """Export Excel des modèles de maintenance"""
    reports = PreventiveReport.query.order_by(PreventiveReport.name).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Modèles"
    
    # En-têtes
    headers = ["Nom", "Machine", "Code Machine", "Périodicité", "Nombre d'éléments"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for report in reports:
        machine_lineage_str = " > ".join([node.name for node in machine_lineage(report.machine)])
        ws.append([
            report.name,
            machine_lineage_str,
            report.machine.code,
            report.periodicity,
            len(report.components)
        ])
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=modeles_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/database-export/machines/excel")
@admin_required
def export_machines_excel():
    """Export Excel de l'arborescence des machines"""
    machines = Machine.query.order_by(Machine.name).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Machines"
    
    # En-têtes
    headers = ["Nom", "Code", "Machine parente", "Compteur activé", "Valeur compteur", "Unité"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for machine in machines:
        parent_name = machine.parent.name if machine.parent else ""
        ws.append([
            machine.name,
            machine.code,
            parent_name,
            "Oui" if machine.hour_counter_enabled else "Non",
            machine.hours if machine.hour_counter_enabled else "",
            machine.counter_unit or "" if machine.hour_counter_enabled else ""
        ])
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=arborescence_machines_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/database-export/releves/excel")
@admin_required
def export_releves_excel():
    """Export Excel des relevés compteur"""
    logs = CounterLog.query.order_by(CounterLog.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Relevés"
    
    # En-têtes
    headers = ["Date", "Machine", "Code", "Ancien compteur", "Nouveau compteur", "Différence", "Unité"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for log in logs:
        unit = log.machine.counter_unit or 'h'
        delta = log.new_hours - log.previous_hours
        ws.append([
            log.created_at.strftime("%d/%m/%Y %H:%M"),
            log.machine.name,
            log.machine.code,
            log.previous_hours,
            log.new_hours,
            delta,
            unit
        ])
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=releves_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/database-export/produits/excel")
@admin_required
def export_produits_excel():
    """Export Excel des produits"""
    products = Product.query.order_by(Product.name).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Produits"
    
    # En-têtes
    headers = ["Nom", "Code", "Prix", "Fournisseur", "Stock minimum"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for product in products:
        ws.append([
            product.name,
            product.code,
            product.price,
            product.supplier_name or "",
            product.minimum_stock
        ])
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=produits_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/database-export/mouvements/excel")
@admin_required
def export_mouvements_excel():
    """Export Excel des mouvements"""
    movements = Movement.query.order_by(Movement.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Mouvements"
    
    # En-têtes
    headers = ["Date", "Type", "Stock source", "Stock destination", "Produits", "Quantités"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for movement in movements:
        products_list = []
        quantities_list = []
        for item in movement.items:
            product_name = item.product.name if item.product else "Produit inconnu"
            product_code = item.product.code if item.product else "-"
            products_list.append(f"{product_name} ({product_code})")
            quantities_list.append(str(item.quantity))
        
        ws.append([
            movement.created_at.strftime("%d/%m/%Y %H:%M"),
            movement.type.capitalize(),
            movement.source_stock.name if movement.source_stock else "-",
            movement.dest_stock.name if movement.dest_stock else "-",
            ", ".join(products_list),
            ", ".join(quantities_list)
        ])
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=mouvements_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/database-export/inventaires/excel")
@admin_required
def export_inventaires_excel():
    """Export Excel des inventaires"""
    inventories = Inventory.query.order_by(Inventory.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaires"
    
    # En-têtes
    headers = ["Date", "Stock", "Code Stock", "Produit", "Code Produit", "Ancienne quantité", "Nouvelle quantité", "Différence", "Commentaire"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for inventory in inventories:
        for item in inventory.items:
            diff = item.new_quantity - item.old_quantity
            ws.append([
                inventory.created_at.strftime("%d/%m/%Y %H:%M"),
                inventory.stock.name,
                inventory.stock.code,
                item.product.name,
                item.product.code,
                item.old_quantity,
                item.new_quantity,
                diff,
                inventory.comment or ""
            ])
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=inventaires_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@app.route("/database-export/all/json")
@admin_required
def export_all_json():
    """Export JSON complet de toute la base de données"""
    
    # Fonction helper pour convertir datetime en string
    def datetime_to_str(obj):
        if isinstance(obj, dt.datetime):
            return obj.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(obj, dt.date):
            return obj.strftime("%Y-%m-%d")
        return obj
    
    # Fonction helper pour convertir un objet SQLAlchemy en dict
    def model_to_dict(model_instance):
        if model_instance is None:
            return None
        result = {}
        for column in model_instance.__table__.columns:
            try:
                value = getattr(model_instance, column.name)
                # Gérer les valeurs None et les types non sérialisables
                if value is None:
                    result[column.name] = None
                else:
                    result[column.name] = datetime_to_str(value)
            except Exception:
                result[column.name] = None
        return result
    
    # Collecter toutes les données
    # Pour les utilisateurs, exclure le password_hash pour des raisons de sécurité
    users_data = []
    for user in User.query.all():
        user_dict = model_to_dict(user)
        if user_dict and 'password_hash' in user_dict:
            user_dict['password_hash'] = None  # Ne pas exporter les mots de passe
        users_data.append(user_dict)
    
    data = {
        "export_date": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "users": users_data,
        "machines": [],
        "products": [model_to_dict(product) for product in Product.query.all()],
        "stocks": [],
        "preventive_reports": [],
        "maintenance_entries": [],
        "corrective_maintenances": [],
        "counter_logs": [],
        "movements": [],
        "inventories": []
    }
    
    # Machines avec arborescence
    for machine in Machine.query.all():
        machine_dict = model_to_dict(machine)
        machine_dict["parent_name"] = machine.parent.name if machine.parent else None
        machine_dict["parent_code"] = machine.parent.code if machine.parent else None
        data["machines"].append(machine_dict)
    
    # Stocks avec produits
    for stock in Stock.query.all():
        stock_dict = model_to_dict(stock)
        stock_dict["products"] = []
        stock_products = StockProduct.query.filter_by(stock_id=stock.id).all()
        for sp in stock_products:
            product = Product.query.get(sp.product_id)
            stock_dict["products"].append({
                "product_id": sp.product_id,
                "product_name": product.name if product else None,
                "product_code": product.code if product else None,
                "quantity": sp.quantity
            })
        data["stocks"].append(stock_dict)
    
    # Modèles de maintenance avec composants
    for report in PreventiveReport.query.all():
        report_dict = model_to_dict(report)
        report_dict["machine_name"] = report.machine.name if report.machine else None
        report_dict["machine_code"] = report.machine.code if report.machine else None
        report_dict["components"] = []
        for component in report.components:
            comp_dict = model_to_dict(component)
            report_dict["components"].append(comp_dict)
        data["preventive_reports"].append(report_dict)
    
    # Maintenances préventives avec valeurs et produits
    for entry in MaintenanceEntry.query.all():
        entry_dict = model_to_dict(entry)
        entry_dict["machine_name"] = entry.machine.name if entry.machine else None
        entry_dict["machine_code"] = entry.machine.code if entry.machine else None
        entry_dict["report_name"] = entry.report.name if entry.report else None
        entry_dict["user_username"] = entry.user.username if entry.user else None
        entry_dict["stock_name"] = entry.stock.name if entry.stock else None
        entry_dict["stock_code"] = entry.stock.code if entry.stock else None
        
        # Valeurs des composants
        entry_dict["values"] = []
        for value in entry.values:
            val_dict = {
                "component_label": value.component.label if value.component else None,
                "component_type": value.component.field_type if value.component else None,
                "value_text": value.value_text,
                "value_number": value.value_number,
                "value_bool": value.value_bool
            }
            entry_dict["values"].append(val_dict)
        
        # Produits utilisés
        entry_dict["products"] = []
        if entry.stock:
            time_window_start = entry.created_at - dt.timedelta(minutes=5)
            time_window_end = entry.created_at + dt.timedelta(minutes=5)
            movements = Movement.query.filter(
                Movement.type == "sortie",
                Movement.source_stock_id == entry.stock.id,
                Movement.created_at >= time_window_start,
                Movement.created_at <= time_window_end
            ).all()
            for movement in movements:
                for item in movement.items:
                    entry_dict["products"].append({
                        "product_name": item.product.name if item.product else None,
                        "product_code": item.product.code if item.product else None,
                        "quantity": item.quantity
                    })
        
        data["maintenance_entries"].append(entry_dict)
    
    # Maintenances correctives avec produits
    for maintenance in CorrectiveMaintenance.query.all():
        maint_dict = model_to_dict(maintenance)
        maint_dict["machine_name"] = maintenance.machine.name if maintenance.machine else None
        maint_dict["machine_code"] = maintenance.machine.code if maintenance.machine else None
        maint_dict["user_username"] = maintenance.user.username if maintenance.user else None
        maint_dict["stock_name"] = maintenance.stock.name if maintenance.stock else None
        maint_dict["stock_code"] = maintenance.stock.code if maintenance.stock else None
        
        # Produits
        maint_dict["products"] = []
        for product_item in maintenance.products:
            maint_dict["products"].append({
                "product_name": product_item.product.name if product_item.product else None,
                "product_code": product_item.product.code if product_item.product else None,
                "quantity": product_item.quantity
            })
        
        data["corrective_maintenances"].append(maint_dict)
    
    # Relevés compteur
    for log in CounterLog.query.all():
        log_dict = model_to_dict(log)
        log_dict["machine_name"] = log.machine.name if log.machine else None
        log_dict["machine_code"] = log.machine.code if log.machine else None
        log_dict["counter_unit"] = log.machine.counter_unit if log.machine else None
        data["counter_logs"].append(log_dict)
    
    # Mouvements avec items
    for movement in Movement.query.all():
        mov_dict = model_to_dict(movement)
        mov_dict["source_stock_name"] = movement.source_stock.name if movement.source_stock else None
        mov_dict["source_stock_code"] = movement.source_stock.code if movement.source_stock else None
        mov_dict["dest_stock_name"] = movement.dest_stock.name if movement.dest_stock else None
        mov_dict["dest_stock_code"] = movement.dest_stock.code if movement.dest_stock else None
        mov_dict["items"] = []
        for item in movement.items:
            mov_dict["items"].append({
                "product_name": item.product.name if item.product else None,
                "product_code": item.product.code if item.product else None,
                "quantity": item.quantity
            })
        data["movements"].append(mov_dict)
    
    # Inventaires avec items
    for inventory in Inventory.query.all():
        inv_dict = model_to_dict(inventory)
        inv_dict["stock_name"] = inventory.stock.name if inventory.stock else None
        inv_dict["stock_code"] = inventory.stock.code if inventory.stock else None
        inv_dict["user_username"] = inventory.user.username if inventory.user else None
        inv_dict["items"] = []
        for item in inventory.items:
            inv_dict["items"].append({
                "product_name": item.product.name if item.product else None,
                "product_code": item.product.code if item.product else None,
                "old_quantity": item.old_quantity,
                "new_quantity": item.new_quantity
            })
        data["inventories"].append(inv_dict)
    
    # Convertir en JSON avec gestion d'erreurs
    try:
        json_output = json.dumps(data, ensure_ascii=False, indent=2, default=str)
        
        response = make_response(json_output)
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=database_export_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        return response
    except Exception as exc:
        flash(f"Erreur lors de l'export JSON : {exc}", "danger")
        return redirect(url_for("database_export"))


@app.route("/counter-logs/export")
@login_required
def export_counter_logs():
    logs = CounterLog.query.order_by(CounterLog.created_at.desc()).all()
    
    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relevés compteur"
    
    # En-têtes
    headers = ["Date", "Machine", "Code", "Compteur", "Ancien compteur", "Nouveau compteur", "Différence", "Unité"]
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    for log in logs:
        if log.counter:
            counter_name = log.counter.name
            unit = log.counter.unit or 'h'
        else:
            counter_name = "Compteur machine"
            unit = log.machine.counter_unit or 'h'
        row = [
            log.created_at.strftime("%d/%m/%Y %H:%M"),
            log.machine.name,
            log.machine.code,
            counter_name,
            log.previous_hours,
            log.new_hours,
            log.new_hours - log.previous_hours,
            unit
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=releves_compteur_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


def apply_movement_rules(movement: Movement):
    def update_stock(stock_id, product_id, delta):
        record = _get_or_create_stock_product(stock_id, product_id)
        new_qty = record.quantity + delta
        if new_qty < 0:
            raise ValueError("Stock insuffisant pour le produit sélectionné")
        record.quantity = new_qty

    try:
        if movement.type == "entree":
            if not movement.dest_stock_id:
                return "Sélectionnez un stock de destination"
            for item in movement.items:
                update_stock(movement.dest_stock_id, item.product_id, item.quantity)
        elif movement.type == "sortie":
            if not movement.source_stock_id:
                return "Sélectionnez un stock source"
            for item in movement.items:
                update_stock(movement.source_stock_id, item.product_id, -item.quantity)
        elif movement.type == "transfert":
            if not movement.source_stock_id or not movement.dest_stock_id:
                return "Sélectionnez les stocks source et destination"
            if movement.source_stock_id == movement.dest_stock_id:
                return "Les stocks source et destination doivent être différents"
            for item in movement.items:
                update_stock(movement.source_stock_id, item.product_id, -item.quantity)
                update_stock(movement.dest_stock_id, item.product_id, item.quantity)
        else:
            return "Type de mouvement invalide"
    except ValueError as exc:
        return str(exc)
    return None


def reverse_movement_rules(movement: Movement):
    """Inverse les effets d'un mouvement sur les stocks"""
    def update_stock(stock_id, product_id, delta):
        record = _get_or_create_stock_product(stock_id, product_id)
        record.quantity = record.quantity + delta

    try:
        if movement.type == "entree":
            # Inverser : retirer du stock de destination
            for item in movement.items:
                update_stock(movement.dest_stock_id, item.product_id, -item.quantity)
        elif movement.type == "sortie":
            # Inverser : remettre dans le stock source
            for item in movement.items:
                update_stock(movement.source_stock_id, item.product_id, item.quantity)
        elif movement.type == "transfert":
            # Inverser : remettre dans source, retirer de destination
            for item in movement.items:
                update_stock(movement.source_stock_id, item.product_id, item.quantity)
                update_stock(movement.dest_stock_id, item.product_id, -item.quantity)
    except Exception as exc:
        raise ValueError(f"Erreur lors de l'inversion du mouvement : {exc}")


def build_machine_tree(node, level=0):
    yield node, level
    for child in sorted(node.children, key=lambda c: c.name):
        yield from build_machine_tree(child, level + 1)


def machine_lineage(machine):
    nodes = []
    current = machine
    while current:
        nodes.append(current)
        current = current.parent
    return list(reversed(nodes))


def get_all_descendants(machine):
    """Récupère la machine et toutes ses sous-machines (descendants)"""
    descendants = [machine]
    for child in machine.children:
        descendants.extend(get_all_descendants(child))
    return descendants


def build_counter_hierarchy(machine, depth=0):
    """Construit une structure hiérarchique des machines avec compteurs pour l'affichage"""
    items = []
    
    # Vérifier si cette machine a des compteurs à afficher
    has_counters = False
    
    # Compteurs multiples pour machines racines
    if machine.is_root() and machine.counters:
        counters_list = sorted(machine.counters, key=lambda c: c.name)
        if counters_list:
            has_counters = True
            items.append({
                'type': 'machine_with_counters',
                'machine': machine,
                'counters': counters_list,
                'depth': depth
            })
    
    # Compteur classique pour sous-machines
    elif machine.hour_counter_enabled:
        has_counters = True
        items.append({
            'type': 'machine_single_counter',
            'machine': machine,
            'depth': depth
        })
    
    # Parcourir récursivement les enfants (trier par nom pour un ordre cohérent)
    for child in sorted(machine.children, key=lambda m: m.name):
        items.extend(build_counter_hierarchy(child, depth + 1))
    
    return items


def has_counter_in_tree(machine):
    """Vérifie si une machine ou une de ses sous-machines a un compteur activé"""
    # Vérifier si la machine a un compteur horaire activé
    if machine.hour_counter_enabled:
        return True
    # Pour les machines racines, vérifier si elles ont des compteurs multiples
    if machine.is_root() and machine.counters:
        return True
    # Vérifier récursivement dans les enfants
    for child in machine.children:
        if has_counter_in_tree(child):
            return True
    return False


def create_chat_message(message_type, content, link_url=None, machine_id=None, user_id=None):
    """Crée un message de chat (automatique ou manuel)"""
    # Supprimer les messages de plus d'une semaine
    week_ago = dt.datetime.utcnow() - dt.timedelta(days=7)
    ChatMessage.query.filter(ChatMessage.created_at < week_ago).delete()
    
    message = ChatMessage(
        message_type=message_type,
        content=content,
        link_url=link_url,
        machine_id=machine_id,
        user_id=user_id
    )
    db.session.add(message)
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        print(f"Error creating chat message: {exc}")


# Routes pour le chat
@app.route("/chat/send", methods=["POST"])
@login_required
def chat_send():
    """Envoyer un message manuel dans le chat"""
    data = request.get_json()
    content = data.get("content", "").strip()
    reply_to_id = data.get("reply_to_id")
    
    if not content:
        return json.dumps({"success": False, "error": "Le message ne peut pas être vide"}), 400
    
    # Vérifier que le message auquel on répond existe
    reply_to = None
    if reply_to_id:
        reply_to = ChatMessage.query.filter_by(id=reply_to_id, deleted_at=None).first()
        if not reply_to:
            return json.dumps({"success": False, "error": "Le message auquel vous répondez n'existe plus"}), 400
    
    message = ChatMessage(
        message_type="manual",
        content=content,
        user_id=current_user.id,
        reply_to_id=reply_to_id if reply_to else None
    )
    db.session.add(message)
    try:
        db.session.commit()
        return json.dumps({"success": True, "message_id": message.id})
    except Exception as exc:
        db.session.rollback()
        return json.dumps({"success": False, "error": str(exc)}), 500


@app.route("/chat/messages")
@login_required
def chat_messages():
    """Récupérer les messages du chat (1 semaine d'historique)"""
    week_ago = dt.datetime.utcnow() - dt.timedelta(days=7)
    messages = ChatMessage.query.filter(
        ChatMessage.created_at >= week_ago,
        ChatMessage.deleted_at == None  # Ne pas afficher les messages supprimés
    ).order_by(ChatMessage.created_at.desc()).limit(100).all()
    
    messages_data = []
    for msg in reversed(messages):  # Inverser pour avoir les plus anciens en premier
        reply_to_data = None
        if msg.reply_to_id and msg.reply_to:
            reply_to_data = {
                "id": msg.reply_to.id,
                "content": msg.reply_to.content[:100] + ("..." if len(msg.reply_to.content) > 100 else ""),
                "user_name": msg.reply_to.user.username if msg.reply_to.user else "Système"
            }
        
        messages_data.append({
            "id": msg.id,
            "type": msg.message_type,
            "content": msg.content,
            "link_url": msg.link_url,
            "machine_name": msg.machine.name if msg.machine else None,
            "user_name": msg.user.username if msg.user else None,
            "user_id": msg.user_id,
            "is_own": msg.user_id == current_user.id if msg.user_id else False,
            "reply_to": reply_to_data,
            "edited": msg.edited_at is not None,
            "created_at": msg.created_at.isoformat(),
            "date": msg.created_at.strftime("%d/%m/%Y"),
            "time": msg.created_at.strftime("%H:%M")
        })
    
    return json.dumps({"success": True, "messages": messages_data})


@app.route("/chat/mark-read", methods=["POST"])
@login_required
def chat_mark_read():
    """Marquer les messages comme lus pour l'utilisateur actuel"""
    read_status = ChatReadStatus.query.filter_by(user_id=current_user.id).first()
    if read_status:
        read_status.last_read_at = dt.datetime.utcnow()
    else:
        read_status = ChatReadStatus(user_id=current_user.id, last_read_at=dt.datetime.utcnow())
        db.session.add(read_status)
    
    try:
        db.session.commit()
        return json.dumps({"success": True})
    except Exception as exc:
        db.session.rollback()
        return json.dumps({"success": False, "error": str(exc)}), 500


@app.route("/chat/unread-count")
@login_required
def chat_unread_count():
    """Compter les messages non lus pour l'utilisateur actuel"""
    read_status = ChatReadStatus.query.filter_by(user_id=current_user.id).first()
    
    # Si l'utilisateur n'a jamais marqué de messages comme lus, 
    # on considère qu'il n'a rien lu depuis toujours (il verra tous les messages de la semaine)
    # Sinon, on utilise sa dernière date de lecture
    week_ago = dt.datetime.utcnow() - dt.timedelta(days=7)
    if read_status:
        last_read_at = read_status.last_read_at
    else:
        # Première connexion : considérer qu'il n'a rien lu depuis toujours
        # On limite à 1 semaine d'historique comme défini dans le système
        last_read_at = dt.datetime.utcnow() - dt.timedelta(days=365)
    
    # Compter les messages manuels non lus (limité à 1 semaine d'historique)
    manual_unread = ChatMessage.query.filter(
        ChatMessage.message_type == "manual",
        ChatMessage.created_at > last_read_at,
        ChatMessage.created_at >= week_ago
    ).count()
    
    # Compter les messages automatiques non lus (limité à 1 semaine d'historique)
    auto_unread = ChatMessage.query.filter(
        ChatMessage.message_type == "auto",
        ChatMessage.created_at > last_read_at,
        ChatMessage.created_at >= week_ago
    ).count()
    
    return json.dumps({
        "success": True,
        "manual_count": manual_unread,
        "auto_count": auto_unread
    })


@app.route("/chat/<int:message_id>/edit", methods=["POST"])
@login_required
def chat_edit_message(message_id):
    """Modifier un message de l'utilisateur"""
    data = request.get_json()
    new_content = data.get("content", "").strip()
    
    if not new_content:
        return json.dumps({"success": False, "error": "Le message ne peut pas être vide"}), 400
    
    message = ChatMessage.query.filter_by(id=message_id, deleted_at=None).first_or_404()
    
    # Vérifier que c'est le message de l'utilisateur
    if message.user_id != current_user.id:
        return json.dumps({"success": False, "error": "Vous ne pouvez modifier que vos propres messages"}), 403
    
    # Vérifier que ce n'est pas un message automatique
    if message.message_type != "manual":
        return json.dumps({"success": False, "error": "Impossible de modifier ce type de message"}), 403
    
    message.content = new_content
    message.edited_at = dt.datetime.utcnow()
    
    try:
        db.session.commit()
        return json.dumps({"success": True})
    except Exception as exc:
        db.session.rollback()
        return json.dumps({"success": False, "error": str(exc)}), 500


@app.route("/chat/<int:message_id>/delete", methods=["POST"])
@login_required
def chat_delete_message(message_id):
    """Supprimer un message de l'utilisateur (soft delete)"""
    message = ChatMessage.query.filter_by(id=message_id, deleted_at=None).first_or_404()
    
    # Vérifier que c'est le message de l'utilisateur
    if message.user_id != current_user.id:
        return json.dumps({"success": False, "error": "Vous ne pouvez supprimer que vos propres messages"}), 403
    
    # Vérifier que ce n'est pas un message automatique
    if message.message_type != "manual":
        return json.dumps({"success": False, "error": "Impossible de supprimer ce type de message"}), 403
    
    message.deleted_at = dt.datetime.utcnow()
    
    try:
        db.session.commit()
        return json.dumps({"success": True})
    except Exception as exc:
        db.session.rollback()
        return json.dumps({"success": False, "error": str(exc)}), 500


app.jinja_env.globals["build_machine_tree"] = build_machine_tree
app.jinja_env.globals["machine_lineage"] = machine_lineage
app.jinja_env.globals["has_counter_in_tree"] = has_counter_in_tree


@app.route("/reports", methods=["GET"])
@login_required
def get_reports():
    """Récupérer les rapports (1 semaine d'historique)"""
    week_ago = dt.datetime.utcnow() - dt.timedelta(days=7)
    reports = Report.query.filter(
        Report.deleted_at.is_(None),
        Report.created_at >= week_ago
    ).order_by(Report.created_at.desc()).all()
    
    reports_data = []
    for report in reports:
        photos_data = []
        for photo in report.photos:
            photos_data.append({
                "id": photo.id,
                "url": url_for("report_photo", photo_id=photo.id),
                "original_filename": photo.original_filename
            })
        
        reports_data.append({
            "id": report.id,
            "user_name": report.user.username,
            "content": report.content,
            "created_at": report.created_at.isoformat(),
            "edited_at": report.edited_at.isoformat() if report.edited_at else None,
            "is_own": report.user_id == current_user.id,
            "photos": photos_data
        })
    
    return jsonify({"success": True, "reports": reports_data})


@app.route("/reports", methods=["POST"])
@login_required
def create_report():
    """Créer un nouveau rapport"""
    content = request.form.get("content", "").strip()
    photos = request.files.getlist("photos")
    
    if not content:
        return jsonify({"success": False, "error": "Le contenu est requis"}), 400
    
    # Créer le rapport
    report = Report(
        user_id=current_user.id,
        content=content
    )
    db.session.add(report)
    db.session.flush()  # Pour obtenir l'ID du rapport
    
    # Traiter les photos - stocker en base de données (BLOB)
    for photo in photos:
        if photo and photo.filename and allowed_image_file(photo.filename):
            original_filename = photo.filename
            # Lire les données binaires de l'image
            photo_data = photo.read()
            # Déterminer le type MIME
            content_type = photo.content_type or 'image/jpeg'
            
            report_photo = ReportPhoto(
                report_id=report.id,
                original_filename=original_filename,
                photo_data=photo_data,
                content_type=content_type
            )
            db.session.add(report_photo)
    
    try:
        db.session.commit()
        return jsonify({"success": True, "report_id": report.id})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/reports/<int:report_id>", methods=["PUT"])
@login_required
def update_report(report_id):
    """Modifier un rapport"""
    report = Report.query.filter_by(id=report_id, deleted_at=None).first_or_404()
    
    # Vérifier que l'utilisateur est le propriétaire
    if report.user_id != current_user.id:
        return jsonify({"success": False, "error": "Vous ne pouvez modifier que vos propres rapports"}), 403
    
    content = request.form.get("content", "").strip()
    if not content:
        return jsonify({"success": False, "error": "Le contenu est requis"}), 400
    
    # Supprimer les photos existantes si demandé
    photos_to_delete = request.form.getlist("delete_photos")
    for photo_id in photos_to_delete:
        try:
            photo_id_int = int(photo_id)
            photo = ReportPhoto.query.filter_by(id=photo_id_int, report_id=report.id).first()
            if photo:
                # Supprimer le fichier si il existe (migration)
                if photo.file_path and os.path.exists(photo.file_path):
                    try:
                        os.remove(photo.file_path)
                    except Exception:
                        pass
                db.session.delete(photo)
        except (ValueError, TypeError):
            pass
    
    # Ajouter de nouvelles photos - stocker en base de données (BLOB)
    new_photos = request.files.getlist("photos")
    for photo in new_photos:
        if photo and photo.filename and allowed_image_file(photo.filename):
            original_filename = photo.filename
            # Lire les données binaires de l'image
            photo_data = photo.read()
            # Déterminer le type MIME
            content_type = photo.content_type or 'image/jpeg'
            
            report_photo = ReportPhoto(
                report_id=report.id,
                original_filename=original_filename,
                photo_data=photo_data,
                content_type=content_type
            )
            db.session.add(report_photo)
    
    # Mettre à jour le contenu
    report.content = content
    report.edited_at = dt.datetime.utcnow()
    
    try:
        db.session.commit()
        return jsonify({"success": True})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/reports/<int:report_id>", methods=["DELETE"])
@login_required
def delete_report(report_id):
    """Supprimer un rapport (soft delete)"""
    report = Report.query.filter_by(id=report_id, deleted_at=None).first_or_404()
    
    # Vérifier que l'utilisateur est le propriétaire
    if report.user_id != current_user.id:
        return jsonify({"success": False, "error": "Vous ne pouvez supprimer que vos propres rapports"}), 403
    
    report.deleted_at = dt.datetime.utcnow()
    
    try:
        db.session.commit()
        return jsonify({"success": True})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/reports/photos/<int:photo_id>")
@login_required
def report_photo(photo_id):
    """Servir une photo de rapport depuis la base de données"""
    photo = ReportPhoto.query.get_or_404(photo_id)
    report = photo.report
    
    # Vérifier que le rapport n'est pas supprimé
    if report.deleted_at:
        abort(404)
    
    # Si la photo est stockée en BLOB dans la base de données
    if photo.photo_data:
        response = make_response(photo.photo_data)
        response.headers['Content-Type'] = photo.content_type or 'image/jpeg'
        return response
    # Fallback : si la photo est encore sur le système de fichiers (migration)
    elif photo.file_path and os.path.exists(photo.file_path):
        return send_from_directory(
            str(REPORT_PHOTOS_FOLDER),
            os.path.basename(photo.file_path),
            as_attachment=False
        )
    else:
        abort(404)


def cleanup_old_reports():
    """Supprime automatiquement les rapports et leurs photos de plus de 7 jours"""
    try:
        week_ago = dt.datetime.utcnow() - dt.timedelta(days=7)
        
        # Trouver tous les rapports de plus de 7 jours (y compris ceux soft-deleted)
        old_reports = Report.query.filter(Report.created_at < week_ago).all()
        
        deleted_count = 0
        for report in old_reports:
            # Supprimer les fichiers photos si ils existent encore (migration)
            for photo in report.photos:
                if photo.file_path and os.path.exists(photo.file_path):
                    try:
                        os.remove(photo.file_path)
                    except Exception:
                        pass
            
            # Supprimer le rapport (cascade supprimera automatiquement les photos)
            db.session.delete(report)
            deleted_count += 1
        
        if deleted_count > 0:
            db.session.commit()
            print(f"Cleanup: {deleted_count} rapports et leurs photos supprimés (plus de 7 jours)")
        
        return deleted_count
    except Exception as exc:
        db.session.rollback()
        print(f"Erreur lors du nettoyage des rapports: {exc}")
        return 0


def cleanup_old_chat_messages():
    """Supprime automatiquement les messages de chat de plus de 7 jours"""
    try:
        week_ago = dt.datetime.utcnow() - dt.timedelta(days=7)
        deleted_count = ChatMessage.query.filter(ChatMessage.created_at < week_ago).count()
        ChatMessage.query.filter(ChatMessage.created_at < week_ago).delete()
        db.session.commit()
        if deleted_count > 0:
            print(f"Cleanup: {deleted_count} messages de chat supprimés (plus de 7 jours)")
        return deleted_count
    except Exception as exc:
        db.session.rollback()
        print(f"Erreur lors du nettoyage des messages: {exc}")
        return 0


def cleanup_old_reports():
    """Supprime automatiquement les rapports et leurs photos de plus de 7 jours"""
    try:
        week_ago = dt.datetime.utcnow() - dt.timedelta(days=7)
        
        # Trouver tous les rapports de plus de 7 jours (y compris ceux soft-deleted)
        old_reports = Report.query.filter(Report.created_at < week_ago).all()
        
        deleted_count = 0
        for report in old_reports:
            # Supprimer les fichiers photos si ils existent encore (migration)
            for photo in report.photos:
                if photo.file_path and os.path.exists(photo.file_path):
                    try:
                        os.remove(photo.file_path)
                    except Exception:
                        pass
            
            # Supprimer le rapport (cascade supprimera automatiquement les photos)
            db.session.delete(report)
            deleted_count += 1
        
        if deleted_count > 0:
            db.session.commit()
            print(f"Cleanup: {deleted_count} rapports et leurs photos supprimés (plus de 7 jours)")
        
        return deleted_count
    except Exception as exc:
        db.session.rollback()
        print(f"Erreur lors du nettoyage des rapports: {exc}")
        return 0


def cleanup_old_chat_messages():
    """Supprime automatiquement les messages de chat de plus de 7 jours"""
    try:
        week_ago = dt.datetime.utcnow() - dt.timedelta(days=7)
        deleted_count = ChatMessage.query.filter(ChatMessage.created_at < week_ago).count()
        ChatMessage.query.filter(ChatMessage.created_at < week_ago).delete()
        db.session.commit()
        if deleted_count > 0:
            print(f"Cleanup: {deleted_count} messages de chat supprimés (plus de 7 jours)")
        return deleted_count
    except Exception as exc:
        db.session.rollback()
        print(f"Erreur lors du nettoyage des messages: {exc}")
        return 0


@app.context_processor
def inject_now():
    return {"now": dt.datetime.utcnow()}


def get_or_create_progress(machine: Machine, report: PreventiveReport):
    progress = MaintenanceProgress.query.filter_by(machine_id=machine.id, report_id=report.id).one_or_none()
    if not progress:
        initial_hours = report.periodicity if machine.hour_counter_enabled else 0.0
        progress = MaintenanceProgress(machine=machine, report=report, hours_since=initial_hours)
        db.session.add(progress)
        db.session.flush()
    return progress


def ensure_all_progress_for_machine(machine: Machine):
    # Récupérer les progress existants (par machine_id ou counter_id)
    existing_by_machine = {
        record.report_id: record
        for record in MaintenanceProgress.query.filter_by(machine_id=machine.id, counter_id=None).all()
    }
    existing_by_counter = {}
    
    # Récupérer la machine racine pour vérifier ses compteurs
    root_machine = machine
    while root_machine.parent:
        root_machine = root_machine.parent
    
    # Récupérer tous les progress avec counter_id pour cette machine
    all_counter_progress = MaintenanceProgress.query.filter_by(machine_id=machine.id).filter(MaintenanceProgress.counter_id.isnot(None)).all()
    for record in all_counter_progress:
        existing_by_counter[record.report_id] = record
    
    reports = PreventiveReport.query.filter_by(machine_id=machine.id).order_by(PreventiveReport.id).all()
    created = False
    for report in reports:
        # Déterminer si le report utilise un compteur spécifique
        if report.counter_id:
            # Utiliser counter_id (compteur de la machine racine)
            if report.id not in existing_by_counter:
                counter = Counter.query.get(report.counter_id)
                initial_hours = report.periodicity if counter else 0.0
                progress = MaintenanceProgress(machine=machine, counter_id=report.counter_id, report=report, hours_since=initial_hours)
                db.session.add(progress)
                existing_by_counter[report.id] = progress
                created = True
        else:
            # Utiliser machine_id (compteur de la machine elle-même)
            if report.id not in existing_by_machine:
                initial_hours = report.periodicity if machine.hour_counter_enabled else 0.0
                progress = MaintenanceProgress(machine=machine, report=report, hours_since=initial_hours)
                db.session.add(progress)
                existing_by_machine[report.id] = progress
                created = True
    if created:
        db.session.flush()
    # Retourner tous les progress (machine et counter)
    all_progress = list(existing_by_machine.values()) + list(existing_by_counter.values())
    return all_progress


@app.route("/maintenance-photo/<int:photo_id>/view")
@login_required
def view_maintenance_photo(photo_id):
    photo = MaintenancePhoto.query.get_or_404(photo_id)
    
    if not os.path.exists(photo.file_path):
        flash("La photo n'existe plus", "danger")
        if photo.maintenance_entry_id:
            return redirect(url_for("maintenance_entry_detail", entry_id=photo.maintenance_entry_id))
        else:
            return redirect(url_for("corrective_maintenance_detail", maintenance_id=photo.corrective_maintenance_id))
    
    return send_from_directory(
        str(MAINTENANCE_PHOTOS_FOLDER),
        photo.filename,
        as_attachment=False,
        download_name=photo.original_filename
    )


@app.route("/maintenance-photo/<int:photo_id>/delete", methods=["POST"])
@login_required
def delete_maintenance_photo(photo_id):
    photo = MaintenancePhoto.query.get_or_404(photo_id)
    
    # Vérifier les permissions : admin ou technicien qui a créé la photo
    if current_user.user_type != "admin" and (current_user.user_type != "technicien" or photo.user_id != current_user.id):
        flash("Accès refusé : vous ne pouvez supprimer que les photos que vous avez uploadées.", "danger")
        if photo.maintenance_entry_id:
            return redirect(url_for("maintenance_entry_detail", entry_id=photo.maintenance_entry_id))
        else:
            return redirect(url_for("corrective_maintenance_detail", maintenance_id=photo.corrective_maintenance_id))
    
    # Supprimer le fichier physique
    try:
        if os.path.exists(photo.file_path):
            os.remove(photo.file_path)
    except Exception as exc:
        flash(f"Erreur lors de la suppression du fichier: {exc}", "warning")
    
    # Déterminer où rediriger
    redirect_entry_id = photo.maintenance_entry_id
    redirect_maintenance_id = photo.corrective_maintenance_id
    
    # Supprimer l'enregistrement en base de données
    db.session.delete(photo)
    try:
        db.session.commit()
        flash("Photo supprimée avec succès", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Erreur lors de la suppression: {exc}", "danger")
    
    if redirect_entry_id:
        return redirect(url_for("maintenance_entry_detail", entry_id=redirect_entry_id))
    else:
        return redirect(url_for("corrective_maintenance_detail", maintenance_id=redirect_maintenance_id))


@app.route("/checklists/manage")
@login_required
def checklists_manage():
    # Récupérer les paramètres de filtrage
    filter_name = request.args.get('filter_name', '').strip().lower()
    filter_date = request.args.get('filter_date', '').strip()
    filter_machine = request.args.get('filter_machine', '').strip().lower()
    filter_user = request.args.get('filter_user', '').strip().lower()
    
    # Récupérer toutes les check lists remplies (ChecklistInstance) avec les relations chargées
    try:
        entries = ChecklistInstance.query.options(
            joinedload(ChecklistInstance.template),
            joinedload(ChecklistInstance.machine),
            joinedload(ChecklistInstance.user),
        ).order_by(ChecklistInstance.created_at.desc()).all()
    except Exception:
        # En cas d'erreur, essayer sans joinedload
        entries = ChecklistInstance.query.order_by(ChecklistInstance.created_at.desc()).all()
    
    # Créer une liste avec nom, machine, date, utilisateur et lien
    checklists = []
    
    for entry in entries:
        try:
            # Vérifier que l'entrée existe
            if not entry:
                continue
            
            # Charger le template si nécessaire
            template = entry.template if hasattr(entry, "template") else None
            if not template and entry.template_id:
                template = ChecklistTemplate.query.get(entry.template_id)
            
            if template:
                # Charger la machine si nécessaire
                machine = entry.machine if hasattr(entry, "machine") else None
                if not machine and entry.machine_id:
                    machine = Machine.query.get(entry.machine_id)
                
                # Charger l'utilisateur si nécessaire
                user = entry.user if hasattr(entry, "user") else None
                if not user and entry.user_id:
                    user = User.query.get(entry.user_id)
                
                checklists.append(
                    {
                        "name": template.name,
                        "date": entry.created_at,
                        "machine": machine,
                        "user": user,
                        "id": entry.id,
                        "url": url_for(
                            "checklist_instance_detail",
                            machine_id=entry.machine_id,
                            template_id=entry.template_id,
                            instance_id=entry.id,
                        ),
                    }
                )
        except Exception as e:
            # Ignorer les entrées avec des erreurs
            import traceback
            print(f"Erreur lors du traitement de l'entrée {entry.id if entry else 'unknown'}: {e}")
            traceback.print_exc()
            continue
    
    # Appliquer les filtres
    filtered_checklists = checklists
    
    if filter_name:
        filtered_checklists = [c for c in filtered_checklists if filter_name in c['name'].lower()]
    
    if filter_date:
        try:
            # Essayer de parser la date (format attendu: YYYY-MM-DD ou DD/MM/YYYY)
            filter_date_obj = None
            if '/' in filter_date:
                # Format DD/MM/YYYY
                parts = filter_date.split('/')
                if len(parts) == 3:
                    filter_date_obj = dt.datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            else:
                # Format YYYY-MM-DD
                filter_date_obj = dt.datetime.strptime(filter_date, '%Y-%m-%d')
            
            if filter_date_obj:
                filtered_checklists = [
                    c for c in filtered_checklists
                    if c['date'].date() == filter_date_obj.date()
                ]
        except (ValueError, AttributeError):
            pass
    
    if filter_machine:
        # Rechercher dans toute l'arborescence de la machine
        filtered_checklists = [
            c for c in filtered_checklists
            if any(
                filter_machine in node.name.lower() or filter_machine in (node.code or '').lower()
                for node in machine_lineage(c['machine'])
            )
        ]
    
    if filter_user:
        filtered_checklists = [
            c for c in filtered_checklists
            if c['user'] and filter_user in (c['user'].username or '').lower()
        ]
    
    # Trier par date décroissante
    filtered_checklists.sort(key=lambda x: x['date'], reverse=True)
    
    return render_template(
        "checklists_manage.html",
        checklists=filtered_checklists,
        filter_name=filter_name,
        filter_date=filter_date,
        filter_machine=filter_machine,
        filter_user=filter_user
    )


@app.route("/api/dashboard")
@login_required
def get_dashboard_data():
    """API pour récupérer les données du tableau de bord"""
    date_start_str = request.args.get('date_start', '')
    date_end_str = request.args.get('date_end', '')
    machine_ids_str = request.args.get('machine_ids', '')
    metrics_str = request.args.get('metrics', '')
    
    # Parser les IDs de machines directement sélectionnées
    selected_machine_ids = []
    if machine_ids_str:
        try:
            selected_machine_ids = [int(id.strip()) for id in machine_ids_str.split(',') if id.strip()]
        except (ValueError, TypeError):
            pass
    
    # Parser les métriques
    metrics = []
    if metrics_str:
        metrics = [m.strip() for m in metrics_str.split(',') if m.strip()]
    
    # Parser les dates
    start_date = None
    end_date = None
    if date_start_str:
        try:
            start_date = dt.datetime.strptime(date_start_str, '%Y-%m-%d')
        except (ValueError, TypeError):
            pass
    if date_end_str:
        try:
            # Ajouter 23h59m59s pour inclure toute la journée
            end_date = dt.datetime.strptime(date_end_str, '%Y-%m-%d') + dt.timedelta(hours=23, minutes=59, seconds=59)
        except (ValueError, TypeError):
            pass
    
    # Récupérer toutes les machines suivies par l'utilisateur si aucune machine spécifiée
    if not selected_machine_ids:
        followed_machines = FollowedMachine.query.filter_by(user_id=current_user.id).all()
        selected_machine_ids = [fm.machine_id for fm in followed_machines]
    
    # Si aucune machine, retourner des données vides
    if not selected_machine_ids:
        return jsonify({
            'success': True,
            'data': [],
            'period': period
        })
    
    # Récupérer les machines directement sélectionnées
    selected_machines = Machine.query.filter(Machine.id.in_(selected_machine_ids)).all()
    
    results = []
    
    # Pour chaque machine directement sélectionnée, calculer les métriques en incluant toutes ses sous-machines
    for machine in selected_machines:
        # Récupérer toutes les sous-machines (descendants)
        all_machines_in_tree = get_all_descendants(machine)
        all_machine_ids_in_tree = [m.id for m in all_machines_in_tree]
        machine_data = {
            'machine_id': machine.id,
            'machine_name': machine.name,
            'machine_code': machine.code,
            'metrics': {}
        }
        
        # 1. Nombre de maintenances préventives (incluant toutes les sous-machines)
        if not metrics or 'maintenances_preventives' in metrics:
            query = MaintenanceEntry.query.filter(MaintenanceEntry.machine_id.in_(all_machine_ids_in_tree))
            if start_date:
                query = query.filter(MaintenanceEntry.created_at >= start_date)
            if end_date:
                query = query.filter(MaintenanceEntry.created_at <= end_date)
            preventive_count = query.count()
            machine_data['metrics']['maintenances_preventives'] = preventive_count
        
        # 2. Nombre de maintenances curatives (incluant toutes les sous-machines)
        if not metrics or 'maintenances_curatives' in metrics:
            query = CorrectiveMaintenance.query.filter(CorrectiveMaintenance.machine_id.in_(all_machine_ids_in_tree))
            if start_date:
                query = query.filter(CorrectiveMaintenance.created_at >= start_date)
            if end_date:
                query = query.filter(CorrectiveMaintenance.created_at <= end_date)
            corrective_count = query.count()
            machine_data['metrics']['maintenances_curatives'] = corrective_count
        
        # 3. Coût des produits utilisés (incluant toutes les sous-machines)
        if not metrics or 'cout_produits' in metrics:
            total_cost = 0.0
            
            # Coût pour les maintenances préventives (via mouvements)
            query = MaintenanceEntry.query.filter(MaintenanceEntry.machine_id.in_(all_machine_ids_in_tree))
            if start_date:
                query = query.filter(MaintenanceEntry.created_at >= start_date)
            if end_date:
                query = query.filter(MaintenanceEntry.created_at <= end_date)
            preventive_entries = query.all()
            
            for entry in preventive_entries:
                if entry.stock_id:
                    # Chercher les mouvements de sortie dans une fenêtre de 5 minutes
                    time_window_start = entry.created_at - dt.timedelta(minutes=5)
                    time_window_end = entry.created_at + dt.timedelta(minutes=5)
                    
                    movements = Movement.query.filter(
                        Movement.type == 'sortie',
                        Movement.source_stock_id == entry.stock_id,
                        Movement.created_at >= time_window_start,
                        Movement.created_at <= time_window_end
                    ).all()
                    
                    for movement in movements:
                        for item in movement.items:
                            if item.product:
                                total_cost += item.quantity * (item.product.price or 0.0)
            
            # Coût pour les maintenances curatives (via CorrectiveMaintenanceProduct)
            query = CorrectiveMaintenance.query.filter(CorrectiveMaintenance.machine_id.in_(all_machine_ids_in_tree))
            if start_date:
                query = query.filter(CorrectiveMaintenance.created_at >= start_date)
            if end_date:
                query = query.filter(CorrectiveMaintenance.created_at <= end_date)
            corrective_maintenances = query.all()
            
            for cm in corrective_maintenances:
                for product_rel in cm.products:
                    if product_rel.product:
                        total_cost += product_rel.quantity * (product_rel.product.price or 0.0)
            
            machine_data['metrics']['cout_produits'] = round(total_cost, 2)
        
        # 4. Nombre de checklists (incluant toutes les sous-machines)
        if not metrics or 'checklists' in metrics:
            query = ChecklistInstance.query.filter(ChecklistInstance.machine_id.in_(all_machine_ids_in_tree))
            if start_date:
                query = query.filter(ChecklistInstance.created_at >= start_date)
            if end_date:
                query = query.filter(ChecklistInstance.created_at <= end_date)
            checklist_count = query.count()
            machine_data['metrics']['checklists'] = checklist_count
        
        # 5. Maintenances préventives en retard vs à l'heure (incluant toutes les sous-machines)
        if not metrics or 'maintenances_retard' in metrics or 'maintenances_a_heure' in metrics:
            query = MaintenanceEntry.query.filter(MaintenanceEntry.machine_id.in_(all_machine_ids_in_tree))
            if start_date:
                query = query.filter(MaintenanceEntry.created_at >= start_date)
            if end_date:
                query = query.filter(MaintenanceEntry.created_at <= end_date)
            preventive_entries = query.all()
            
            on_time_count = 0
            late_count = 0
            
            for entry in preventive_entries:
                # Si hours_before_maintenance est négatif ou None, c'est en retard
                if entry.hours_before_maintenance is not None:
                    if entry.hours_before_maintenance < 0:
                        late_count += 1
                    else:
                        on_time_count += 1
                else:
                    # Si pas d'info, considérer comme à l'heure
                    on_time_count += 1
            
            if 'maintenances_retard' in (metrics or []):
                machine_data['metrics']['maintenances_retard'] = late_count
            if 'maintenances_a_heure' in (metrics or []):
                machine_data['metrics']['maintenances_a_heure'] = on_time_count
        
        # 6. Nombre de mises à jour de compteur (incluant toutes les sous-machines)
        if not metrics or 'mises_a_jour_compteur' in metrics:
            query = CounterLog.query.filter(CounterLog.machine_id.in_(all_machine_ids_in_tree))
            if start_date:
                query = query.filter(CounterLog.created_at >= start_date)
            if end_date:
                query = query.filter(CounterLog.created_at <= end_date)
            counter_count = query.count()
            machine_data['metrics']['mises_a_jour_compteur'] = counter_count
        
        results.append(machine_data)
    
    return jsonify({
        'success': True,
        'data': results
    })


@app.route("/api/dashboard-chart")
@login_required
def get_dashboard_chart_data():
    """API pour récupérer les données du tableau de bord groupées par période temporelle"""
    date_start_str = request.args.get('date_start', '')
    date_end_str = request.args.get('date_end', '')
    machine_ids_str = request.args.get('machine_ids', '')
    metrics_str = request.args.get('metrics', '')
    
    # Parser les IDs de machines directement sélectionnées
    selected_machine_ids = []
    if machine_ids_str:
        try:
            selected_machine_ids = [int(id.strip()) for id in machine_ids_str.split(',') if id.strip()]
        except (ValueError, TypeError):
            pass
    
    # Parser les métriques
    metrics = []
    if metrics_str:
        metrics = [m.strip() for m in metrics_str.split(',') if m.strip()]
    
    # Parser les dates
    start_date = None
    end_date = None
    if date_start_str:
        try:
            start_date = dt.datetime.strptime(date_start_str, '%Y-%m-%d')
        except (ValueError, TypeError):
            pass
    if date_end_str:
        try:
            # Ajouter 23h59m59s pour inclure toute la journée
            end_date = dt.datetime.strptime(date_end_str, '%Y-%m-%d') + dt.timedelta(hours=23, minutes=59, seconds=59)
        except (ValueError, TypeError):
            pass
    
    # Déterminer le groupement temporel selon la période
    if start_date and end_date:
        # S'assurer que end_date est après start_date
        if end_date < start_date:
            end_date, start_date = start_date, end_date
        delta = end_date - start_date
        if delta.days <= 31:
            time_group = 'day'  # Grouper par jour pour les périodes courtes
        elif delta.days <= 365:
            time_group = 'week'  # Grouper par semaine pour les périodes moyennes
        else:
            time_group = 'month'  # Grouper par mois pour les périodes longues
    elif start_date:
        # Si seulement date de début, utiliser le groupement par défaut
        time_group = 'month'
    else:
        # Si aucune date, utiliser le groupement par mois
        time_group = 'month'
    
    # Récupérer toutes les machines suivies par l'utilisateur si aucune machine spécifiée
    if not selected_machine_ids:
        followed_machines = FollowedMachine.query.filter_by(user_id=current_user.id).all()
        selected_machine_ids = [fm.machine_id for fm in followed_machines]
    
    # Si aucune machine, retourner des données vides
    if not selected_machine_ids:
        return jsonify({
            'success': True,
            'data': [],
            'time_group': time_group
        })
    
    # Récupérer les machines directement sélectionnées et leurs descendants
    selected_machines = Machine.query.filter(Machine.id.in_(selected_machine_ids)).all()
    all_machine_ids_in_trees = set()
    for machine in selected_machines:
        descendants = get_all_descendants(machine)
        all_machine_ids_in_trees.update([m.id for m in descendants])
        # Ajouter aussi la machine elle-même
        all_machine_ids_in_trees.add(machine.id)
    all_machine_ids_in_trees = list(all_machine_ids_in_trees)
    
    # Si aucune machine trouvée, retourner vide
    if not all_machine_ids_in_trees:
        return jsonify({
            'success': True,
            'data': [],
            'time_group': time_group
        })
    
    # Générer les périodes temporelles
    periods = []
    if start_date:
        current = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_limit = end_date if end_date else dt.datetime.utcnow()
        while current <= end_limit:
            if time_group == 'day':
                periods.append(current.date())
                current += dt.timedelta(days=1)
            elif time_group == 'week':
                # Premier jour de la semaine (lundi)
                days_since_monday = current.weekday()
                week_start = current - dt.timedelta(days=days_since_monday)
                week_start_date = week_start.date()
                if not periods or periods[-1] != week_start_date:
                    periods.append(week_start_date)
                current = week_start + dt.timedelta(days=7)
            elif time_group == 'month':
                # Premier jour du mois
                periods.append(dt.date(current.year, current.month, 1))
                # Passer au mois suivant
                if current.month == 12:
                    current = dt.datetime(current.year + 1, 1, 1)
                else:
                    current = dt.datetime(current.year, current.month + 1, 1)
    else:
        # Depuis le début : trouver la première date dans les données
        first_dates = []
        # Toujours vérifier toutes les sources de données pour trouver la première date
        first_entry = MaintenanceEntry.query.filter(
            MaintenanceEntry.machine_id.in_(all_machine_ids_in_trees)
        ).order_by(MaintenanceEntry.created_at.asc()).first()
        if first_entry:
            first_dates.append(first_entry.created_at.date())
        
        first_corrective = CorrectiveMaintenance.query.filter(
            CorrectiveMaintenance.machine_id.in_(all_machine_ids_in_trees)
        ).order_by(CorrectiveMaintenance.created_at.asc()).first()
        if first_corrective:
            first_dates.append(first_corrective.created_at.date())
        
        first_checklist = ChecklistInstance.query.filter(
            ChecklistInstance.machine_id.in_(all_machine_ids_in_trees)
        ).order_by(ChecklistInstance.created_at.asc()).first()
        if first_checklist:
            first_dates.append(first_checklist.created_at.date())
        
        first_counter = CounterLog.query.filter(
            CounterLog.machine_id.in_(all_machine_ids_in_trees)
        ).order_by(CounterLog.created_at.asc()).first()
        if first_counter:
            first_dates.append(first_counter.created_at.date())
        
        if first_dates:
            first_date = min(first_dates)
            current = dt.datetime.combine(first_date, dt.time.min)
            # Commencer au premier jour du mois de la première date
            current = dt.datetime(current.year, current.month, 1)
            now = dt.datetime.utcnow()
            while current <= now:
                periods.append(dt.date(current.year, current.month, 1))
                if current.month == 12:
                    current = dt.datetime(current.year + 1, 1, 1)
                else:
                    current = dt.datetime(current.year, current.month + 1, 1)
        else:
            # Pas de données, retourner vide
            return jsonify({
                'success': True,
                'data': [],
                'time_group': time_group
            })
    
    results = []
    
    for period_date in periods:
        period_start = dt.datetime.combine(period_date, dt.time.min)
        if time_group == 'day':
            period_end = period_start + dt.timedelta(days=1)
            period_label = period_date.strftime('%d/%m/%Y')
        elif time_group == 'week':
            period_end = period_start + dt.timedelta(days=7)
            period_label = period_date.strftime('%d/%m/%Y')
        else:  # month
            if period_date.month == 12:
                period_end = dt.datetime(period_date.year + 1, 1, 1)
            else:
                period_end = dt.datetime(period_date.year, period_date.month + 1, 1)
            period_label = period_date.strftime('%m/%Y')
        
        period_data = {
            'period': period_date.isoformat(),
            'period_label': period_label,
            'metrics': {}
        }
        
        # 1. Nombre de maintenances préventives
        if not metrics or 'maintenances_preventives' in metrics:
            query = MaintenanceEntry.query.filter(
                MaintenanceEntry.machine_id.in_(all_machine_ids_in_trees),
                MaintenanceEntry.created_at >= period_start,
                MaintenanceEntry.created_at < period_end
            )
            preventive_count = query.count()
            period_data['metrics']['maintenances_preventives'] = preventive_count
        
        # 2. Nombre de maintenances curatives
        if not metrics or 'maintenances_curatives' in metrics:
            query = CorrectiveMaintenance.query.filter(
                CorrectiveMaintenance.machine_id.in_(all_machine_ids_in_trees),
                CorrectiveMaintenance.created_at >= period_start,
                CorrectiveMaintenance.created_at < period_end
            )
            corrective_count = query.count()
            period_data['metrics']['maintenances_curatives'] = corrective_count
        
        # 3. Coût des produits utilisés
        if not metrics or 'cout_produits' in metrics:
            total_cost = 0.0
            
            # Coût pour les maintenances préventives (via mouvements)
            preventive_entries = MaintenanceEntry.query.filter(
                MaintenanceEntry.machine_id.in_(all_machine_ids_in_trees),
                MaintenanceEntry.created_at >= period_start,
                MaintenanceEntry.created_at < period_end
            ).all()
            
            for entry in preventive_entries:
                if entry.stock_id:
                    time_window_start = entry.created_at - dt.timedelta(minutes=5)
                    time_window_end = entry.created_at + dt.timedelta(minutes=5)
                    
                    movements = Movement.query.filter(
                        Movement.type == 'sortie',
                        Movement.source_stock_id == entry.stock_id,
                        Movement.created_at >= time_window_start,
                        Movement.created_at <= time_window_end
                    ).all()
                    
                    for movement in movements:
                        for item in movement.items:
                            if item.product:
                                total_cost += item.quantity * (item.product.price or 0.0)
            
            # Coût pour les maintenances curatives
            corrective_maintenances = CorrectiveMaintenance.query.filter(
                CorrectiveMaintenance.machine_id.in_(all_machine_ids_in_trees),
                CorrectiveMaintenance.created_at >= period_start,
                CorrectiveMaintenance.created_at < period_end
            ).all()
            
            for cm in corrective_maintenances:
                for product_rel in cm.products:
                    if product_rel.product:
                        total_cost += product_rel.quantity * (product_rel.product.price or 0.0)
            
            period_data['metrics']['cout_produits'] = round(total_cost, 2)
        
        # 4. Nombre de checklists
        if not metrics or 'checklists' in metrics:
            query = ChecklistInstance.query.filter(
                ChecklistInstance.machine_id.in_(all_machine_ids_in_trees),
                ChecklistInstance.created_at >= period_start,
                ChecklistInstance.created_at < period_end
            )
            checklist_count = query.count()
            period_data['metrics']['checklists'] = checklist_count
        
        # 5. Maintenances préventives en retard vs à l'heure
        if not metrics or 'maintenances_retard' in metrics or 'maintenances_a_heure' in metrics:
            preventive_entries = MaintenanceEntry.query.filter(
                MaintenanceEntry.machine_id.in_(all_machine_ids_in_trees),
                MaintenanceEntry.created_at >= period_start,
                MaintenanceEntry.created_at < period_end
            ).all()
            
            on_time_count = 0
            late_count = 0
            
            for entry in preventive_entries:
                if entry.hours_before_maintenance is not None:
                    if entry.hours_before_maintenance < 0:
                        late_count += 1
                    else:
                        on_time_count += 1
                else:
                    on_time_count += 1
            
            if 'maintenances_retard' in (metrics or []):
                period_data['metrics']['maintenances_retard'] = late_count
            if 'maintenances_a_heure' in (metrics or []):
                period_data['metrics']['maintenances_a_heure'] = on_time_count
        
        # 6. Nombre de mises à jour de compteur
        if not metrics or 'mises_a_jour_compteur' in metrics:
            query = CounterLog.query.filter(
                CounterLog.machine_id.in_(all_machine_ids_in_trees),
                CounterLog.created_at >= period_start,
                CounterLog.created_at < period_end
            )
            counter_count = query.count()
            period_data['metrics']['mises_a_jour_compteur'] = counter_count
        
        results.append(period_data)
    
    return jsonify({
        'success': True,
        'data': results,
        'time_group': time_group
    })


@app.route("/chat")
@login_required
def chat_full_page():
    """Page complète dédiée au chat et informations"""
    return render_template("chat_full.html")


# Routes pour la gestion des fichiers Excel
@app.route("/database-export/excel/upload", methods=["POST"])
@admin_required
def upload_excel_file():
    """Uploader un fichier Excel"""
    if 'file' not in request.files:
        flash("Aucun fichier sélectionné", "danger")
        return redirect(url_for("database_export"))
    
    file = request.files['file']
    name = request.form.get('name', '').strip()
    
    if file.filename == '':
        flash("Aucun fichier sélectionné", "danger")
        return redirect(url_for("database_export"))
    
    if not name:
        flash("Veuillez donner un nom au fichier", "danger")
        return redirect(url_for("database_export"))
    
    # Vérifier l'extension
    filename = file.filename
    if '.' not in filename:
        flash("Fichier invalide : extension manquante", "danger")
        return redirect(url_for("database_export"))
    
    ext = filename.rsplit('.', 1)[1].lower()
    if ext not in ALLOWED_EXCEL_EXTENSIONS:
        flash(f"Format de fichier non autorisé. Formats acceptés : {', '.join(ALLOWED_EXCEL_EXTENSIONS)}", "danger")
        return redirect(url_for("database_export"))
    
    try:
        # Générer un nom de fichier unique
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_filename = secure_filename(filename)
        unique_filename = f"{timestamp}_{safe_filename}"
        file_path = EXCEL_FILES_FOLDER / unique_filename
        
        # Sauvegarder le fichier
        file.save(str(file_path))
        
        # Créer l'entrée en base de données
        excel_file = ExcelFile(
            name=name,
            filename=unique_filename,
            original_filename=filename,
            user_id=current_user.id
        )
        db.session.add(excel_file)
        db.session.commit()
        
        flash(f"Fichier '{name}' uploadé avec succès", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de l'upload : {str(e)}", "danger")
    
    return redirect(url_for("database_export"))


@app.route("/database-export/excel/<int:file_id>/download")
@admin_required
def download_excel_file(file_id):
    """Télécharger un fichier Excel"""
    excel_file = ExcelFile.query.get_or_404(file_id)
    file_path = EXCEL_FILES_FOLDER / excel_file.filename
    
    if not file_path.exists():
        flash("Fichier introuvable", "danger")
        return redirect(url_for("database_export"))
    
    return send_from_directory(
        str(EXCEL_FILES_FOLDER),
        excel_file.filename,
        as_attachment=True,
        download_name=excel_file.original_filename
    )


@app.route("/database-export/excel/<int:file_id>/delete", methods=["POST"])
@admin_required
def delete_excel_file(file_id):
    """Supprimer un fichier Excel"""
    excel_file = ExcelFile.query.get_or_404(file_id)
    file_path = EXCEL_FILES_FOLDER / excel_file.filename
    
    try:
        # Supprimer le fichier du disque
        if file_path.exists():
            file_path.unlink()
        
        # Supprimer l'entrée en base de données
        db.session.delete(excel_file)
        db.session.commit()
        
        flash(f"Fichier '{excel_file.name}' supprimé avec succès", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de la suppression : {str(e)}", "danger")
    
    return redirect(url_for("database_export"))


def run_cleanup_scheduler():
    """Lance le scheduler de nettoyage automatique en arrière-plan"""
    def cleanup_loop():
        while True:
            try:
                # Exécuter le nettoyage toutes les heures
                cleanup_old_reports()
                cleanup_old_chat_messages()
            except Exception as exc:
                print(f"Erreur dans le scheduler de nettoyage: {exc}")
            # Attendre 1 heure avant le prochain nettoyage
            time.sleep(3600)
    
    # Démarrer le thread de nettoyage
    cleanup_thread = threading.Thread(target=cleanup_loop, daemon=True)
    cleanup_thread.start()
    print("Scheduler de nettoyage automatique démarré (nettoyage toutes les heures)")


if __name__ == "__main__":
    # Démarrer le scheduler de nettoyage
    run_cleanup_scheduler()
    app.run(debug=True)
else:
    # Pour Gunicorn et autres serveurs WSGI
    run_cleanup_scheduler()

